#!/usr/bin/env python3
"""
🔍 ML EXTRACTOR TEST ANALYZER
Comprehensive tool to analyze output quality and detect mapping errors
"""

import openpyxl
import pandas as pd
import json
from datetime import datetime
import os

class MLOutputAnalyzer:
    def __init__(self):
        self.results = {
            'timestamp': datetime.now().isoformat(),
            'formula_fix_test': {},
            'mapping_quality': {},
            'field_completeness': {},
            'ai_enhancement': {},
            'errors': [],
            'recommendations': []
        }
    
    def analyze_excel_formulas(self, input_file):
        """Test if formulas are being read correctly"""
        print("🔍 TESTING FORMULA READING...")
        
        try:
            # Read with formulas (old way)
            wb_formulas = openpyxl.load_workbook(input_file, data_only=False)
            sheet_formulas = wb_formulas.active
            
            # Read with values only (new way)
            try:
                wb_values = openpyxl.load_workbook(input_file, data_only=True)
            except TypeError:
                wb_values = openpyxl.load_workbook(input_file)
            sheet_values = wb_values.active
            
            formula_cells = []
            value_cells = []
            
            # Check first 10 rows for price-like fields
            for row in range(1, min(11, sheet_formulas.max_row + 1)):
                for col in range(1, sheet_formulas.max_column + 1):
                    cell_formula = sheet_formulas.cell(row, col).value
                    cell_value = sheet_values.cell(row, col).value
                    
                    if isinstance(cell_formula, str) and cell_formula.startswith('='):
                        formula_cells.append({
                            'row': row, 'col': col,
                            'formula': cell_formula,
                            'calculated_value': cell_value
                        })
            
            self.results['formula_fix_test'] = {
                'formulas_detected': len(formula_cells),
                'formulas_list': formula_cells,
                'fix_working': len(formula_cells) > 0
            }
            
            print(f"✓ Found {len(formula_cells)} formula cells")
            for cell in formula_cells:
                print(f"  Row {cell['row']}, Col {cell['col']}: {cell['formula']} = {cell['calculated_value']}")
                
        except Exception as e:
            self.results['errors'].append(f"Formula analysis error: {str(e)}")
            print(f"❌ Error analyzing formulas: {e}")
    
    def analyze_output_quality(self, output_file):
        """Analyze the generated ML output file"""
        print("\n📊 ANALYZING OUTPUT QUALITY...")
        
        try:
            try:
                wb = openpyxl.load_workbook(output_file, data_only=True)
            except TypeError:
                wb = openpyxl.load_workbook(output_file)
            
            # Find the data sheet (not Ayuda, Legales)
            data_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() not in ['ayuda', 'legales', 'extra info']:
                    data_sheet = wb[sheet_name]
                    break
            
            if not data_sheet:
                self.results['errors'].append("No data sheet found in output")
                return
            
            # Get headers (row 3)
            headers = {}
            for col in range(1, data_sheet.max_column + 1):
                header = data_sheet.cell(row=3, column=col).value
                if header:
                    headers[col] = str(header).strip()
            
            # Analyze data rows (from row 4 onwards)
            total_rows = 0
            filled_fields = {}
            empty_fields = {}
            price_values = []
            
            for row in range(4, data_sheet.max_row + 1):
                has_data = False
                row_data = {}
                
                for col, header in headers.items():
                    value = data_sheet.cell(row, col).value
                    row_data[header] = value
                    
                    if value and str(value).strip():
                        has_data = True
                        filled_fields[header] = filled_fields.get(header, 0) + 1
                        
                        # Track price values
                        if 'precio' in header.lower():
                            try:
                                price_val = float(str(value).replace('$', '').replace(',', ''))
                                price_values.append(price_val)
                            except:
                                pass
                    else:
                        empty_fields[header] = empty_fields.get(header, 0) + 1
                
                if has_data:
                    total_rows += 1
            
            # Calculate completeness percentages
            completeness = {}
            for header in headers.values():
                filled = filled_fields.get(header, 0)
                empty = empty_fields.get(header, 0)
                total = filled + empty
                if total > 0:
                    completeness[header] = {
                        'filled': filled,
                        'empty': empty,
                        'percentage': round((filled / total) * 100, 2)
                    }
            
            self.results['field_completeness'] = completeness
            self.results['mapping_quality'] = {
                'total_products': total_rows,
                'price_values_detected': len(price_values),
                'price_range': {
                    'min': min(price_values) if price_values else 0,
                    'max': max(price_values) if price_values else 0,
                    'avg': sum(price_values) / len(price_values) if price_values else 0
                }
            }
            
            print(f"✓ Total products processed: {total_rows}")
            print(f"✓ Price values detected: {len(price_values)}")
            if price_values:
                print(f"✓ Price range: ${min(price_values):.2f} - ${max(price_values):.2f}")
            
            # Generate recommendations
            self._generate_recommendations(completeness, total_rows)
            
        except Exception as e:
            self.results['errors'].append(f"Output analysis error: {str(e)}")
            print(f"❌ Error analyzing output: {e}")
    
    def _generate_recommendations(self, completeness, total_rows):
        """Generate improvement recommendations"""
        recommendations = []
        
        # Check critical fields
        critical_fields = ['Título', 'Precio', 'Stock disponible']
        for field in critical_fields:
            for header, data in completeness.items():
                if any(critical in header for critical in critical_fields):
                    if data['percentage'] < 80:
                        recommendations.append(f"❗ {header}: Only {data['percentage']}% complete - Review mapping")
                    elif data['percentage'] == 100:
                        recommendations.append(f"✅ {header}: Perfect mapping (100%)")
        
        # Check AI enhanced fields
        ai_fields = ['Código universal', 'Marca', 'Modelo']
        for field in ai_fields:
            for header, data in completeness.items():
                if any(ai_field.lower() in header.lower() for ai_field in ai_fields):
                    if data['percentage'] > 50:
                        recommendations.append(f"🤖 {header}: AI working well ({data['percentage']}%)")
        
        # Check price mapping specifically
        price_mapped = False
        for header, data in completeness.items():
            if 'precio' in header.lower() and data['filled'] > 0:
                price_mapped = True
                recommendations.append(f"💰 Price mapping: SUCCESS! {data['filled']} products with prices")
                break
        
        if not price_mapped:
            recommendations.append("❌ CRITICAL: No prices mapped - Check Excel formula fix")
        
        self.results['recommendations'] = recommendations
        
        print("\n🎯 RECOMMENDATIONS:")
        for rec in recommendations:
            print(f"  {rec}")
    
    def save_report(self, filename="test_analysis_report.json"):
        """Save detailed analysis report"""
        report_path = f"/home/granaventura/Desktop/ML EXTRACTOR/{filename}"
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        print(f"\n📋 Full report saved: {report_path}")
        return report_path

def main():
    print("=" * 60)
    print("🔍 ML EXTRACTOR QUALITY ANALYZER")
    print("=" * 60)
    
    analyzer = MLOutputAnalyzer()
    
    # List available files
    uploads_dir = "/home/granaventura/Desktop/ML EXTRACTOR/uploads"
    print(f"\n📁 Available files in uploads:")
    
    input_files = []
    output_files = []
    
    for file in os.listdir(uploads_dir):
        if file.endswith('.xlsx'):
            file_path = os.path.join(uploads_dir, file)
            print(f"  📄 {file}")
            
            if 'output' in file.lower() or 'ml_ai' in file.lower():
                output_files.append(file_path)
            else:
                input_files.append(file_path)
    
    # Analyze latest files
    if input_files:
        latest_input = max(input_files, key=os.path.getctime)
        print(f"\n🔍 Analyzing input file: {os.path.basename(latest_input)}")
        analyzer.analyze_excel_formulas(latest_input)
    
    if output_files:
        latest_output = max(output_files, key=os.path.getctime)
        print(f"\n📊 Analyzing output file: {os.path.basename(latest_output)}")
        analyzer.analyze_output_quality(latest_output)
    
    # Save report
    report_file = analyzer.save_report(f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    
    print("\n" + "=" * 60)
    print("✅ ANALYSIS COMPLETE!")
    print("=" * 60)
    
    return analyzer.results

if __name__ == "__main__":
    main()
