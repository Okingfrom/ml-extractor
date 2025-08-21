#!/usr/bin/env python3
"""
Simple command-line version of the ML Extractor that doesn't require NumPy/Pandas/Streamlit
"""

import argparse
import sys
import os
import openpyxl
import csv
import docx
import PyPDF2
import yaml

def read_excel_simple(file_path):
    """Read Excel file without pandas"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = {}
    headers = []
    
    # Get headers from first row
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Get first data row
    if sheet.max_row > 1:
        for i, cell in enumerate(sheet[2]):
            if i < len(headers) and headers[i]:
                data[headers[i]] = cell.value
    
    return data

def read_csv_simple(file_path):
    """Read CSV file without pandas"""
    data = {}
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        # Get first row
        try:
            data = next(reader)
        except StopIteration:
            pass
    return data

def read_docx_simple(file_path):
    """Read DOCX file"""
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def read_pdf_simple(file_path):
    """Read PDF file"""
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text

def read_txt_simple(file_path):
    """Read text file"""
    with open(file_path, "r", encoding='utf-8') as f:
        return f.read()

def load_mapping_simple(config_path):
    """Load mapping configuration"""
    try:
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
            return config.get('template_columns', []), config.get('mapping', {})
    except Exception as e:
        print(f"Error loading config: {e}")
        return [], {}

def apply_mapping_simple(data, mapping, template_columns):
    """Apply mapping without pandas"""
    mapped_data = {}
    for template_col in template_columns:
        if template_col in mapping:
            source_field = mapping[template_col]
            if isinstance(data, dict) and source_field in data:
                mapped_data[template_col] = data[source_field]
            else:
                mapped_data[template_col] = ""
        else:
            mapped_data[template_col] = ""
    return mapped_data

def write_excel_simple(data, template_columns, output_path):
    """Write Excel file without pandas"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Write headers
    for i, col in enumerate(template_columns, 1):
        sheet.cell(row=1, column=i, value=col)
    
    # Write data
    for i, col in enumerate(template_columns, 1):
        sheet.cell(row=2, column=i, value=data.get(col, ""))
    
    workbook.save(output_path)

def main():
    parser = argparse.ArgumentParser(description='ML Extractor - Simple version without NumPy/Pandas')
    parser.add_argument('input_file', help='Input file (Excel, CSV, PDF, DOCX, TXT)')
    parser.add_argument('output_file', help='Output Excel file')
    parser.add_argument('--config', '-c', default='config/mapping.yaml', 
                       help='Mapping configuration file (default: config/mapping.yaml)')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' not found")
        sys.exit(1)
    
    # Check if config file exists
    if not os.path.exists(args.config):
        print(f"Error: Config file '{args.config}' not found")
        sys.exit(1)
    
    print(f"Processing: {args.input_file}")
    print(f"Output: {args.output_file}")
    print(f"Config: {args.config}")
    
    # Load mapping configuration
    template_columns, mapping = load_mapping_simple(args.config)
    
    if not mapping:
        print("Error: Could not load mapping configuration")
        sys.exit(1)
    
    print(f"Found {len(template_columns)} template columns and {len(mapping)} mappings")
    
    # Process input file based on extension
    ext = args.input_file.split('.')[-1].lower()
    
    try:
        if ext in ['xlsx', 'xls']:
            print("Reading Excel file...")
            data = read_excel_simple(args.input_file)
        elif ext == 'csv':
            print("Reading CSV file...")
            data = read_csv_simple(args.input_file)
        elif ext == 'txt':
            print("Reading text file...")
            text_content = read_txt_simple(args.input_file)
            data = {'content': text_content}  # Simple mapping for text
        elif ext == 'docx':
            print("Reading DOCX file...")
            doc_content = read_docx_simple(args.input_file)
            data = {'content': doc_content}  # Simple mapping for docx
        elif ext == 'pdf':
            print("Reading PDF file...")
            pdf_content = read_pdf_simple(args.input_file)
            data = {'content': pdf_content}  # Simple mapping for pdf
        else:
            print(f"Error: Unsupported file type: {ext}")
            sys.exit(1)
            
        print(f"Extracted {len(data)} fields from input file")
        
        # Apply mapping
        print("Applying mapping...")
        mapped_data = apply_mapping_simple(data, mapping, template_columns)
        
        # Write output
        print(f"Writing output to {args.output_file}...")
        write_excel_simple(mapped_data, template_columns, args.output_file)
        
        print("Success! Processing completed.")
        print(f"Mapped {len([v for v in mapped_data.values() if v])} fields")
        
    except Exception as e:
        print(f"Error processing files: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
