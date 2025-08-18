#!/usr/bin/env python3
"""
Simple Flask-based alternative to Streamlit that doesn't require NumPy/Pandas
"""

from flask import Flask, request, render_template_string, send_file, redirect, url_for
import csv
import os
import tempfile
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Mercado Libre Bulk Mapper</title>
    <style>
        body { font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }
        .form-group { margin-bottom: 15px; }
        label { display: block; margin-bottom: 5px; font-weight: bold; }
        input[type="file"], input[type="text"] { width: 100%; padding: 8px; }
        button { background-color: #4CAF50; color: white; padding: 10px 20px; border: none; cursor: pointer; }
        button:hover { background-color: #45a049; }
        .success { color: green; margin-top: 10px; }
        .error { color: red; margin-top: 10px; }
    </style>
</head>
<body>
    <h1>Mercado Libre Bulk Mapper</h1>
    
    {% if message %}
        <div class="{{ message_type }}">{{ message }}</div>
    {% endif %}
    
    <form method="post" enctype="multipart/form-data">
        <div class="form-group">
            <label for="template">Upload Mercado Libre Template (.xlsx):</label>
            <input type="file" name="template" accept=".xlsx" required>
        </div>
        
        <div class="form-group">
            <label for="content">Upload Product Data:</label>
            <input type="file" name="content" accept=".xlsx,.xls,.csv,.pdf,.docx,.txt" required>
        </div>
        
        <div class="form-group">
            <label for="config_path">Mapping Config Path:</label>
            <input type="text" name="config_path" value="config/mapping.yaml" required>
        </div>
        
        <button type="submit">Process Files</button>
    </form>
    
    {% if output_file %}
        <p><a href="{{ url_for('download_file', filename=output_file) }}">Download Processed File</a></p>
    {% endif %}
</body>
</html>
'''

def read_excel_simple(file_path):
    """Read Excel file without pandas"""
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is required to read Excel files. Install it or run using the fallback environment.") from e

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = {}
    headers = []
    
    # Get headers from first row
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Get first data row
    if sheet.max_row > 1:
        for i, cell in enumerate(sheet[2]):
            if i < len(headers) and headers[i]:
                data[headers[i]] = cell.value
    
    return data

def read_csv_simple(file_path):
    """Read CSV file without pandas"""
    data = {}
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        # Get first row
        try:
            data = next(reader)
        except StopIteration:
            pass
    return data

def read_docx_simple(file_path):
    """Read DOCX file"""
    try:
        import docx
    except Exception as e:
        raise RuntimeError("python-docx is required to read .docx files. Install it or use a different input format.") from e

    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def read_pdf_simple(file_path):
    """Read PDF file"""
    try:
        import PyPDF2
    except Exception as e:
        raise RuntimeError("PyPDF2 is required to read PDF files. Install it or use a different input format.") from e

    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text

def read_txt_simple(file_path):
    """Read text file"""
    with open(file_path, "r", encoding='utf-8') as f:
        return f.read()

def load_mapping_simple(config_path):
    """Load mapping configuration"""
    try:
        try:
            import yaml
        except Exception:
            # If yaml is not available, return empty mapping to avoid import-time failures
            return [], {}

        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
            return config.get('template_columns', []), config.get('mapping', {})
    except Exception:
        return [], {}

def apply_mapping_simple(data, mapping, template_columns):
    """Apply mapping without pandas"""
    mapped_data = {}
    for template_col in template_columns:
        if template_col in mapping:
            source_field = mapping[template_col]
            if isinstance(data, dict) and source_field in data:
                mapped_data[template_col] = data[source_field]
            else:
                mapped_data[template_col] = ""
        else:
            mapped_data[template_col] = ""
    return mapped_data

def write_excel_simple(data, template_columns, output_path):
    """Write Excel file without pandas"""
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is required to write Excel files. Install it or run using the fallback environment.") from e

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Write headers
    for i, col in enumerate(template_columns, 1):
        sheet.cell(row=1, column=i, value=col)
    
    # Write data
    for i, col in enumerate(template_columns, 1):
        sheet.cell(row=2, column=i, value=data.get(col, ""))
    
    workbook.save(output_path)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            # Get uploaded files
            template_file = request.files.get('template')
            content_file = request.files.get('content')
            config_path = request.form.get('config_path', 'config/mapping.yaml')
            
            if not template_file or not content_file:
                return render_template_string(HTML_TEMPLATE, 
                                           message="Please select both files", 
                                           message_type="error")
            
            # Save uploaded files
            template_filename = secure_filename(template_file.filename)
            content_filename = secure_filename(content_file.filename)
            
            template_path = os.path.join(app.config['UPLOAD_FOLDER'], template_filename)
            content_path = os.path.join(app.config['UPLOAD_FOLDER'], content_filename)
            
            template_file.save(template_path)
            content_file.save(content_path)
            
            # Load mapping configuration
            template_columns, mapping = load_mapping_simple(config_path)
            
            if not mapping:
                return render_template_string(HTML_TEMPLATE, 
                                           message=f"Could not load mapping from {config_path}", 
                                           message_type="error")
            
            # Process content file based on extension
            ext = content_filename.split('.')[-1].lower()
            
            if ext in ['xlsx', 'xls']:
                data = read_excel_simple(content_path)
            elif ext == 'csv':
                data = read_csv_simple(content_path)
            elif ext == 'txt':
                text_content = read_txt_simple(content_path)
                data = {'content': text_content}  # Simple mapping for text
            elif ext == 'docx':
                doc_content = read_docx_simple(content_path)
                data = {'content': doc_content}  # Simple mapping for docx
            elif ext == 'pdf':
                pdf_content = read_pdf_simple(content_path)
                data = {'content': pdf_content}  # Simple mapping for pdf
            else:
                return render_template_string(HTML_TEMPLATE, 
                                           message="Unsupported file type", 
                                           message_type="error")
            
            # Apply mapping
            mapped_data = apply_mapping_simple(data, mapping, template_columns)
            
            # Write output
            output_filename = f"output_{template_filename}"
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            write_excel_simple(mapped_data, template_columns, output_path)
            
            return render_template_string(HTML_TEMPLATE, 
                                       message="Files processed successfully!", 
                                       message_type="success",
                                       output_file=output_filename)
            
        except Exception as e:
            return render_template_string(HTML_TEMPLATE, 
                                       message=f"Error processing files: {str(e)}", 
                                       message_type="error")
    
    return render_template_string(HTML_TEMPLATE)

@app.route('/download/<filename>')
def download_file(filename):
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return f"Error downloading file: {str(e)}", 404

if __name__ == '__main__':
    print("Starting Mercado Libre Bulk Mapper on http://localhost:5000")
    app.run(debug=True, host='localhost', port=5000)
