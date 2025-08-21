import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Crear datos de ejemplo para plantilla ML
data = {
    'Título': ['iPhone 14 Pro Max 256GB', 'Samsung Galaxy S23 Ultra', 'MacBook Air M2', 'Dell XPS 13', 'Sony WH-1000XM5'],
    'Precio': [1200000, 1100000, 2500000, 1800000, 450000],
    'Stock': [15, 8, 5, 12, 25],
    'Categoría': ['Celulares', 'Celulares', 'Computadoras', 'Computadoras', 'Auriculares'],
    'Marca': ['Apple', 'Samsung', 'Apple', 'Dell', 'Sony'],
    'Modelo': ['iPhone 14 Pro Max', 'Galaxy S23 Ultra', 'MacBook Air M2', 'XPS 13', 'WH-1000XM5'],
    'Descripción': [
        'Smartphone Apple iPhone 14 Pro Max con 256GB de almacenamiento',
        'Smartphone Samsung Galaxy S23 Ultra con S Pen', 
        'Laptop Apple MacBook Air con chip M2',
        'Laptop Dell XPS 13 ultradelgada',
        'Auriculares inalámbricos con cancelación de ruido'
    ],
    'Peso': [238, 234, 1240, 1200, 250],
    'Garantía': ['12 meses', '12 meses', '12 meses', '24 meses', '24 meses'],
    'Condición': ['Nuevo', 'Nuevo', 'Nuevo', 'Nuevo', 'Nuevo'],
    'Imagen Principal': [
        'https://example.com/iphone.jpg',
        'https://example.com/samsung.jpg', 
        'https://example.com/macbook.jpg',
        'https://example.com/dell.jpg',
        'https://example.com/sony.jpg'
    ]
}

# Crear Excel con estructura ML
output_path = 'C:/Users/equipo/Desktop/ml-extractor/samples/plantilla_ml_ejemplo.xlsx'

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Plantilla ML"

# Headers en fila 1
headers = list(data.keys())
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)
    ws.cell(row=1, column=col).font = Font(bold=True)

# Fila 2-3 vacías (como en plantillas ML reales)
# Fila 4: Obligatorio markers
obligatory_markers = ['obligatorio', 'obligatorio', 'obligatorio', 'obligatorio', 'opcional', 'opcional', 'obligatorio', 'opcional', 'opcional', 'obligatorio', 'opcional']
for col, marker in enumerate(obligatory_markers, 1):
    cell = ws.cell(row=4, column=col, value=marker)
    if marker == 'obligatorio':
        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        cell.font = Font(color="CC0000", bold=True)

# Fila 5: Tipos de datos
data_types = ['texto', 'número', 'número', 'texto', 'texto', 'texto', 'texto', 'número', 'texto', 'texto', 'url']
for col, dtype in enumerate(data_types, 1):
    cell = ws.cell(row=5, column=col, value=dtype)
    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    cell.font = Font(color="0066CC", italic=True)

# Filas 6-7 vacías (separación como en plantillas ML)

# Datos de productos desde fila 8
for i, row_num in enumerate(range(8, 8 + len(data['Título']))):
    for col, key in enumerate(data.keys(), 1):
        ws.cell(row=row_num, column=col, value=data[key][i])

# Ajustar ancho de columnas
for col in range(1, len(headers) + 1):
    max_length = 0
    column = openpyxl.utils.get_column_letter(col)
    for row in range(1, ws.max_row + 1):
        try:
            if len(str(ws[column + str(row)].value)) > max_length:
                max_length = len(str(ws[column + str(row)].value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Guardar archivo
wb.save(output_path)
print(f"Plantilla ML de ejemplo creada en: {output_path}")
