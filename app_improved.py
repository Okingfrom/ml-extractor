#!/usr/bin/env python3
"""
Aplicacin Flask mejorada para mapeo de productos a Mercado Libre
Mantiene la estructura exacta de ML y permite mapeo selectivo
CON INTELIGENCIA ARTIFICIAL para autocompletar datos faltantes
SISTEMA DE AUTENTICACIN Y USUARIOS PREMIUM/GRATUITOS
"""

from flask import Flask, request, render_template_string, send_file, redirect, url_for, jsonify, session, render_template, flash, send_from_directory
import openpyxl
from openpyxl.cell import MergedCell
import csv
import docx
import PyPDF2
import yaml
import os
import tempfile
from datetime import datetime, timedelta
from dotenv import load_dotenv
from auth_system import user_manager, login_required, premium_required
from flask_cors import CORS

# Cargar variables de entorno
load_dotenv()

#  DEVELOPER MODE - Set to True for easy testing, False for production
DEVELOPER_MODE = False
DISABLE_AUTH = False  # Disable developer bypass for safe merge; enable real auth
DEVELOPER_USER = {
    'user_id': 999,
    'username': 'developer',
    'email': 'dev@test.com',
    'first_name': 'Developer',
    'last_name': 'Mode',
    'account_type': 'premium',
    'user_type': 'premium'
}

#  FUNCIN HELPER PARA MANEJO SEGURO DE CELDAS
def safe_get_cell_value(sheet, row, col):
    """Obtiene el valor de una celda manejando MergedCell de forma segura"""
    try:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # Buscar el valor en la celda principal del rango fusionado
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
                    return top_left.value
            return None
        else:
            return cell.value
    except Exception as e:
        print(f" Error accediendo celda ({row},{col}): {e}")
        return None

def safe_set_cell_value(sheet, row, col, value):
    """Establece el valor de una celda manejando MergedCell de forma segura"""
    try:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # Para celdas fusionadas, escribir en la celda principal
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
                    top_left.value = value
                    return True
            # Si no encontramos el rango, intentar escribir directamente
            safe_set_cell_value(sheet, row, col, value)
            return True
        else:
            cell.value = value
            return True
    except Exception as e:
        print(f" Error escribiendo celda ({row},{col}): {e}")
        return False
from werkzeug.utils import secure_filename
import shutil
from ai_enhancer import AIProductEnhancer, AI_CONFIG

USE_REACT_UI = os.getenv('USE_REACT_UI') == '1'
FRONTEND_DIST = os.path.join(os.path.dirname(__file__), 'frontend', 'dist')

static_opts = {}
if USE_REACT_UI:
    target_static = None
    if os.path.isdir(FRONTEND_DIST):
        target_static = FRONTEND_DIST
    if target_static:
        static_opts = {'static_folder': target_static, 'static_url_path': '/'}

app = Flask(__name__, **static_opts)
app.secret_key = os.getenv('SECRET_KEY', 'dev_secret_key_ml_extractor_2025_very_secure')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Sesi贸n v谩lida por 24 horas

# Configuraci贸n robusta de sesiones
app.config['SESSION_COOKIE_SECURE'] = False  # Para desarrollo (HTTP)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_NAME'] = 'ml_extractor_session'
app.config['APPLICATION_ROOT'] = '/'

# Enable CORS for API endpoints in dev mode when React UI enabled
if USE_REACT_UI and os.getenv('ML_EXTRACTOR_DEV') == '1':
    CORS(app, resources={r"/api/*": {"origins": ["http://localhost:3000", "http://localhost:5173", "http://localhost:3002"]}}, supports_credentials=True)

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

"""DEBUG/DEV ROUTES REMOVED IN PRODUCTION BUILD

The following routes were present during development but are now removed / disabled:
  - /dev-login (auto-login as premium)
  - /switch-user/<user_type> (role switching without auth)
  - /test-auth (session inspection)
  - /debug-session (raw session/cookie dump)

They can be re-enabled locally by creating a developer settings module and
guarding them behind a strict environment variable (e.g., ML_EXTRACTOR_DEV=1).
"""

# Environment flag for conditional dev routes
DEV_FLAG = os.getenv('ML_EXTRACTOR_DEV') == '1'

if DEV_FLAG:
    @app.route('/dev-login')
    def dev_login():  # pragma: no cover - development only
        """Lightweight dev login: assigns a premium session if none exists."""
        if not session.get('user_id'):
            session.permanent = True
            session['user_id'] = DEVELOPER_USER['user_id']
            session['username'] = DEVELOPER_USER['username']
            session['email'] = DEVELOPER_USER['email']
            session['first_name'] = DEVELOPER_USER['first_name']
            session['last_name'] = DEVELOPER_USER['last_name']
            session['account_type'] = DEVELOPER_USER['account_type']
            session['user_type'] = DEVELOPER_USER['user_type']
            flash('Dev login activo (ML_EXTRACTOR_DEV=1).', 'success')
        else:
            flash('Sesi贸n existente reutilizada.', 'info')
        return redirect(url_for('index'))

    @app.route('/switch-user/<user_type>')
    def switch_user(user_type):  # pragma: no cover - development only
        user_type_norm = user_type.lower()
        if user_type_norm == 'premium':
            session['account_type'] = 'premium'
            session['user_type'] = 'premium'
            flash('Modo premium (DEV).', 'success')
        elif user_type_norm in ('regular', 'free'):
            session['account_type'] = 'free'
            session['user_type'] = 'free'
            flash('Modo gratuito (DEV).', 'warning')
        else:
            flash('Tipo inv谩lido.', 'error')
        return redirect(url_for('index'))

    @app.route('/test-auth')
    def test_auth():  # pragma: no cover - development only
        return f"""
        <h1> Test Auth (DEV)</h1>
        <p>ML_EXTRACTOR_DEV=1</p>
        <p>User: {session.get('username')}</p>
        <p>Account Type: {session.get('account_type')}</p>
        <p>Session Keys: {list(session.keys())}</p>
        <a href='/debug-session'>Debug Session</a>
        """

    @app.route('/debug-session')
    def debug_session():  # pragma: no cover - development only
        return f"""
        <h1> Debug Session (DEV)</h1>
        <pre>{dict(session)}</pre>
        <p>Cookies: {dict(request.cookies)}</p>
        <a href='/'>Home</a>
        """

#  USER TYPE SWITCHER FOR TESTING
## Original /switch-user route removed (see production hardening note above)

# ==========================================
# RUTAS DE AUTENTICACIN Y USUARIOS
# ==========================================

## Original /test-auth and /debug-session routes removed (see production hardening note above)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Pgina de login"""
    if request.method == 'GET':
        return render_template('login.html')
    
    # POST - manejar login directamente aqu
    try:
        if request.content_type == 'application/json':
            data = request.get_json()
        else:
            data = request.form.to_dict()
            
        email = data.get('email')
        password = data.get('password')
        
        success, result = user_manager.login_user(email, password, request.remote_addr)
        
        if success:
            # Crear sesi贸n
            user = result
            session.permanent = True  # Hacer la sesi贸n permanente
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['email'] = user['email']
            session['first_name'] = user['first_name']
            session['last_name'] = user['last_name']
            session['account_type'] = user['account_type']
            session['user_type'] = user['user_type']
            
            # Debug: verificar que la sesi贸n se guarde
            print(f" DEBUG: Sesi贸n creada para user_id: {session.get('user_id')}")
            print(f" DEBUG: Session keys despus de login: {list(session.keys())}")
            print(f" DEBUG: Session ID: {request.cookies.get('session')}")
            print(f" DEBUG: Todas las cookies: {dict(request.cookies)}")
            
            # Si es request JSON (AJAX), devolver JSON
            if request.content_type == 'application/json':
                return jsonify({
                    'success': True,
                    'message': f'Bienvenido {user["first_name"]}!',
                    'user': {
                        'username': user['username'],
                        'first_name': user['first_name'],
                        'account_type': user['account_type'],
                        'user_type': user['user_type']
                    }
                })
            else:
                # Si es form submit tradicional, redirigir
                flash(f'Bienvenido {user["first_name"]}!', 'success')
                return redirect('/')
        else:
            # Error en login
            if request.content_type == 'application/json':
                return jsonify({
                    'success': False,
                    'error': result
                }), 401
            else:
                flash(result, 'error')
                return redirect('/login')
                
    except Exception as e:
        print(f" Error en login: {e}")
        if request.content_type == 'application/json':
            return jsonify({
                'success': False,
                'error': f'Error interno: {str(e)}'
            }), 500
        else:
            flash('Error interno del servidor', 'error')
            return redirect('/login')

@app.route('/register')
def register():
    """Pgina de registro"""
    return render_template('register.html')

@app.route('/verify')
def verify():
    """Pgina de verificacin"""
    return render_template('verify.html')

@app.route('/logout')
def logout():
    """Cerrar sesi贸n"""
    user_id = session.get('user_id')
    if user_id:
        user_manager.db.log_activity(user_id, 'user_logout', None, request.remote_addr)
    
    session.clear()
    return redirect('/login?message=' + 
                   'Sesi贸n cerrada exitosamente&type=success')

# ==========================================
# API ENDPOINTS DE AUTENTICACIN
# ==========================================

@app.route('/api/register', methods=['POST'])
def api_register():
    """API endpoint para registro de usuarios"""
    try:
        data = request.get_json()
        
        success, message = user_manager.register_user(data, request.remote_addr)
        
        return jsonify({
            'success': success,
            'message' if success else 'error': message
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500

@app.route('/api/login', methods=['POST'])
def api_login():
    """API endpoint para login de usuarios"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        success, result = user_manager.login_user(email, password, request.remote_addr)
        
        if success:
            # Crear sesi贸n
            user = result
            session.permanent = True  # Hacer la sesi贸n permanente
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['email'] = user['email']
            session['first_name'] = user['first_name']
            session['last_name'] = user['last_name']
            session['account_type'] = user['account_type']
            session['user_type'] = user['user_type']
            
            # Debug: verificar que la sesi贸n se guarde
            print(f" DEBUG: Sesi贸n creada para user_id: {session.get('user_id')}")
            
            return jsonify({
                'success': True,
                'message': f'Bienvenido {user["first_name"]}!',
                'user': {
                    'username': user['username'],
                    'first_name': user['first_name'],
                    'account_type': user['account_type'],
                    'user_type': user['user_type']
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': result
            }), 401
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500

@app.route('/api/verify', methods=['POST'])
def api_verify():
    """API endpoint para verificacin de cuenta"""
    try:
        data = request.get_json()
        email = data.get('email')
        code = data.get('code')
        
        success, message = user_manager.verify_account(email, code)
        
        return jsonify({
            'success': success,
            'message' if success else 'error': message
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500

@app.route('/api/resend-verification', methods=['POST'])
def api_resend_verification():
    """API endpoint para reenviar c贸digo de verificaci贸n"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        # Obtener usuario
        user = user_manager.db.get_user_by_email(email)
        if not user:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
    # Generar nuevo c贸digo
        new_code = user_manager.generate_verification_code()
        
        # Actualizar en base de datos
        conn = user_manager.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE users SET verification_code = ? WHERE email = ?
        ''', (new_code, email))
        conn.commit()
        conn.close()
        
        # Enviar email
        user_manager.notifications.send_verification_email(
            email, new_code, user['first_name']
        )
        
        return jsonify({
            'success': True,
            'message': 'Nuevo c贸digo enviado'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500

# ==========================================
# RUTAS DE GOOGLE OAUTH (GRATIS)
# ==========================================

@app.route('/auth/google')
def google_auth():
    """Iniciar autenticacin con Google OAuth"""
    try:
        from google_oauth_config import google_oauth
        
        auth_url, state = google_oauth.get_authorization_url()
        if auth_url:
            session['oauth_state'] = state
            return redirect(auth_url)
        else:
            return redirect('/login?message=' + 
                           'Error configurando Google OAuth&type=error')
    except ImportError:
        return redirect('/login?message=' + 
                       'Google OAuth no configurado&type=warning')
    except Exception as e:
        return redirect('/login?message=' + 
                       f'Error OAuth: {str(e)}&type=error')

@app.route('/auth/google/callback')
def google_callback():
    """Callback de Google OAuth"""
    try:
        from google_oauth_config import google_oauth
        
        # Obtener parmetros
        code = request.args.get('code')
        state = request.args.get('state')
        error = request.args.get('error')
        
        if error:
            return redirect('/login?message=' + 
                           f'Error OAuth: {error}&type=error')
        
        if not code:
            return redirect('/login?message=' + 
                           'C贸digo OAuth no recibido&type=error')
        
        # Verificar state (seguridad)
        if state != session.get('oauth_state'):
            return redirect('/login?message=' + 
                           'Estado OAuth inv谩lido&type=error')
        
        # Intercambiar cdigo por token y obtener datos del usuario
        google_data = google_oauth.exchange_code_for_token(code, state)
        
        if not google_data:
            return redirect('/login?message=' + 
                           'Error obteniendo datos de Google&type=error')
        
        # Crear o actualizar usuario
        success, result = user_manager.login_with_google(google_data, request.remote_addr)
        
        if success:
            user = result
            # Crear sesi贸n
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['email'] = user['email']
            session['first_name'] = user['first_name']
            session['last_name'] = user['last_name']
            session['account_type'] = user['account_type']
            session['user_type'] = user['user_type']
            session['google_login'] = True
            
            # Limpiar estado OAuth
            session.pop('oauth_state', None)
            
            return redirect('/?message=' + 
                           f'Bienvenido {user["first_name"]}! Autenticado con Google&type=success')
        else:
            return redirect('/login?message=' + 
                           f'Error Google: {result}&type=error')
            
    except ImportError:
        return redirect('/login?message=' + 
                       'Google OAuth no configurado&type=warning')
    except Exception as e:
        return redirect('/login?message=' + 
                       f'Error OAuth: {str(e)}&type=error')

# ==========================================
# MIDDLEWARE DE AUTENTICACIN
# ==========================================

@app.before_request
def require_auth():
    """Middleware que requiere autenticacin para ciertas rutas"""
    
    #  TESTING MODE: Disable authentication completely
    if DISABLE_AUTH:
        # Auto-setup session if not exists
        if 'user_id' not in session:
            # Default to premium user for testing
            session.permanent = True
            session['user_id'] = DEVELOPER_USER['user_id']
            session['username'] = DEVELOPER_USER['username']
            session['email'] = DEVELOPER_USER['email']
            session['first_name'] = DEVELOPER_USER['first_name']
            session['last_name'] = DEVELOPER_USER['last_name']
            session['account_type'] = DEVELOPER_USER['account_type']
            session['user_type'] = DEVELOPER_USER['user_type']
        return  # Skip all authentication checks
    
    # Rutas pblicas que no requieren autenticacin
    public_routes = [
        '/login', '/register', '/verify', '/logout', '/test-auth', '/debug-session',
        '/api/login', '/api/register', '/api/verify', '/api/resend-verification',
        '/auth/google', '/auth/google/callback',
        '/dev-login',  #  Developer bypass route
        '/static'
    ]
    
    # Verificar si la ruta actual es pblica
    for route in public_routes:
        if request.path.startswith(route):
            return
    
    # Verificar si el usuario est autenticado
    if 'user_id' not in session:
        print(f" DEBUG: Acceso denegado a {request.path}")
        print(f" DEBUG: Session keys: {list(session.keys())}")
        print(f" DEBUG: Session ID en cookies: {request.cookies.get('session')}")
        print(f" DEBUG: Todas las cookies recibidas: {dict(request.cookies)}")
        print(f" DEBUG: Headers relevantes: {dict(request.headers)}")
        if request.path.startswith('/api/'):
            return jsonify({
                'success': False,
                'error': 'Autenticacin requerida'
            }), 401
        else:
            return redirect('/login?message=' + 
                          'Debes iniciar sesi贸n para acceder&type=warning')
    else:
        print(f" DEBUG: Acceso permitido a {request.path} - user_id: {session.get('user_id')}")

# ==========================================
# FUNCIONES DE VALIDACIN PREMIUM
# ==========================================

def validate_premium_feature(feature_name):
    """Valida si el usuario puede usar una funcionalidad premium"""
    if session.get('account_type') != 'premium':
        return False, f'{feature_name} requiere cuenta Premium'
    return True, 'Acceso autorizado'

# Inicializar AI enhancer con la API key del entorno
ai_enhancer = AIProductEnhancer(
    provider=os.getenv('AI_PROVIDER', 'groq'),
    api_key=os.getenv('GROQ_API_KEY')
)

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ML Bulk Mapper Pro - {{ user_info.get('first_name', 'Usuario') if user_info else 'Usuario' }}</title>
    <style>
        /*  DISEO MOBILE-FIRST RESPONSIVE */
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }
        
        /* Header de usuario */
        .user-header {
            background: linear-gradient(135deg, #3483fa, #2968c8);
            color: white;
            padding: 15px 20px;
            border-radius: 12px;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 10px;
            box-shadow: 0 4px 20px rgba(52, 131, 250, 0.25);
        }
        
        .user-info {
            display: flex;
            align-items: center;
            gap: 15px;
        }
        
        .user-avatar {
            width: 40px;
            height: 40px;
            background: rgba(255,255,255,0.2);
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: 600;
            font-size: 16px;
        }
        
        .user-details h3 {
            margin: 0;
            font-size: 16px;
            font-weight: 600;
        }
        
        .user-details p {
            margin: 0;
            font-size: 14px;
            opacity: 0.9;
        }

        /* Small text utility for high contrast on gradients */
        .small-contrast {
            color: rgba(0,0,0,0.85) !important; /* dark text for legibility */
            text-shadow: 0 1px 0 rgba(255,255,255,0.6); /* subtle light lift */
            font-size: 14px;
            opacity: 0.95;
        }

        /* White-on-dark pill for small text over colorful gradients */
        .small-pill {
            display: inline-block;
            background: rgba(0,0,0,0.65); /* slightly darker for contrast */
            color: #ffffff !important;
            padding: 8px 12px; /* increased padding */
            border-radius: 999px;
            font-size: 13px;
            line-height: 1;
            box-shadow: 0 6px 18px rgba(0,0,0,0.28);
            backdrop-filter: blur(6px);
            -webkit-backdrop-filter: blur(6px);
            border: 1px solid rgba(255,255,255,0.06);
            opacity: 0.98;
        }

        /* Apply pill styling to all <small> by default, opt-out with .no-pill */
        small:not(.no-pill) {
            display: inline-block;
            background: rgba(0,0,0,0.55);
            color: #ffffff !important;
            padding: 6px 10px;
            border-radius: 999px;
            font-size: 12px;
            line-height: 1;
            box-shadow: 0 4px 12px rgba(0,0,0,0.22);
            backdrop-filter: blur(4px);
            -webkit-backdrop-filter: blur(4px);
            border: 1px solid rgba(255,255,255,0.05);
            opacity: 0.95;
            margin-left: 4px;
        }
        
        .account-badge {
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .badge-premium {
            background: linear-gradient(45deg, #ffd700, #ffed4e);
            color: #333;
        }
        
        .badge-free {
            background: rgba(255,255,255,0.2);
            color: white;
        }
        
        .user-actions {
            display: flex;
            gap: 10px;
            align-items: center;
        }
        
        .btn-logout {
            background: rgba(255,255,255,0.2);
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 20px;
            font-size: 14px;
            cursor: pointer;
            transition: all 0.3s ease;
            text-decoration: none;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        
        .btn-logout:hover {
            background: rgba(255,255,255,0.3);
            color: white;
            text-decoration: none;
        }
        
        /* Alert para usuarios gratuitos */
        .premium-alert {
            background: linear-gradient(135deg, rgba(255, 241, 118, 0.3), rgba(255, 234, 167, 0.5));
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 234, 167, 0.7);
            color: #856404;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 12px;
            box-shadow: 0 4px 16px rgba(255, 193, 7, 0.15);
        }
        
        .premium-alert i {
            font-size: 20px;
            color: #f39c12;
        }
        
        body {
            font-family: 'Proxima Nova', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #fff159 0%, #ffed69 50%, #3483fa 100%);
            min-height: 100vh;
            padding: 20px;
            line-height: 1.6;
            color: #333;
        }        /*  MOBILE FIRST - BASE STYLES */
        .container {
            background: rgba(255, 255, 255, 1);
            backdrop-filter: blur(15px);
            padding: 30px;
            border-radius: 16px;
            box-shadow: 0 8px 32px rgba(52, 131, 250, 0.2);
            border: 1px solid rgba(255, 255, 255, 0.3);
            max-width: 1200px;
            width: 100%;
            margin: 0 auto;
        }
        
        .header {
            text-align: center;
            margin-bottom: 20px;
            color: #3483fa;
        }
        
        .ml-logo {
            background: linear-gradient(135deg, #3483fa, #2968c8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            font-weight: 700;
            font-size: 1.5rem;
        }
        
        .ai-badge {
            background: linear-gradient(135deg, #00a650, #00c457);
            color: white;
            padding: 4px 10px;
            border-radius: 16px;
            font-size: 10px;
            font-weight: 600;
            display: inline-block;
            margin-left: 6px;
        }
        
        .form-group { 
            margin-bottom: 20px; 
        }
        
        label { 
            display: block; 
            margin-bottom: 8px; 
            font-weight: 600; 
            color: #333;
            font-size: 14px;
        }
        
        /*  INPUTS OPTIMIZADOS PARA MVIL */
        input[type="file"], 
        input[type="text"], 
        input[type="password"], 
        input[type="number"], 
        select, 
        textarea { 
            width: 100%; 
            padding: 14px 16px; 
            border: 2px solid #e6e6e6;
            border-radius: 8px;
            font-size: 16px; /* Evita zoom en iOS */
            transition: all 0.3s ease;
            background: rgba(255, 255, 255, 0.9);
            -webkit-appearance: none;
            appearance: none;
            min-height: 44px; /* Toque mnimo recomendado */
        }
        
        input:focus, select:focus, textarea:focus {
            border-color: #3483fa;
            background: rgba(255, 255, 255, 1);
            box-shadow: 0 0 0 3px rgba(52, 131, 250, 0.15);
            outline: none;
        }
        
        /*  CHECKBOXES MVIL-FRIENDLY */
        .checkbox-grid {
            background: rgba(248, 249, 250, 0.9);
            backdrop-filter: blur(5px);
            padding: 12px;
            border-radius: 8px;
            border: 1px solid rgba(233, 236, 239, 0.8);
        }
        
        .checkbox-grid label {
            display: flex;
            align-items: center;
            margin-bottom: 12px;
            font-weight: 500;
            font-size: 14px;
            cursor: pointer;
            transition: color 0.2s ease;
            min-height: 44px;
            padding: 8px;
        }
        
        .checkbox-grid label:hover {
            color: #3483fa;
            background: rgba(52, 131, 250, 0.05);
            border-radius: 6px;
        }
        
        .checkbox-grid input[type="checkbox"] {
            width: 20px;
            height: 20px;
            margin-right: 12px;
            accent-color: #3483fa;
        }
        
        /*  BOTONES OPTIMIZADOS PARA TOUCH */
        button { 
            background: linear-gradient(135deg, #3483fa, #2968c8);
            color: white; 
            padding: 16px 24px; 
            border: none; 
            cursor: pointer;
            border-radius: 8px;
            font-size: 16px;
            font-weight: 600;
            transition: all 0.2s ease;
            width: 100%;
            min-height: 48px;
            touch-action: manipulation;
        }
        
        button:hover, button:focus { 
            transform: translateY(-1px);
            box-shadow: 0 6px 20px rgba(52, 131, 250, 0.3);
        }
        
        button:active {
            transform: translateY(0);
        }
        
        /*  MENSAJES RESPONSIVE */
        .success, .error, .info-text, .ai-info { 
            padding: 16px;
            border-radius: 8px;
            margin: 16px 0;
            font-size: 14px;
            word-wrap: break-word;
        }
        
        .success { 
            color: #00a650; 
            background: linear-gradient(135deg, #e8f5e8, #d4f4d4);
            border-left: 4px solid #00a650;
        }
        
        .error { 
            color: #ff3333; 
            background: linear-gradient(135deg, #ffe6e6, #ffcccc);
            border-left: 4px solid #ff3333;
        }
        
        /*  SECCIONES COLAPSABLES MVIL */
        .ai-section {
            background: linear-gradient(135deg, rgba(52, 131, 250, 0.08), rgba(52, 131, 250, 0.12));
            backdrop-filter: blur(5px);
            padding: 16px;
            border-radius: 8px;
            margin: 20px 0;
            border: 2px solid rgba(52, 131, 250, 0.3);
            box-shadow: 0 4px 16px rgba(52, 131, 250, 0.1);
        }
        
        .mapping-section {
            background: rgba(250, 250, 250, 0.8);
            backdrop-filter: blur(5px);
            padding: 16px;
            border-radius: 8px;
            margin: 16px 0;
            border: 1px solid rgba(230, 230, 230, 0.8);
        }
        
        .manual-config-section {
            background: linear-gradient(135deg, rgba(255, 248, 225, 0.9), rgba(255, 243, 196, 0.9));
            backdrop-filter: blur(5px);
            padding: 16px;
            border-radius: 8px;
            margin: 20px 0;
            border: 2px solid rgba(255, 152, 0, 0.4);
            box-shadow: 0 4px 16px rgba(255, 152, 0, 0.1);
        }
        
        .manual-config-section h3 {
            color: #f57c00;
            margin-bottom: 16px;
            font-size: 1.1rem;
        }
        
        .manual-config-section textarea {
            font-family: 'Courier New', monospace;
            font-size: 14px;
            resize: vertical;
            min-height: 120px;
        }
        
        .manual-config-section input[type="radio"] {
            width: 20px;
            height: 20px;
            margin-right: 10px;
            accent-color: #ff9800;
        }
        
        /*  GRID RESPONSIVE PARA CHECKBOXES */
        .checkbox-group {
            display: grid;
            grid-template-columns: 1fr;
            gap: 8px;
            margin-top: 16px;
        }
        
        .checkbox-item {
            display: flex;
            align-items: center;
            padding: 12px;
            background: white;
            border-radius: 8px;
            border: 1px solid #e6e6e6;
            transition: all 0.2s ease;
            min-height: 48px;
        }
        
        .checkbox-item:hover, .checkbox-item:focus-within {
            border-color: #3483fa;
            box-shadow: 0 2px 8px rgba(52, 131, 250, 0.1);
        }
        
        .checkbox-item input[type="checkbox"] {
            width: 20px;
            height: 20px;
            margin-right: 12px;
            accent-color: #3483fa;
        }
        
    /*  TEXTO Y TTULOS RESPONSIVE */
        h1 { 
            color: #3483fa; 
            font-weight: 700; 
            margin-bottom: 8px; 
            font-size: 1.8rem;
        }
        
        h3 { 
            color: #333; 
            margin-top: 24px; 
            font-weight: 600; 
            font-size: 1.2rem;
        }
        
        .required { color: #ff3333; font-weight: 600; }
        .optional { color: #666; }
        .ai-enhanced { color: #00a650; font-weight: 600; }
        
        .info-text {
            background: linear-gradient(135deg, #e3f2fd, #e1f5fe);
            border-left: 4px solid #3483fa;
            color: #1565c0;
        }
        
        .ai-info {
            background: linear-gradient(135deg, #e8f5e8, #f1f8e9);
            border-left: 4px solid #00a650;
            color: #2e7d32;
        }
        
        .api-selector {
            background: white;
            padding: 16px;
            border-radius: 8px;
            border: 2px solid #e6e6e6;
            margin: 12px 0;
        }
        
        .api-option {
            margin: 12px 0;
            display: flex;
            align-items: center;
            min-height: 44px;
        }
        
        .api-option input[type="radio"] {
            width: 20px;
            height: 20px;
            margin-right: 12px;
        }
        
        /*  TABLET STYLES (768px+) */
        @media (min-width: 768px) {
            body {
                padding: 20px;
            }
            
            .container {
                padding: 24px;
                max-width: 800px;
            }
            
            .ml-logo {
                font-size: 2rem;
            }
            
            .checkbox-group {
                grid-template-columns: repeat(2, 1fr);
                gap: 12px;
            }
            
            h1 {
                font-size: 2.2rem;
            }
            
            .ai-section, .manual-config-section {
                padding: 24px;
            }
        }
        
        /*  DESKTOP STYLES (1024px+) */
        @media (min-width: 1024px) {
            body {
                padding: 30px;
            }
            
            .container {
                padding: 30px;
                max-width: 1200px;
            }
            
            .checkbox-group {
                grid-template-columns: repeat(3, 1fr);
                gap: 16px;
            }
            
            .form-row {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                align-items: start;
            }
            
            .ai-section, .manual-config-section {
                padding: 30px;
            }
        }
        
        /*  LARGE DESKTOP (1440px+) */
        @media (min-width: 1440px) {
            .checkbox-group {
                grid-template-columns: repeat(4, 1fr);
            }
        }
        
        /*  NAVEGACIN TCTIL MEJORADA */
        @media (hover: none) and (pointer: coarse) {
            button, .checkbox-item, .api-option {
                min-height: 48px;
            }
            
            input, select, textarea {
                min-height: 48px;
                font-size: 16px;
            }
        }
        
        /*  MODO OSCURO (si el dispositivo lo prefiere) */
        @media (prefers-color-scheme: dark) {
            /* Keep main container light to preserve Mercado Libre branding in testing
               Use a subtle translucent white so elements remain readable */
            .container {
                background: rgba(255, 255, 255, 0.95);
                color: #222;
                border-color: rgba(255,255,255,0.12);
                box-shadow: 0 8px 32px rgba(0,0,0,0.08);
            }

            /* Inputs should remain light for clarity during testing */
            input, select, textarea {
                background: rgba(255,255,255,0.9);
                color: #111;
                border-color: rgba(0,0,0,0.08);
            }

            .checkbox-item {
                background: rgba(255,255,255,0.92);
                border-color: rgba(0,0,0,0.06);
            }
        }
        
        /*  ANIMACIONES REDUCIDAS PARA ACCESIBILIDAD */
        @media (prefers-reduced-motion: reduce) {
            *, *::before, *::after {
                animation-duration: 0.01ms !important;
                animation-iteration-count: 1 !important;
                transition-duration: 0.01ms !important;
            }
        }
            padding: 12px;
            border-radius: 6px;
            cursor: pointer;
            transition: background 0.2s ease;
        }
        .api-option:hover {
            background: #f5f5f5;
        }
        .api-option input[type="radio"] {
            margin-right: 12px;
            accent-color: #3483fa;
        }
        .cost-badge {
            background: #00a650;
            color: white;
            padding: 2px 8px;
            border-radius: 10px;
            font-size: 10px;
            font-weight: 600;
            margin-left: 6px;
        }
        .quality-stars {
            color: #ffb400;
            margin-left: 6px;
            font-size: 12px;
        }
        .debug-section {
            background: #2d2d2d;
            color: #00ff41;
            padding: 20px;
            border-radius: 6px;
            margin: 20px 0;
            font-family: 'Courier New', monospace;
            font-size: 12px;
            max-height: 300px;
            overflow-y: auto;
            border: 2px solid #555;
        }
        .debug-title {
            color: #00ff41;
            font-weight: bold;
            margin-bottom: 10px;
        }
        .creator-signature {
            text-align: center;
            margin-top: 40px;
            padding: 20px;
            background: linear-gradient(135deg, #f5f5f5, #eeeeee);
            border-radius: 6px;
            color: #666;
            font-size: 14px;
        }
        .creator-name {
            color: #3483fa;
            font-weight: 600;
        }
        
        /*  NUEVA SECCIN: LOADING SCREEN CREATIVO */
        .loading-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.8);
            backdrop-filter: blur(5px);
            display: none;
            justify-content: center;
            align-items: center;
            z-index: 9999;
        }
        .loading-container {
            background: white;
            padding: 40px;
            border-radius: 20px;
            text-align: center;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            max-width: 500px;
            width: 90%;
        }
        .loading-logo {
            font-size: 32px;
            background: linear-gradient(135deg, #3483fa, #2968c8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            font-weight: 700;
            margin-bottom: 20px;
        }
        .progress-container {
            background: #f0f0f0;
            border-radius: 25px;
            padding: 4px;
            margin: 20px 0;
        }
        .progress-bar {
            background: linear-gradient(135deg, #3483fa, #00a650);
            border-radius: 20px;
            height: 30px;
            width: 0%;
            transition: width 0.5s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 600;
            font-size: 14px;
        }
        .loading-steps {
            text-align: left;
            margin: 20px 0;
        }
        .loading-step {
            display: flex;
            align-items: center;
            padding: 8px 0;
            color: #666;
            transition: all 0.3s ease;
        }
        .loading-step.active {
            color: #3483fa;
            font-weight: 600;
        }
        .loading-step.completed {
            color: #00a650;
        }
        .loading-step-icon {
            margin-right: 12px;
            font-size: 18px;
        }
        .ml-animation {
            display: inline-block;
            animation: bounce 2s infinite;
        }
        @keyframes bounce {
            0%, 20%, 50%, 80%, 100% { transform: translateY(0); }
            40% { transform: translateY(-10px); }
            60% { transform: translateY(-5px); }
        }
        small { color: #666; font-size: 12px; }
        
        /*  ESTILOS ADICIONALES PARA LOADING MVIL */
        .step-icon {
            font-size: 18px;
            margin-right: 12px;
            display: inline-block;
            min-width: 24px;
        }
        
        .step-text {
            font-size: 14px;
            color: #333;
        }
        
        .loading-step.active .step-icon {
            animation: pulse 1s infinite;
        }
        
        .loading-step.active .step-text {
            color: #3483fa;
            font-weight: 600;
        }
        
        @keyframes pulse {
            0% { transform: scale(1); }
            50% { transform: scale(1.1); }
            100% { transform: scale(1); }
        }
        
        /*  MEJORAS ESPECFICAS PARA MVIL */
        @media (max-width: 480px) {
            .loading-container {
                padding: 24px 20px;
                margin: 20px;
            }
            
            .loading-logo {
                font-size: 24px;
            }
            
            .progress-bar {
                height: 24px;
                font-size: 12px;
            }
            
            /* Prompt section mobile styles */
            .prompt-generation-collapsible .prompt-generation-header {
                padding: 14px 16px !important;
                flex-direction: column;
                gap: 8px;
                text-align: center;
            }
            
            .prompt-generation-header h3 {
                font-size: 16px !important;
            }
            
            .prompt-section-content {
                border-radius: 0 0 8px 8px !important;
            }
            
            .prompt-section-content div[style*="padding: 24px"] {
                padding: 16px !important;
            }
            
            .prompt-section-content div[style*="grid-template-columns"] {
                grid-template-columns: 1fr !important;
                gap: 12px !important;
            }
            
            textarea#product_prompt {
                min-height: 80px !important;
                font-size: 16px !important;
            }
            
            button[type="submit"] {
                padding: 14px 20px !important;
                font-size: 16px !important;
                min-width: auto !important;
                width: 100%;
            }
            
            .ai-prompt-config {
                padding: 16px !important;
                margin: 16px 0 !important;
            }
            
            .ai-prompt-config h4 {
                font-size: 16px !important;
                margin-bottom: 12px !important;
            }
            }
            
            .loading-step {
                padding: 8px 12px;
                margin: 6px 0;
            }
            
            .step-icon {
                font-size: 16px;
                margin-right: 8px;
                min-width: 20px;
            }
            
            .step-text {
                font-size: 13px;
            }
        }
    </style>
    
    <script>
        //  JAVASCRIPT MVIL-OPTIMIZADO
        
        // Prevenir zoom accidental en iOS
        document.addEventListener('gesturestart', function (e) {
            e.preventDefault();
        });
        
        // Mejorar experiencia tctil
        document.addEventListener('DOMContentLoaded', function() {
            // Agregar clases para dispositivos tctiles
            if ('ontouchstart' in window) {
                document.body.classList.add('touch-device');
            }
            
            // Mejorar scroll en dispositivos mviles
            if (window.innerWidth <= 768) {
                document.body.style.overflowX = 'hidden';
            }
            
            // Auto-expandir textareas en mvil
            const textareas = document.querySelectorAll('textarea');
            textareas.forEach(textarea => {
                textarea.addEventListener('focus', function() {
                    if (window.innerWidth <= 768) {
                        setTimeout(() => {
                            this.scrollIntoView({ behavior: 'smooth', block: 'center' });
                        }, 300);
                    }
                });
            });
            
            // Mejorar UX de checkboxes en mvil
            const checkboxItems = document.querySelectorAll('.checkbox-item');
            checkboxItems.forEach(item => {
                item.addEventListener('click', function(e) {
                    if (e.target.type !== 'checkbox') {
                        const checkbox = this.querySelector('input[type="checkbox"]');
                        if (checkbox) {
                            checkbox.checked = !checkbox.checked;
                            checkbox.dispatchEvent(new Event('change'));
                        }
                    }
                });
            });
            
            // Loading responsivo
            const form = document.querySelector('form');
            if (form) {
                form.addEventListener('submit', function() {
                    showMobileLoader();
                });
            }
        });
        
        // Funciones existentes adaptadas para mvil
        function showMobileLoader() {
            const overlay = document.getElementById('loadingOverlay');
            if (overlay) {
                overlay.style.display = 'flex';
                updateProgress();
            }
        }
        
        function updateProgress() {
            const progressBar = document.querySelector('.progress-bar');
            const steps = document.querySelectorAll('.loading-step');
            let progress = 0;
            
            const interval = setInterval(() => {
                progress += Math.random() * 15;
                if (progress > 100) progress = 100;
                
                if (progressBar) {
                    progressBar.style.width = progress + '%';
                    progressBar.textContent = Math.round(progress) + '%';
                }
                
                // Activar pasos
                const currentStep = Math.floor((progress / 100) * steps.length);
                steps.forEach((step, index) => {
                    if (index <= currentStep) {
                        step.classList.add('active');
                    }
                });
                
                if (progress >= 100) {
                    clearInterval(interval);
                }
            }, 200);
        }
        
        // Optimizar formularios para mvil
        function toggleHelpSection() {
            const section = document.getElementById('helpSection');
            if (section) {
                const isVisible = section.style.display !== 'none';
                section.style.display = isVisible ? 'none' : 'block';
                
                // Scroll suave en mvil
                if (!isVisible && window.innerWidth <= 768) {
                    setTimeout(() => {
                        section.scrollIntoView({ behavior: 'smooth', block: 'start' });
                    }, 100);
                }
            }
        }
        
        function toggleSection(sectionId) {
            const section = document.getElementById(sectionId);
            const button = event.target;
            
            if (section) {
                const isVisible = section.style.display !== 'none';
                section.style.display = isVisible ? 'none' : 'block';
                button.textContent = isVisible ? 
                    button.textContent.replace('', '') : 
                    button.textContent.replace('', '');
                
                // Scroll suave en mvil
                if (!isVisible && window.innerWidth <= 768) {
                    setTimeout(() => {
                        section.scrollIntoView({ behavior: 'smooth', block: 'start' });
                    }, 100);
                }
            }
        }
        
        // Toggle prompt generation section
        function togglePromptSection() {
            const content = document.getElementById('promptSectionContent');
            const icon = document.getElementById('promptToggleIcon');
            const badge = document.getElementById('promptStatusBadge');
            
            if (content.style.maxHeight === '0px' || content.style.maxHeight === '') {
                // Expand
                content.style.maxHeight = '1000px'; // Large enough to fit content
                icon.style.transform = 'rotate(180deg)';
                icon.textContent = '';
                badge.textContent = 'ABIERTO';
                badge.style.background = 'rgba(76, 175, 80, 0.3)';
            } else {
                // Collapse
                content.style.maxHeight = '0px';
                icon.style.transform = 'rotate(0deg)';
                icon.textContent = '';
                badge.textContent = 'CONFIGURAR';
                badge.style.background = 'rgba(255,255,255,0.2)';
            }
        }
        
        // Optimizar seleccin de APIs para mvil
        function selectAIProvider(provider) {
            const buttons = document.querySelectorAll('.api-option input[type="radio"]');
            buttons.forEach(btn => {
                if (btn.value === provider) {
                    btn.checked = true;
                    btn.dispatchEvent(new Event('change'));
                }
            });
        }
        
        // Manejar orientacin en mvil
        window.addEventListener('orientationchange', function() {
            setTimeout(() => {
                // Reajustar layout despus del cambio de orientacin
                window.scrollTo(0, 0);
            }, 100);
        });
        
        // Optimizar rendimiento en mvil
        function debounce(func, wait) {
            let timeout;
            return function executedFunction(...args) {
                const later = () => {
                    clearTimeout(timeout);
                    func(...args);
                };
                clearTimeout(timeout);
                timeout = setTimeout(later, wait);
            };
        }
        
        // Lazy loading para imgenes en mvil
        if ('IntersectionObserver' in window) {
            const imageObserver = new IntersectionObserver((entries, observer) => {
                entries.forEach(entry => {
                    if (entry.isIntersecting) {
                        const img = entry.target;
                        img.src = img.dataset.src;
                        img.classList.remove('lazy');
                        imageObserver.unobserve(img);
                    }
                });
            });
            
            document.querySelectorAll('img[data-src]').forEach(img => {
                imageObserver.observe(img);
            });
        }
    </script>
</head>
<body>
    <!-- HEADER DE USUARIO -->
    {% if user_info %}
    <div class="user-header">
        <div class="user-info">
            <div class="user-avatar">
                {{ user_info.first_name[0] if user_info.first_name else 'U' }}
            </div>
            <div class="user-details">
                <h3>{{ user_info.first_name }} {{ user_info.last_name }}</h3>
                <p>{{ user_info.email }}  {{ user_info.user_type.title() }}</p>
            </div>
        </div>
        <div class="user-actions">
            <span class="account-badge {{ 'badge-premium' if user_info.account_type == 'premium' else 'badge-free' }}">
                {{ ' Premium' if user_info.account_type == 'premium' else ' Gratuito' }}
            </span>
            <a href="/logout" class="btn-logout">
                <i class="fas fa-sign-out-alt"></i> Cerrar sesi贸n
            </a>
        </div>
    </div>
    {% endif %}
    
    <!--  TESTING CONTROLS - Only shown when auth is disabled -->
    {% if user_info %}
    <div class="testing-controls" style="background: rgba(255, 165, 0, 0.1); border: 2px dashed #ffa500; padding: 15px; margin: 20px auto; max-width: 1200px; border-radius: 10px; text-align: center;">
    <h4 style="color: #ff8c00; margin-bottom: 10px;"> MODO TESTING - Cambiar tipo de usuario</h4>
        <div style="display: flex; gap: 15px; justify-content: center; flex-wrap: wrap;">
            <a href="/switch-user/premium" class="btn" style="background: linear-gradient(135deg, #3483fa, #2968c8); color: white; text-decoration: none; padding: 10px 20px; border-radius: 8px; font-weight: bold;">
                 Modo PREMIUM
            </a>
            <a href="/switch-user/regular" class="btn" style="background: linear-gradient(135deg, #6c757d, #495057); color: white; text-decoration: none; padding: 10px 20px; border-radius: 8px; font-weight: bold;">
                 Modo GRATUITO
            </a>
        </div>
                <p style="color: #666; font-size: 0.9rem; margin-top: 10px; font-style: italic;">
            Estos controles solo est谩n disponibles durante las pruebas (DISABLE_AUTH = True)
        </p>
    </div>
    {% endif %}
    
    <!-- ALERTA PARA USUARIOS GRATUITOS -->
    {% if user_info and user_info.account_type == 'free' %}
    <div class="premium-alert">
        <i class="fas fa-info-circle"></i>
        <div>
            <strong>Cuenta gratuita:</strong> Tienes acceso al modo manual. 
            <strong>Actualiza a Premium</strong> para usar IA autom谩tica y generaci贸n desde prompt.
        </div>
    </div>
    {% endif %}
    
    <div class="container">
        <div class="header">
            <h1 class="ml-logo">ML Bulk Mapper Pro <span class="ai-badge">AI POWERED</span></h1>
            <p style="color: #666; margin: 0;">Herramienta profesional para carga masiva en Mercado Libre</p>
            
            <!-- BOTONES DE AYUDA Y DONACIN -->
            <div style="display: flex; justify-content: center; gap: 16px; margin-top: 20px;">
                <button type="button" onclick="toggleHelpSection()" style="
                    background: linear-gradient(135deg, #17a2b8 0%, #138496 100%);
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 25px;
                    cursor: pointer;
                    font-size: 14px;
                    font-weight: 600;
                    transition: all 0.3s ease;
                    box-shadow: 0 3px 10px rgba(23, 162, 184, 0.3);
                " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(23, 162, 184, 0.4)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(23, 162, 184, 0.3)'">
                     Gu铆a de Uso
                </button>
                
                <button type="button" onclick="toggleDonationSection()" style="
                    background: linear-gradient(135deg, #e91e63 0%, #c2185b 100%);
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 25px;
                    cursor: pointer;
                    font-size: 14px;
                    font-weight: 600;
                    transition: all 0.3s ease;
                    box-shadow: 0 3px 10px rgba(233, 30, 99, 0.3);
                " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(233, 30, 99, 0.4)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(233, 30, 99, 0.3)'">
                     Donar para Mascotas
                </button>
            </div>
        </div>
        
        <!--  SECCIN GUA DE USO COLAPSABLE -->
        <div class="help-section-collapsible">
            <div id="helpSectionContent" class="help-section-content" style="
                max-height: 0;
                overflow: hidden;
                transition: max-height 0.4s ease;
                background: linear-gradient(135deg, #f0fdff, #e0f7ff);
                border: 2px solid #17a2b8;
                border-radius: 12px;
                margin-bottom: 24px;
            ">
                <div style="padding: 24px;">
                    <div style="text-align: center; margin-bottom: 20px;">
                        <h3 style="color: #17a2b8; margin: 0; font-size: 20px;"> Gu铆a completa de uso</h3>
                        <small style="color: #666;">Todo lo que necesitas saber para usar ML Bulk Mapper Pro</small>
                    </div>
                    
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 20px;">
                        <!-- Funcionalidades Principales -->
                        <div style="background: rgba(23, 162, 184, 0.1); padding: 16px; border-radius: 8px; border-left: 4px solid #17a2b8;">
                            <h4 style="color: #17a2b8; margin-top: 0;"> Funcionalidades Principales</h4>
                            <ul style="margin: 0; padding-left: 20px; color: #333;">
                                <li><strong>Detecci贸n inteligente:</strong> Lee autom谩ticamente plantillas ML</li>
                                <li><strong>Mapeo Anti-Errores:</strong> Cada campo va exactamente donde debe</li>
                                <li><strong>IA avanzada:</strong> Completa datos faltantes autom谩ticamente</li>
                                <li><strong>Configuraci贸n manual:</strong> Valores masivos para toda la tienda</li>
                                <li><strong>C贸digos EAN-13:</strong> Generaci贸n autom谩tica de c贸digos</li>
                            </ul>
                        </div>
                        
                        <!-- C贸mo Evitar Errores -->
                        <div style="background: rgba(220, 53, 69, 0.1); padding: 16px; border-radius: 8px; border-left: 4px solid #dc3545;">
                            <h4 style="color: #dc3545; margin-top: 0;"> C贸mo evitar errores</h4>
                            <ul style="margin: 0; padding-left: 20px; color: #333;">
                                <li><strong>Plantilla ML:</strong> Descarga la plantilla oficial de tu categor铆a</li>
                                <li><strong>Datos limpios:</strong> Aseg煤rate de que los precios sean solo n煤meros</li>
                                <li><strong>SKU 煤nicos:</strong> Cada producto debe tener SKU diferente</li>
                                <li><strong>Campos obligatorios:</strong> T铆tulo, Precio, Stock son requeridos</li>
                                <li><strong>Formato Excel:</strong> Usa .xlsx para mejores resultados</li>
                            </ul>
                        </div>
                    </div>
                    
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 20px;">
                        <!-- Garant铆a -->
                        <div class="form-group">
                            <label> Garant铆a:</label>
                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                                <div>
                                    <label for="tipo_garantia">Tipo de garant铆a:</label>
                                    <select name="tipo_garantia" id="tipo_garantia" style="width: 100%; max-width: 100%;">
                                        <option value="">Sin configurar</option>
                                        <option value="Garantia del vendedor">Garant铆a del vendedor</option>
                                        <option value="Garantia de fabrica">Garant铆a de f谩brica</option>
                                        <option value="Sin garantia">Sin garant铆a</option>
                                    </select>
                                </div>
                                <div>
                                    <label for="tiempo_garantia">Tiempo garant铆a:</label>
                                    <input type="number" name="tiempo_garantia" id="tiempo_garantia" placeholder="ej: 12" min="0">
                                    <label for="unidad_garantia">Unidad de tiempo:</label>
                                    <select name="unidad_garantia" id="unidad_garantia">
                                        <option value="dias">D铆as</option>
                                        <option value="meses">Meses</option>
                                        <option value="anios">A帽os</option>
                                    </select>
                                </div>
                            </div>
                        </div>
                    
                    <div style="background: rgba(255, 193, 7, 0.1); padding: 16px; border-radius: 8px; border-left: 4px solid #ffc107; margin-bottom: 20px;">
                            <h4 style="color: #e67e22; margin-top: 0;"> Campos obligatorios de ML</h4>
                        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; color: #333;">
                            <div><strong> T铆tulo:</strong> Nombre del producto</div>
                            <div><strong> Precio:</strong> Solo n煤meros (sin $ ni s铆mbolos)</div>
                            <div><strong> Stock:</strong> Cantidad disponible</div>
                            <div><strong> Marca:</strong> Fabricante del producto</div>
                            <div><strong> Modelo:</strong> Versi贸n espec铆fica</div>
                            <div><strong> Env铆o:</strong> Configuraci贸n autom谩tica</div>
                        </div>
                    </div>
                    
                    <!-- Botn de cierre -->
                    <div style="text-align: center; padding: 16px 0 8px 0; border-top: 1px solid rgba(23, 162, 184, 0.2); margin-top: 20px;">
                        <button type="button" onclick="toggleHelpSection()" style="
                            background: linear-gradient(135deg, #17a2b8 0%, #138496 100%);
                            color: white;
                            border: none;
                            padding: 10px 24px;
                            border-radius: 25px;
                            cursor: pointer;
                            font-size: 14px;
                            font-weight: 600;
                            transition: all 0.3s ease;
                            box-shadow: 0 3px 10px rgba(23, 162, 184, 0.3);
                        " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(23, 162, 184, 0.4)'" 
                           onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(23, 162, 184, 0.3)'">
                             Cerrar Gu铆a de Uso
                        </button>
                    </div>
                </div>
            </div>
        </div>
        
        <!--  SECCIN DONACIN PARA MASCOTAS COLAPSABLE -->
        <div class="donation-section-collapsible">
            <div id="donationSectionContent" class="donation-section-content" style="
                max-height: 0;
                overflow: hidden;
                transition: max-height 0.4s ease;
                background: linear-gradient(135deg, #fef0f5, #fce4ec);
                border: 2px solid #e91e63;
                border-radius: 12px;
                margin-bottom: 24px;
            ">
                <div style="padding: 24px;">
                    <div style="text-align: center; margin-bottom: 20px;">
                        <h3 style="color: #e91e63; margin: 0; font-size: 20px;"> Apoya Mascotas de Montevideo</h3>
                        <small style="color: #666;">Tu donacin salva vidas de perros y gatos de la calle</small>
                    </div>
                    
                    <div style="background: rgba(233, 30, 99, 0.1); padding: 16px; border-radius: 8px; border-left: 4px solid #e91e63; margin-bottom: 20px; text-align: center;">
                        <p style="margin: 0; font-size: 16px; color: #333; line-height: 1.6;">
                             <strong>Tu ayuda hace la diferencia!</strong> <br>
                            Cada donacin que hagas apoya directamente a <strong>fundaciones reales de Montevideo</strong> que rescatan, curan y buscan hogares para mascotas abandonadas.
                        </p>
                    </div>
                    
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                        <!-- Fundacin 1 -->
                        <div style="background: white; padding: 16px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border: 1px solid #e91e63;">
                            <h4 style="color: #e91e63; margin: 0 0 8px 0;"> Fundacin Peludo Feliz</h4>
                            <p style="margin: 0 0 12px 0; font-size: 14px; color: #555;">
                                Rescate y rehabilitacin de perros y gatos en situacin de calle. Ms de 500 mascotas salvadas.
                            </p>
                            <div style="display: flex; gap: 8px;">
                                <a href="https://instagram.com/peludo_feliz_uy" target="_blank" style="
                                    background: #e91e63; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Instagram</a>
                                <a href="https://facebook.com/peludofelizuy" target="_blank" style="
                                    background: #3b5998; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Facebook</a>
                            </div>
                        </div>
                        
                        <!-- Fundacin 2 -->
                        <div style="background: white; padding: 16px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border: 1px solid #e91e63;">
                            <h4 style="color: #e91e63; margin: 0 0 8px 0;"> Refugio Cuatro Patas</h4>
                            <p style="margin: 0 0 12px 0; font-size: 14px; color: #555;">
                                Hogar temporal y adopcin responsable. Atencin veterinaria gratuita para mascotas rescatadas.
                            </p>
                            <div style="display: flex; gap: 8px;">
                                <a href="https://instagram.com/cuatro_patas_uy" target="_blank" style="
                                    background: #e91e63; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Instagram</a>
                                <a href="https://facebook.com/cuatropatasuy" target="_blank" style="
                                    background: #3b5998; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Facebook</a>
                            </div>
                        </div>
                    </div>
                    
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                        <!-- Fundacin 3 -->
                        <div style="background: white; padding: 16px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border: 1px solid #e91e63;">
                            <h4 style="color: #e91e63; margin: 0 0 8px 0;"> Rescate Amor Animal</h4>
                            <p style="margin: 0 0 12px 0; font-size: 14px; color: #555;">
                                Enfoque en casos de emergencia veterinaria. Cirugas y tratamientos para mascotas heridas.
                            </p>
                            <div style="display: flex; gap: 8px;">
                                <a href="https://instagram.com/amor_animal_uy" target="_blank" style="
                                    background: #e91e63; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Instagram</a>
                                <a href="https://facebook.com/amoranimaluy" target="_blank" style="
                                    background: #3b5998; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Facebook</a>
                            </div>
                        </div>
                        
                        <!-- Fundacin 4 -->
                        <div style="background: white; padding: 16px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border: 1px solid #e91e63;">
                            <h4 style="color: #e91e63; margin: 0 0 8px 0;"> Hogar Esperanza</h4>
                            <p style="margin: 0 0 12px 0; font-size: 14px; color: #555;">
                                Especialistas en gatos ferales y socializacin. Programas de castracin masiva.
                            </p>
                            <div style="display: flex; gap: 8px;">
                                <a href="https://instagram.com/hogar_esperanza_uy" target="_blank" style="
                                    background: #e91e63; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Instagram</a>
                                <a href="https://facebook.com/hogaresperanzauy" target="_blank" style="
                                    background: #3b5998; color: white; padding: 6px 12px; border-radius: 15px; 
                                    text-decoration: none; font-size: 12px; font-weight: 600;
                                "> Facebook</a>
                            </div>
                        </div>
                    </div>
                    
                    <div style="background: rgba(255, 193, 7, 0.1); padding: 16px; border-radius: 8px; border-left: 4px solid #ffc107; margin-bottom: 20px; text-align: center;">
                        <p style="margin: 0; font-size: 14px; color: #333; line-height: 1.6;">
                             <strong>驴C贸mo ayudar?</strong><br>
                            Visita las redes sociales de estas fundaciones, comparte sus publicaciones, haz donaciones directas o adopta una mascota. 
                            <strong>Cada accin cuenta!</strong> 
                        </p>
                    </div>
                    
                    <!-- Botn de cierre -->
                    <div style="text-align: center; padding: 16px 0 8px 0; border-top: 1px solid rgba(233, 30, 99, 0.2); margin-top: 20px;">
                        <button type="button" onclick="toggleDonationSection()" style="
                            background: linear-gradient(135deg, #e91e63 0%, #c2185b 100%);
                            color: white;
                            border: none;
                            padding: 10px 24px;
                            border-radius: 25px;
                            cursor: pointer;
                            font-size: 14px;
                            font-weight: 600;
                            transition: all 0.3s ease;
                            box-shadow: 0 3px 10px rgba(233, 30, 99, 0.3);
                        " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(233, 30, 99, 0.4)'" 
                           onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(233, 30, 99, 0.3)'">
                             Cerrar Seccin de Donaciones
                        </button>
                    </div>
                </div>
            </div>
        </div>
        
        {% if message %}
            <div class="{{ message_type }}">{{ message }}</div>
        {% endif %}
        
        {% if debug_info %}
            <div class="debug-section">
                <div class="debug-title">DEBUG LOG:</div>
                <pre>{{ debug_info }}</pre>
            </div>
        {% endif %}
        
        <!--  SECCIN GENERACIN POR PROMPT COLAPSABLE -->
        <div class="prompt-generation-collapsible">
            <div class="prompt-generation-header" onclick="togglePromptSection()" style="
                background: linear-gradient(135deg, #3483fa 0%, #2968c8 100%);
                color: white;
                padding: 16px 24px;
                border-radius: 12px;
                cursor: pointer;
                display: flex;
                align-items: center;
                justify-content: space-between;
                margin: 24px 0 0 0;
                transition: all 0.3s ease;
                box-shadow: 0 4px 15px rgba(52, 131, 250, 0.2);
            " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 6px 20px rgba(52, 131, 250, 0.3)'" 
               onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 4px 15px rgba(52, 131, 250, 0.2)'">
                <div style="display: flex; align-items: center; gap: 12px;">
                    <span style="font-size: 24px;"></span>
                    <div>
                        <h3 style="margin: 0; font-size: 18px; font-weight: 600;">Generaci贸n por prompt (IA)</h3>
                        <small style="opacity: 0.9; font-size: 14px;">Crear plantillas desde descripci贸n de texto</small>
                    </div>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <span id="promptStatusBadge" style="
                        background: rgba(255,255,255,0.2);
                        padding: 4px 12px;
                        border-radius: 20px;
                        font-size: 12px;
                        font-weight: 600;
                    ">CONFIGURAR</span>
                    <span id="promptToggleIcon" style="
                        font-size: 20px;
                        transition: transform 0.3s ease;
                    "></span>
                </div>
            </div>
            
            <div id="promptSectionContent" class="prompt-section-content" style="
                max-height: 0;
                overflow: hidden;
                transition: max-height 0.4s ease;
                background: white;
                border: 2px solid #3483fa;
                border-top: none;
                border-radius: 0 0 12px 12px;
                margin-bottom: 24px;
            ">
                <div style="padding: 24px;">
                        <div style="
                        background: linear-gradient(135deg, #e3f2fd, #e1f5fe);
                        padding: 16px;
                        border-radius: 8px;
                        margin-bottom: 20px;
                        border-left: 4px solid #3483fa;
                        color: #1565c0;
                    ">
                        <strong>Generaci贸n inteligente:</strong> Describe tus productos y la IA crear谩 una plantilla ML completa con datos realistas y precios competitivos.
                    </div>
                    
                    <form method="post" action="/generate-from-prompt" enctype="multipart/form-data">
                        <div class="form-group">
                            <label for="product_prompt" style="font-weight: 600; color: #333; margin-bottom: 8px; display: block;">
                                Descripci贸n de productos:
                            </label>
                            <textarea 
                                name="product_prompt" 
                                id="product_prompt" 
                                placeholder="Ejemplo: 10 iPhones 14 Pro de 128GB en colores negro, blanco y dorado. 5 Samsung Galaxy S24 Ultra de 256GB. Todos nuevos con precios competitivos para Uruguay."
                                style="
                                    width: 100%;
                                    min-height: 100px;
                                    padding: 14px 16px;
                                    border: 2px solid #e6e6e6;
                                    border-radius: 8px;
                                    background: #fafafa;
                                    color: #333;
                                    font-size: 16px;
                                    resize: vertical;
                                    font-family: inherit;
                                    transition: all 0.3s ease;
                                "
                                onfocus="this.style.borderColor='#3483fa'; this.style.background='#fff';"
                                onblur="this.style.borderColor='#e6e6e6'; this.style.background='#fafafa';"
                                required
                            ></textarea>
                            <small style="color: #666; display: block; margin-top: 6px;">
                                Incluye cantidad, modelos, colores, capacidades y detalles espec铆ficos
                            </small>
                        </div>
                        
                        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin: 20px 0;">
                            <div class="form-group">
                                <label style="font-weight: 600; color: #333;">Cantidad estimada:</label>
                                <input 
                                    type="number" 
                                    name="total_products" 
                                    min="1" 
                                    max="100" 
                                    value="10"
                                    style="
                                        width: 100%;
                                        padding: 14px 16px;
                                        border: 2px solid #e6e6e6;
                                        border-radius: 8px;
                                        background: #fafafa;
                                        color: #333;
                                        font-size: 16px;
                                        transition: all 0.3s ease;
                                    "
                                    onfocus="this.style.borderColor='#3483fa'; this.style.background='#fff';"
                                    onblur="this.style.borderColor='#e6e6e6'; this.style.background='#fafafa';"
                                >
                            </div>
                            
                            <div class="form-group">
                                <label style="font-weight: 600; color: #333;">Categora principal:</label>
                                <select 
                                    name="main_category"
                                    style="
                                        width: 100%;
                                        padding: 14px 16px;
                                        border: 2px solid #e6e6e6;
                                        border-radius: 8px;
                                        background: #fafafa;
                                        color: #333;
                                        font-size: 16px;
                                        transition: all 0.3s ease;
                                        -webkit-appearance: none;
                                        appearance: none;
                                    "
                                    onfocus="this.style.borderColor='#3483fa'; this.style.background='#fff';"
                                    onblur="this.style.borderColor='#e6e6e6'; this.style.background='#fafafa';"
                                >
                                    <option value="electronica">Electrnicos</option>
                                    <option value="ropa">Ropa y Accesorios</option>
                                    <option value="hogar">Hogar y Jardn</option>
                                    <option value="deportes">Deportes y Fitness</option>
                                    <option value="belleza">Belleza y Cuidado</option>
                                    <option value="juguetes">Juguetes y Juegos</option>
                                    <option value="libros">Libros y Msica</option>
                                    <option value="autos">Autos y Motos</option>
                                    <option value="otros">Otros</option>
                                </select>
                            </div>
                            
                            <div class="form-group">
                                <label style="font-weight: 600; color: #333;">Estilo de venta:</label>
                                <select 
                                    name="generation_style"
                                    style="
                                        width: 100%;
                                        padding: 14px 16px;
                                        border: 2px solid #e6e6e6;
                                        border-radius: 8px;
                                        background: #fafafa;
                                        color: #333;
                                        font-size: 16px;
                                        transition: all 0.3s ease;
                                        -webkit-appearance: none;
                                        appearance: none;
                                    "
                                    onfocus="this.style.borderColor='#3483fa'; this.style.background='#fff';"
                                    onblur="this.style.borderColor='#e6e6e6'; this.style.background='#fafafa';"
                                >
                                    <option value="professional">Profesional</option>
                                    <option value="casual">Casual</option>
                                    <option value="premium">Premium</option>
                                    <option value="wholesale">Mayorista</option>
                                </select>
                            </div>
                        </div>
                        
                        <!-- Configuraci贸n de IA - Solo Premium -->
                        {% if not user_info or user_info.account_type == 'premium' %}
                        <div style="
                            background: #f8f9fa;
                            padding: 20px;
                            border-radius: 8px;
                            margin: 20px 0;
                            border: 1px solid #e9ecef;
                        ">
                            <h4 style="color: #333; margin: 0 0 16px 0; font-size: 16px; font-weight: 600;">Configuraci贸n de IA</h4>
                            
                            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px;">
                                <div class="form-group">
                                    <label style="font-weight: 600; color: #333;">Proveedor IA:</label>
                                    <select 
                                        name="ai_provider_prompt"
                                        style="
                                            width: 100%;
                                            padding: 12px 16px;
                                            border: 2px solid #e6e6e6;
                                            border-radius: 6px;
                                            background: white;
                                            color: #333;
                                            font-size: 14px;
                                        "
                                    >
                                        <option value="groq">Groq (Gratis)</option>
                                        <option value="deepseek">DeepSeek (Econmico)</option>
                                    </select>
                                </div>
                                
                                <div class="form-group">
                                    <label style="font-weight: 600; color: #333;">API Key:</label>
                                    <input 
                                        type="password" 
                                        name="ai_api_key_prompt" 
                                        placeholder="Ingresa tu API key"
                                        style="
                                            width: 100%;
                                            padding: 12px 16px;
                                            border: 2px solid #e6e6e6;
                                            border-radius: 6px;
                                            background: white;
                                            color: #333;
                                            font-size: 14px;
                                        "
                                        required
                                    >
                                </div>
                            </div>
                        </div>
                        {% else %}
                        <!-- Mensaje para usuarios gratuitos -->
                        <div style="
                            background: linear-gradient(135deg, #fff3cd, #ffeaa7);
                            border: 1px solid #ffeaa7;
                            color: #856404;
                            padding: 20px;
                            border-radius: 8px;
                            margin: 20px 0;
                            text-align: center;
                        ">
                            <h4 style="color: #856404; margin: 0 0 10px 0;"> Funcionalidad Premium</h4>
                            <p style="margin: 0;">La configuraci贸n de IA requiere cuenta Premium. Actualiza tu cuenta para acceder a esta funcionalidad.</p>
                        </div>
                        {% endif %}
                        
                        <div style="text-align: center; margin-top: 24px;">
                            <button type="submit" style="
                                background: linear-gradient(135deg, #3483fa, #2968c8);
                                color: white;
                                border: none;
                                padding: 16px 32px;
                                border-radius: 8px;
                                font-size: 16px;
                                font-weight: 600;
                                cursor: pointer;
                                box-shadow: 0 4px 15px rgba(52, 131, 250, 0.3);
                                transition: all 0.3s ease;
                                min-width: 200px;
                            " 
                            onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 6px 20px rgba(52, 131, 250, 0.4)'" 
                            onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 4px 15px rgba(52, 131, 250, 0.3)'">
                                Generar Plantilla
                            </button>
                        </div>
                        
                        <div style="text-align: center; margin-top: 12px;">
                            <small style="color: #666;">
                                La IA crear una plantilla ML completa con productos realistas
                            </small>
                        </div>
                    </form>
                    
                    <!-- Botn de cierre -->
                    <div style="
                        text-align: center;
                        padding: 16px 0 8px 0;
                        border-top: 1px solid rgba(52, 131, 250, 0.2);
                        margin-top: 20px;
                    ">
                        <button type="button" onclick="togglePromptSection()" style="
                            background: linear-gradient(135deg, #3483fa, #2968c8);
                            color: white;
                            border: none;
                            padding: 10px 24px;
                            border-radius: 25px;
                            cursor: pointer;
                            font-size: 14px;
                            font-weight: 600;
                            transition: all 0.3s ease;
                            box-shadow: 0 3px 10px rgba(52, 131, 250, 0.3);
                        " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(52, 131, 250, 0.4)'" 
                           onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(52, 131, 250, 0.3)'">
                            Cerrar generaci贸n por prompt
                        </button>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- SEPARADOR VISUAL -->
        <div style="
            display: flex;
            align-items: center;
            margin: 32px 0;
            text-align: center;
        ">
            <div style="flex: 1; height: 1px; background: linear-gradient(to right, transparent, #e6e6e6, transparent);"></div>
            <div style="
                padding: 0 20px;
                background: white;
                color: #666;
                font-weight: 600;
                font-size: 14px;
            ">
                O PROCESA UN ARCHIVO EXISTENTE
            </div>
            <div style="flex: 1; height: 1px; background: linear-gradient(to left, transparent, #e6e6e6, transparent);"></div>
        </div>
        
        <form method="post" enctype="multipart/form-data">
            
            <!-- SECCIN IA COLAPSABLE - Solo Premium -->
            {% if not user_info or user_info.account_type == 'premium' %}
            <div class="ai-section-collapsible">
                <div class="ai-section-header" onclick="toggleAISection()" style="
                    background: linear-gradient(135deg, #66b3ff 0%, #4da6ff 100%);
                    color: white;
                    padding: 16px 24px;
                    border-radius: 12px;
                    cursor: pointer;
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    margin: 24px 0 0 0;
                    transition: all 0.3s ease;
                    box-shadow: 0 4px 15px rgba(52, 131, 250, 0.2);
                " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 6px 20px rgba(52, 131, 250, 0.3)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 4px 15px rgba(52, 131, 250, 0.2)'">
                    <div style="display: flex; align-items: center; gap: 12px;">
                        <span style="font-size: 24px;"></span>
                        <div>
                            <h3 style="margin: 0; font-size: 18px; font-weight: 600;">Configuraci贸n de Inteligencia Artificial</h3>
                            <small class="small-pill">Haz clic para configurar la IA (Opcional)</small>
                        </div>
                    </div>
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <span id="aiStatusBadge" style="
                            background: rgba(255,255,255,0.2);
                            padding: 4px 12px;
                            border-radius: 20px;
                            font-size: 12px;
                            font-weight: 600;
                        ">CONFIGURAR</span>
                        <span id="aiToggleIcon" style="
                            font-size: 20px;
                            transition: transform 0.3s ease;
                        "></span>
                    </div>
                </div>
                
                <div id="aiSectionContent" class="ai-section-content" style="
                    max-height: 0;
                    overflow: hidden;
                    transition: max-height 0.4s ease;
                    background: linear-gradient(135deg, #f8fffe, #f0f9ff);
                    border: 2px solid #66b3ff;
                    border-top: none;
                    border-radius: 0 0 12px 12px;
                    margin-bottom: 24px;
                ">
                    <div style="padding: 24px;">
                        <div class="ai-info" style="
                            background: rgba(102, 179, 255, 0.1);
                            padding: 16px;
                            border-radius: 8px;
                            margin-bottom: 20px;
                            border-left: 4px solid #66b3ff;
                        ">
                            <strong> La IA ayuda a completar autom谩ticamente:</strong> c贸digos universales, marcas, modelos, descripciones y caracter铆sticas t茅cnicas que falten en tus datos.
                        </div>
                        
                        <div class="form-group">
                            <label>Selecciona tu proveedor de IA:</label>
                            <div class="api-selector">
                                <div class="api-option">
                                    <input type="radio" name="ai_provider" value="groq" id="groq" checked onchange="updateAIStatus()">
                                    <label for="groq">
                                        <strong>Groq</strong> <span class="cost-badge">GRATIS</span>
                                        <span class="quality-stars"></span><br>
                                        <small>R谩pido y gratuito. Recomendado para empezar.</small>
                                    </label>
                                </div>
                                
                                <div class="api-option">
                                    <input type="radio" name="ai_provider" value="deepseek" id="deepseek" onchange="updateAIStatus()">
                                    <label for="deepseek">
                                        <strong>DeepSeek</strong> <span class="cost-badge">$0.14/1M</span>
                                        <span class="quality-stars"></span><br>
                                        <small>Muy econmico y excelente calidad.</small>
                                    </label>
                                </div>
                                
                                <div class="api-option">
                                    <input type="radio" name="ai_provider" value="manual" id="manual" onchange="updateAIStatus()">
                                    <label for="manual">
                                        <strong>Modo Manual</strong> <span class="cost-badge">SIN IA</span><br>
                                        <small>Valores por defecto sin usar IA.</small>
                                    </label>
                                </div>
                            </div>
                        </div>
                        
                        <div class="form-group" id="api_key_section">
                            <label for="ai_api_key">API Key (requerida para Groq/DeepSeek):</label>
                            <input type="password" name="ai_api_key" id="ai_api_key" placeholder="Ingresa tu API key aqu" onchange="updateAIStatus()">
                            <small>Para Groq: Regstrate gratis en <a href="https://groq.com" target="_blank" style="color: #66b3ff;">groq.com</a></small>
                        </div>
                        
                        <div class="form-group">
                            <label>Campos que la IA debe completar autom谩ticamente:</label>
                            <div class="checkbox-group">
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="codigo_universal" id="ai_codigo" checked>
                                    <label for="ai_codigo"><span class="ai-enhanced">C贸digo Universal EAN-13</span></label>
                                </div>
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="marca" id="ai_marca" checked>
                                    <label for="ai_marca"><span class="ai-enhanced">Marca</span> (si no est en datos)</label>
                                </div>
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="modelo" id="ai_modelo" checked>
                                    <label for="ai_modelo"><span class="ai-enhanced">Modelo</span> (si no est en datos)</label>
                                </div>
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="descripcion" id="ai_desc">
                                    <label for="ai_desc"><span class="ai-enhanced">Descripci贸n atractiva</span></label>
                                </div>
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="peso" id="ai_peso">
                                    <label for="ai_peso"><span class="ai-enhanced">Peso estimado</span></label>
                                </div>
                                <div class="checkbox-item">
                                    <input type="checkbox" name="ai_fields" value="color" id="ai_color">
                                    <label for="ai_color"><span class="ai-enhanced">Color inferido</span></label>
                                </div>
                            </div>
                        </div>
                        
                        <!-- Botn de cierre en la parte inferior -->
                        <div style="
                            text-align: center;
                            padding: 16px 0 8px 0;
                            border-top: 1px solid rgba(102, 179, 255, 0.2);
                            margin-top: 20px;
                        ">
                            <button type="button" onclick="toggleAISection()" style="
                                background: linear-gradient(135deg, #66b3ff 0%, #4da6ff 100%);
                                color: white;
                                border: none;
                                padding: 10px 24px;
                                border-radius: 25px;
                                cursor: pointer;
                                font-size: 14px;
                                font-weight: 600;
                                transition: all 0.3s ease;
                                box-shadow: 0 3px 10px rgba(102, 179, 255, 0.3);
                            " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(102, 179, 255, 0.4)'" 
                               onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(102, 179, 255, 0.3)'">
                                 Cerrar Configuraci贸n IA
                            </button>
                        </div>
                    </div>
                </div>
            </div>
            
            <h3>Archivos</h3>
            
            <div class="form-group">
                <label for="template">Plantilla de Mercado Libre (.xlsx):</label>
                <input type="file" name="template" accept=".xlsx" required>
                <small>Sube la plantilla oficial descargada de Mercado Libre</small>
            </div>
            
            <div class="form-group">
                <label for="content">Datos de Productos:</label>
                <input type="file" name="content" accept=".xlsx,.xls,.csv,.pdf,.docx,.txt" required>
                <small>Recomendado: Excel o CSV para mejor mapeo</small>
            </div>
            
            <h3>Configuraci贸n de Mapeo</h3>
            
            <div class="mapping-section">
                <label>Selecciona los campos que quieres mapear desde tus datos:</label>
                
                <div class="checkbox-group">
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="titulo" id="titulo" checked>
                        <label for="titulo"><span class="required">T铆tulo</span> (Nombre del producto)</label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="precio" id="precio" checked>
                        <label for="precio"><span class="required">Precio</span></label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="stock" id="stock" checked>
                        <label for="stock"><span class="required">Stock</span></label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="sku" id="sku" checked>
                        <label for="sku"><span class="optional">SKU</span></label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="descripcion" id="descripcion">
                        <label for="descripcion"><span class="optional">Descripci贸n</span> (existente en datos)</label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="marca" id="marca_existing">
                        <label for="marca_existing"><span class="required">Marca</span> (existente en datos)</label>
                    </div>
                    
                    <div class="checkbox-item">
                        <input type="checkbox" name="map_fields" value="modelo" id="modelo_existing">
                        <label for="modelo_existing"><span class="required">Modelo</span> (existente en datos)</label>
                    </div>
                </div>
            </div>
            
            <h3>Valores por Defecto</h3>
            
            <div class="form-group">
                <label for="condicion">Condicin del producto:</label>
                <select name="condicion">
                    <option value="Nuevo" selected>Nuevo</option>
                    <option value="Usado">Usado</option>
                    <option value="Reacondicionado">Reacondicionado</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="moneda">Moneda:</label>
                <select name="moneda">
                    <option value="$" selected>$ (Peso Uruguayo)</option>
                    <option value="USD">USD (Dlar)</option>
                </select>
            </div>
            
            {% else %}
            <!-- Mensaje para usuarios gratuitos sobre IA -->
            <div style="
                background: linear-gradient(135deg, #fff3cd, #ffeaa7);
                border: 1px solid #ffeaa7;
                color: #856404;
                padding: 24px;
                border-radius: 12px;
                margin: 24px 0;
                text-align: center;
            ">
                <h3 style="color: #856404; margin: 0 0 15px 0;"> Inteligencia Artificial - Premium</h3>
                <p style="margin: 0 0 15px 0; font-size: 16px;">
                    La IA automtica para autocompletar productos requiere cuenta Premium.
                </p>
                <div style="
                    background: rgba(255,255,255,0.7);
                    padding: 15px;
                    border-radius: 8px;
                    margin-top: 15px;
                ">
                    <strong>Con Premium obtienes:</strong><br>
                     Autocompletado inteligente de t铆tulos<br>
                     Detecci贸n autom谩tica de marcas y modelos<br>
                     Optimizaci贸n de precios con IA<br>
                     Generaci贸n de descripciones atractivas<br>
                     C贸digos EAN-13 autom谩ticos
                </div>
            </div>
            {% endif %}
            
            <!-- NUEVA SECCIN: CONFIGURACIN MANUAL MASIVA -->
            {% if user_info and user_info.account_type == 'premium' %}
            <div class="manual-config-collapsible">
                <div class="manual-config-header" onclick="toggleManualSection()" style="
                    background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                    color: white;
                    padding: 16px 24px;
                    border-radius: 12px;
                    cursor: pointer;
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    margin: 24px 0 0 0;
                    transition: all 0.3s ease;
                    box-shadow: 0 4px 15px rgba(40, 167, 69, 0.3);
                " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 6px 20px rgba(40, 167, 69, 0.4)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 4px 15px rgba(40, 167, 69, 0.3)'">
                    <div style="display: flex; align-items: center; gap: 12px;">
                        <span style="font-size: 24px;"></span>
                        <div>
                            <h3 style="margin: 0; font-size: 18px; font-weight: 600;">Configuraci贸n Manual Masiva</h3>
                            <small style="opacity: 0.9; font-size: 14px;">Haz clic para configurar valores masivos (Opcional)</small>
                        </div>
                    </div>
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <span id="manualStatusBadge" style="
                            background: rgba(255,255,255,0.2);
                            padding: 4px 12px;
                            border-radius: 20px;
                            font-size: 12px;
                            font-weight: 600;
                        ">CONFIGURAR</span>
                        <span id="manualToggleIcon" style="
                            font-size: 20px;
                            transition: transform 0.3s ease;
                        "></span>
                    </div>
                </div>
                
                <div id="manualSectionContent" class="manual-section-content" style="
                    max-height: 0;
                    overflow: hidden;
                    transition: max-height 0.4s ease;
                    background: linear-gradient(135deg, #f8fff8, #f0fff4);
                    border: 2px solid #28a745;
                    border-top: none;
                    border-radius: 0 0 12px 12px;
                    margin-bottom: 24px;
                ">
                    <div style="padding: 24px;">
                        <div class="info-text" style="
                            background: rgba(40, 167, 69, 0.1);
                            padding: 16px;
                            border-radius: 8px;
                            margin-bottom: 20px;
                            border-left: 4px solid #28a745;
                        ">
                            <strong> Configura valores que se aplicarn a todos los productos o selectivamente por nmero de fila.</strong><br>
                             <em>Perfecto para tiendas que manejan valores estndar en todos sus productos.</em>
                        </div>
                
                <!-- Stock Masivo -->
                <div class="form-group">
                    <label> Stock:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="stock_global">Stock para todos los productos:</label>
                            <input type="number" name="stock_global" id="stock_global" placeholder="ej: 100" min="0" style="width: 100%; max-width: 100%;">
                            <small>Aplicar a todos los productos detectados</small>
                        </div>
                        <div>
                            <label for="stock_selective">Stock selectivo (Fila#:Cantidad):</label>
                            <textarea name="stock_selective" id="stock_selective" rows="3" placeholder="ej: 8:50, 10:70, 12:25" style="width: 100%; max-width: 100%;"></textarea>
                            <small>Formato: Fila_Excel:Stock (separados por comas)</small>
                        </div>
                    </div>
                </div>
                
                <!-- Marca y Modelo -->
                <div class="form-group">
                    <label> Marca y Modelo:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="marca_global">Marca para todos:</label>
                            <input type="text" name="marca_global" id="marca_global" placeholder="ej: Samsung, Apple, Sony" style="width: 100%; max-width: 100%;">
                            <small>Se aplicar a todos los productos</small>
                        </div>
                        <div>
                            <label for="modelo_global">Modelo para todos:</label>
                            <input type="text" name="modelo_global" id="modelo_global" placeholder="ej: Pro Max, Standard" style="width: 100%; max-width: 100%;">
                            <small>Se aplicar a todos los productos</small>
                        </div>
                    </div>
                    <div>
                        <label for="marca_modelo_selective">Excepciones selectivas (Fila#:Marca:Modelo):</label>
                        <textarea name="marca_modelo_selective" id="marca_modelo_selective" rows="2" placeholder="ej: 8:iPhone:14 Pro, 10:Samsung:Galaxy S24" style="width: 100%; max-width: 100%;"></textarea>
                        <small>Formato: Fila_Excel:Marca:Modelo (separados por comas)</small>
                    </div>
                </div>
                
                <!-- Retiro en Persona -->
                <div class="form-group">
                    <label> Retiro en Persona:</label>
                    <div style="display: flex; gap: 20px; align-items: center;">
                        <label><input type="radio" name="retiro_persona" value="Acepto">  Acepto (todos los productos)</label>
                        <label><input type="radio" name="retiro_persona" value="No acepto">  No acepto (todos los productos)</label>
                        <label><input type="radio" name="retiro_persona" value="" checked> Sin configurar</label>
                    </div>
                </div>
                
                <!-- NUEVA SECCIN: FORMAS DE ENVO -->
                <div class="form-group">
                    <label> Formas de env铆o:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="forma_envio_global">Forma de env铆o para todos:</label>
                            <select name="forma_envio_global" id="forma_envio_global" style="width: 100%; max-width: 100%;">
                                <option value="">Sin configurar</option>
                                <option value="Mercado Envios"> Mercado Env铆os</option>
                                <option value="Mercado Envios + Mercado Envios Flex"> Mercado Env铆os + Mercado Env铆os Flex</option>
                                <option value="Acordar con el vendedor"> Acordar con el vendedor</option>
                            </select>
                            <small>Aplica a todos los productos</small>
                        </div>
                        <div>
                            <label for="forma_envio_selective">Excepciones selectivas:</label>
                            <textarea name="forma_envio_selective" id="forma_envio_selective" rows="2" placeholder="ej: 8:Mercado Env铆os, 10:Mercado Env铆os + Mercado Env铆os Flex" style="width: 100%; max-width: 100%;"></textarea>
                            <small>Formato: Fila_Excel:Forma_Env铆o</small>
                        </div>
                    </div>
                </div>
                
                <!-- NUEVA SECCIN: COSTO DE ENVO -->
                <div class="form-group">
                    <label> Costo de env铆o:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="costo_envio_global">Costo para todos los productos:</label>
                            <select name="costo_envio_global" id="costo_envio_global" style="width: 100%; max-width: 100%;">
                                <option value="">Sin configurar</option>
                                <option value="A cargo del comprador"> A cargo del comprador</option>
                                <option value="Envio gratis"> Env铆o gratis</option>
                            </select>
                            <small>Pol铆tica de costo de env铆o general</small>
                        </div>
                        <div>
                            <label for="costo_envio_selective">Excepciones por fila:</label>
                            <textarea name="costo_envio_selective" id="costo_envio_selective" rows="2" placeholder="ej: 8:Env铆o gratis, 10:A cargo del comprador" style="width: 100%; max-width: 100%;"></textarea>
                            <small>Formato: Fila_Excel:Tipo_Costo</small>
                        </div>
                    </div>
                </div>
                
                <!-- NUEVA SECCIN: VARIACIONES POR COLOR -->
                <div class="form-group">
                    <label> Variaciones por Color:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="color_global">Color base para todos:</label>
                            <select name="color_global" id="color_global" style="width: 100%; max-width: 100%;">
                                <option value="">Sin configurar</option>
                                <option value="Negro"> Negro</option>
                                <option value="Blanco"> Blanco</option>
                                <option value="Azul"> Azul</option>
                                <option value="Rojo"> Rojo</option>
                                <option value="Verde"> Verde</option>
                                <option value="Amarillo"> Amarillo</option>
                                <option value="Rosa"> Rosa</option>
                                <option value="Gris"> Gris</option>
                                <option value="Marron"> Marr贸n</option>
                                <option value="Multicolor"> Multicolor</option>
                                <option value="Transparente"> Transparente</option>
                            </select>
                        </div>
                        <div>
                            <label for="color_comercial_global">Nombre comercial del color:</label>
                            <input type="text" name="color_comercial_global" id="color_comercial_global" placeholder="ej: Azul Marino, Rojo Cereza" style="width: 100%; max-width: 100%;">
                            <small>Nombre espec铆fico o comercial del color</small>
                        </div>
                    </div>
                    <div>
                        <label for="color_selective">Colores selectivos (Fila#:Color:Nombre_Comercial):</label>
                        <textarea name="color_selective" id="color_selective" rows="2" placeholder="ej: 8:Azul:Azul Marino, 10:Rojo:Rojo Cereza" style="width: 100%; max-width: 100%;"></textarea>
                        <small>Formato: Fila_Excel:Color_Base:Nombre_Comercial (separados por comas)</small>
                    </div>
                </div>
                
                <!-- Garant铆a -->
                <div class="form-group">
                    <label> Garant铆a:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px;">
                        <div>
                            <label for="tipo_garantia">Tipo de garant铆a:</label>
                            <select name="tipo_garantia" id="tipo_garantia">
                                <option value="">Sin configurar</option>
                                <option value="Garantia del vendedor">Garant铆a del vendedor</option>
                                <option value="Garantia de fabrica">Garant铆a de f谩brica</option>
                                <option value="Sin garantia">Sin garant铆a</option>
                            </select>
                        </div>
                        <div>
                            <label for="tiempo_garantia">Tiempo de garant铆a:</label>
                            <input type="number" name="tiempo_garantia" id="tiempo_garantia" placeholder="ej: 12" min="0">
                        </div>
                        <div>
                            <label for="unidad_garantia">Unidad de tiempo:</label>
                            <select name="unidad_garantia" id="unidad_garantia">
                                <option value="">Seleccionar</option>
                                <option value="das">Das</option>
                                <option value="meses" selected>Meses</option>
                                <option value="aos">Aos</option>
                            </select>
                        </div>
                    </div>
                </div>
                
                <!-- C贸digos Universales -->
                <div class="form-group">
                    <label> C贸digos Universales:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 24px; align-items: start;">
                        <div style="padding-right: 8px;">
                            <label for="codigo_universal_masivo">C贸digo base para todos:</label>
                            <input type="text" name="codigo_universal_masivo" id="codigo_universal_masivo" placeholder="ej: PROD, SKU, UPC">
                            <small>Se aplicar谩 como base a todos los productos</small>
                        </div>
                        <div style="padding-left: 8px;">
                            <label for="codigo_universal_secuencial" style="display: block; margin-bottom: 8px;">Numeraci贸n autom谩tica:</label>
                            <label style="display: flex; align-items: center; gap: 8px; margin-bottom: 8px;">
                                <input type="checkbox" name="codigo_universal_secuencial" id="codigo_universal_secuencial">
                                Agregar numeraci贸n secuencial
                            </label>
                            <input type="number" name="codigo_universal_offset" id="codigo_universal_offset" placeholder="Empezar en (ej: 1001)" min="1">
                            <small>Ej: PROD0001, PROD0002, PROD0003...</small>
                        </div>
                    </div>
                    <div style="margin-top: 12px;">
                        <label for="codigo_universal_selective">C贸digos espec铆ficos (Fila#:C贸digo):</label>
                        <textarea name="codigo_universal_selective" id="codigo_universal_selective" rows="2" placeholder="ej: 8:UPC123456, 10:EAN987654, 12:CODE555"></textarea>
                        <small>Formato: Fila_Excel:C贸digo_Especifico (separados por comas)</small>
                    </div>
                </div>
                
                <!-- Catlogo -->
                <div class="form-group">
                    <label> Catlogo:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="tiene_catalogo">Los productos tienen catlogo?</label>
                            <select name="tiene_catalogo" id="tiene_catalogo" style="width: 100%; max-width: 100%;">
                                <option value="">Sin configurar</option>
                                <option value="Si">S铆, tienen cat谩logo</option>
                                <option value="No">No tienen catlogo</option>
                            </select>
                        </div>
                        <div>
                            <label for="numero_catalogo_selective">N煤meros de cat谩logo selectivos:</label>
                            <textarea name="numero_catalogo_selective" id="numero_catalogo_selective" rows="2" placeholder="ej: 8:CAT001, 10:CAT002, 12:CAT003" style="width: 100%; max-width: 100%;"></textarea>
                            <small>Formato: Fila_Excel:C贸digo_Cat谩logo</small>
                        </div>
                    </div>
                </div>
                
                <!-- Descripci贸n global -->
                <div class="form-group">
                    <label> Descripci贸n global de la tienda:</label>
                    <textarea name="descripcion_global" id="descripcion_global" rows="4" placeholder="ej: Horario: Lunes a Viernes 9-18h | Ubicaci贸n: Centro de Montevideo | WhatsApp: 099123456 | Env铆os a todo el pa铆s"></textarea>
                    <small>Esta informaci贸n se agregar谩 al final de cada descripci贸n de producto. Usa emojis para mejor visualizaci贸n.</small>
                </div>
                
                        <!-- Botn de cierre en la parte inferior -->
                        <div style="
                            text-align: center;
                            padding: 16px 0 8px 0;
                            border-top: 1px solid rgba(40, 167, 69, 0.2);
                            margin-top: 20px;
                        ">
                            <button type="button" onclick="toggleManualSection()" style="
                                background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                                color: white;
                                border: none;
                                padding: 10px 24px;
                                border-radius: 25px;
                                cursor: pointer;
                                font-size: 14px;
                                font-weight: 600;
                                transition: all 0.3s ease;
                                box-shadow: 0 3px 10px rgba(40, 167, 69, 0.3);
                            " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(40, 167, 69, 0.4)'" 
                               onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(40, 167, 69, 0.3)'">
                                 Cerrar Configuraci贸n Manual
                            </button>
                        </div>
                    </div>
                </div>
            </div>
            
            {% else %}
            <!-- Mensaje para usuarios gratuitos sobre configuraci贸n manual -->
            <div style="
                background: linear-gradient(135deg, #f8f9fa, #e9ecef);
                border: 1px solid #dee2e6;
                color: #6c757d;
                padding: 24px;
                border-radius: 12px;
                margin: 24px 0;
                text-align: center;
            ">
                <h3 style="color: #6c757d; margin: 0 0 15px 0;"> Configuraci贸n Manual Masiva - Premium</h3>
                <p style="margin: 0 0 15px 0; font-size: 16px;">
                    Las opciones de configuraci贸n masiva estn disponibles solo para cuentas Premium.
                </p>
                <div style="
                    background: rgba(255,255,255,0.8);
                    padding: 15px;
                    border-radius: 8px;
                    margin-top: 15px;
                ">
                    <strong>Con Premium puedes configurar:</strong><br>
                     Categoras masivas automticas<br>
                     Precios y comisiones globales<br>
                     Env铆os configurados en lote<br>
                     Garant铆as estandarizadas<br>
                     C贸digos universales autom谩ticos
                </div>
            </div>
            {% endif %}
            
            <!-- NUEVA SECCIN: CONFIGURACIN AVANZADA DE IA -->
            {% if user_info and user_info.account_type == 'premium' %}
            <div class="advanced-ai-collapsible">
                <div class="advanced-ai-header" onclick="toggleAdvancedAISection()" style="
                    background: linear-gradient(135deg, #6f42c1 0%, #8e44ad 100%);
                    color: white;
                    padding: 16px 24px;
                    border-radius: 12px;
                    cursor: pointer;
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    margin: 24px 0 0 0;
                    transition: all 0.3s ease;
                    box-shadow: 0 4px 15px rgba(111, 66, 193, 0.3);
                " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 6px 20px rgba(111, 66, 193, 0.4)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 4px 15px rgba(111, 66, 193, 0.3)'">
                    <div style="display: flex; align-items: center; gap: 12px;">
                        <span style="font-size: 24px;"></span>
                        <div>
                            <h3 style="margin: 0; font-size: 18px; font-weight: 600;">Configuraci贸n Avanzada de IA</h3>
                            <small class="small-pill">Haz clic para personalizar prompts de IA (Opcional)</small>
                        </div>
                    </div>
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <span id="advancedAIStatusBadge" style="
                            background: rgba(255,255,255,0.2);
                            padding: 4px 12px;
                            border-radius: 20px;
                            font-size: 12px;
                            font-weight: 600;
                        ">CONFIGURAR</span>
                        <span id="advancedAIToggleIcon" style="
                            font-size: 20px;
                            transition: transform 0.3s ease;
                        "></span>
                    </div>
                </div>
                
                <div id="advancedAISectionContent" class="advanced-ai-section-content" style="
                    max-height: 0;
                    overflow: hidden;
                    transition: max-height 0.4s ease;
                    background: linear-gradient(135deg, #faf8ff, #f4f0ff);
                    border: 2px solid #6f42c1;
                    border-top: none;
                    border-radius: 0 0 12px 12px;
                    margin-bottom: 24px;
                ">
                    <div style="padding: 24px;">
                        <div class="info-text" style="
                            background: rgba(111, 66, 193, 0.1);
                            padding: 16px;
                            border-radius: 8px;
                            margin-bottom: 20px;
                            border-left: 4px solid #6f42c1;
                        ">
                            <strong> Personaliza c贸mo la IA procesar谩 y mejorar谩 tus productos autom谩ticamente.</strong><br>
                             <em>La IA se enfocar谩 primero en campos OBLIGATORIOS y luego en mejoras adicionales.</em>
                        </div>
                
                <!-- Prompts Personalizados -->
                <div class="form-group">
                    <label> Mejora de T铆tulos (SEO):</label>
                    <textarea name="ai_titulo_prompt" id="ai_titulo_prompt" rows="3" placeholder="ej: Mejora este t铆tulo para que sea m谩s atractivo y tenga mejor SEO. Incluye palabras clave relevantes sin hacer spam. Mant茅n el estilo profesional y claro."></textarea>
                    <small>Instrucciones para que la IA mejore los t铆tulos con SEO optimizado</small>
                </div>
                
                <div class="form-group">
                    <label> Mejora de Descripciones:</label>
                    <textarea name="ai_descripcion_prompt" id="ai_descripcion_prompt" rows="4" placeholder="ej: Crea una descripci贸n detallada y persuasiva. Incluye beneficios del producto, caracter铆sticas t茅cnicas importantes y llamadas a la acci贸n. Usa un tono profesional pero cercano."></textarea>
                    <small>Instrucciones para generar descripciones m谩s completas y atractivas</small>
                </div>
                
                <div class="form-group">
                    <label> Investigaci贸n de Productos:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label>
                                <input type="checkbox" name="ai_auto_research" id="ai_auto_research">
                                Activar investigaci贸n autom谩tica
                            </label>
                            <small>La IA investigar谩 autom谩ticamente especificaciones t茅cnicas</small>
                        </div>
                        <div>
                            <label for="ai_research_prompt">Instrucciones de investigaci贸n:</label>
                            <textarea name="ai_research_prompt" id="ai_research_prompt" rows="3" placeholder="ej: Busca informacin tcnica como peso, dimensiones, material, compatibilidad. Agrega solo datos verificables y relevantes para la categora del producto." style="width: 100%; max-width: 100%;"></textarea>
                        </div>
                    </div>
                </div>
                
                <!-- Campos Obligatorios IA -->
                <div class="form-group">
                    <label> Enfoque en Campos Obligatorios:</label>
                    <div class="checkbox-grid" style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px;">
                        <label><input type="checkbox" name="ai_enforce_titulo" checked> T铆tulo optimizado</label>
                        <label><input type="checkbox" name="ai_enforce_precio" checked> Validacin de precios</label>
                        <label><input type="checkbox" name="ai_enforce_stock" checked> Stock realista</label>
                        <label><input type="checkbox" name="ai_enforce_marca" checked> Marca detectada</label>
                        <label><input type="checkbox" name="ai_enforce_modelo" checked> Modelo identificado</label>
                        <label><input type="checkbox" name="ai_enforce_envio" checked> Opciones de env铆o</label>
                    </div>
                    <small>La IA se asegurar谩 de que estos campos obligatorios est茅n correctamente completados</small>
                </div>
                
                <!-- Mejoras Opcionales -->
                <div class="form-group">
                    <label> Mejoras Adicionales Opcionales:</label>
                    <div class="checkbox-grid" style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px;">
                        <label><input type="checkbox" name="ai_add_peso"> Peso del producto</label>
                        <label><input type="checkbox" name="ai_add_dimensiones"> Dimensiones</label>
                        <label><input type="checkbox" name="ai_add_material"> Material</label>
                        <label><input type="checkbox" name="ai_add_compatibilidad"> Compatibilidad</label>
                        <label><input type="checkbox" name="ai_add_garantia"> Informaci贸n de garant铆a</label>
                        <label><input type="checkbox" name="ai_add_usos"> Usos recomendados</label>
                    </div>
                    <small>Campos adicionales que la IA puede completar si encuentra informacin relevante</small>
                </div>
                
                <!-- Estilo de Comunicacin -->
                <div class="form-group">
                    <label> Estilo de Comunicaci贸n:</label>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 12px;">
                        <div>
                            <label for="ai_tone">Tono de voz:</label>
                            <select name="ai_tone" id="ai_tone" style="width: 100%; max-width: 100%;">
                                <option value="profesional"> Profesional</option>
                                <option value="amigable" selected> Amigable</option>
                                <option value="tecnico"> Tcnico</option>
                                <option value="persuasivo"> Persuasivo</option>
                                <option value="minimalista"> Minimalista</option>
                            </select>
                        </div>
                        <div>
                            <label for="ai_target">Pblico objetivo:</label>
                            <select name="ai_target" id="ai_target" style="width: 100%; max-width: 100%;">
                                <option value="general" selected> Pblico general</option>
                                <option value="profesional"> Profesionales</option>
                                <option value="hogar"> Uso dom茅stico</option>
                                <option value="tecnico"> Usuarios tcnicos</option>
                                <option value="empresarial"> Empresarial</option>
                            </select>
                        </div>
                    </div>
                </div>
                
                <!-- Gestin de Fotos - Coming Soon -->
                <div class="form-group" style="background: linear-gradient(135deg, #f5f5f5, #e8e8e8); padding: 16px; border-radius: 6px; border: 2px dashed #ccc;">
                    <label> Gestin Automtica de Fotos:</label>
                    <div style="text-align: center; color: #666; padding: 20px;">
                        <h4 style="color: #999;"> COMING SOON </h4>
                        <p>Pr贸ximamente podr谩s conectar tu cuenta de ML para obtener autom谩ticamente las URLs de las fotos de tus productos.</p>
                        <small>Esta funcionalidad estar disponible en futuras actualizaciones.</small>
                    </div>
                </div>
                
                        <!-- Botn de cierre en la parte inferior -->
                        <div style="
                            text-align: center;
                            padding: 16px 0 8px 0;
                            border-top: 1px solid rgba(111, 66, 193, 0.2);
                            margin-top: 20px;
                        ">
                            <button type="button" onclick="toggleAdvancedAISection()" style="
                                background: linear-gradient(135deg, #6f42c1 0%, #8e44ad 100%);
                                color: white;
                                border: none;
                                padding: 10px 24px;
                                border-radius: 25px;
                                cursor: pointer;
                                font-size: 14px;
                                font-weight: 600;
                                transition: all 0.3s ease;
                                box-shadow: 0 3px 10px rgba(111, 66, 193, 0.3);
                            " onmouseover="this.style.transform='translateY(-1px)'; this.style.boxShadow='0 5px 15px rgba(111, 66, 193, 0.4)'" 
                               onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 3px 10px rgba(111, 66, 193, 0.3)'">
                                 Cerrar Configuraci贸n Avanzada
                            </button>
                        </div>
                    </div>
                </div>
            </div>
            
            {% else %}
            <!-- Mensaje para usuarios gratuitos sobre configuraci贸n avanzada de IA -->
            <div style="
                background: linear-gradient(135deg, #f3e7ff, #e8d5ff);
                border: 1px solid #d4b3ff;
                color: #6b21a8;
                padding: 24px;
                border-radius: 12px;
                margin: 24px 0;
                text-align: center;
            ">
                <h3 style="color: #6b21a8; margin: 0 0 15px 0;"> Configuraci贸n Avanzada de IA - Premium</h3>
                <p style="margin: 0 0 15px 0; font-size: 16px;">
                    La personalizacin avanzada de prompts de IA est disponible solo para cuentas Premium.
                </p>
                <div style="
                    background: rgba(255,255,255,0.8);
                    padding: 15px;
                    border-radius: 8px;
                    margin-top: 15px;
                ">
                    <strong>Con Premium puedes personalizar:</strong><br>
                     Prompts de mejora de t铆tulos SEO<br>
                     Instrucciones de descripci贸n autom谩tica<br>
                     Investigaci贸n autom谩tica de productos<br>
                     Estilo de comunicacin personalizado<br>
                     Enfoque en campos espec铆ficos
                </div>
            </div>
            {% endif %}
            
            <!-- SUBMIT BUTTON - CENTERED AND PROFESSIONAL -->
            {% if user_info and user_info.account_type == 'premium' %}
            <div style="text-align: center; margin: 40px 0; padding: 20px;">
                <button type="submit" style="
                    background: linear-gradient(135deg, #3483fa 0%, #1e6acc 100%);
                    color: white;
                    border: none;
                    padding: 18px 40px;
                    font-size: 18px;
                    font-weight: 600;
                    border-radius: 12px;
                    cursor: pointer;
                    box-shadow: 0 6px 20px rgba(52, 131, 250, 0.3);
                    transition: all 0.3s ease;
                    text-transform: uppercase;
                    letter-spacing: 1px;
                    min-width: 300px;
                " onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 8px 25px rgba(52, 131, 250, 0.4)'" 
                   onmouseout="this.style.transform='translateY(0px)'; this.style.boxShadow='0 6px 20px rgba(52, 131, 250, 0.3)'">
                     Procesar y Generar Archivo ML
                </button>
                <div style="margin-top: 12px; color: #666; font-size: 14px;">
                     El procesamiento puede tomar entre 10-30 segundos
                </div>
            </div>
            
            {% else %}
            <!-- Botn deshabilitado para usuarios gratuitos -->
            <div style="text-align: center; margin: 40px 0; padding: 20px;">
                <div style="
                    background: linear-gradient(135deg, #ffc107, #ffb300);
                    border: 2px solid #ff8f00;
                    color: #663c00;
                    padding: 24px;
                    border-radius: 12px;
                    margin-bottom: 20px;
                ">
                    <h3 style="margin: 0 0 15px 0;"> Procesamiento Premium Requerido</h3>
                    <p style="margin: 0 0 15px 0; font-size: 16px;">
                        Para procesar archivos y usar las funciones de IA necesitas una cuenta Premium.
                    </p>
                    <a href="/register" style="
                        background: linear-gradient(135deg, #28a745, #20c997);
                        color: white;
                        text-decoration: none;
                        padding: 12px 30px;
                        border-radius: 8px;
                        font-weight: 600;
                        display: inline-block;
                        margin-top: 10px;
                    ">
                         Obtener Premium
                    </a>
                </div>
                
                <button type="button" disabled style="
                    background: #6c757d;
                    color: #dee2e6;
                    border: none;
                    padding: 18px 40px;
                    font-size: 18px;
                    font-weight: 600;
                    border-radius: 12px;
                    cursor: not-allowed;
                    text-transform: uppercase;
                    letter-spacing: 1px;
                    min-width: 300px;
                    opacity: 0.6;
                ">
                     Procesamiento Bloqueado
                </button>
                <div style="margin-top: 12px; color: #6c757d; font-size: 14px;">
                     Requiere cuenta Premium para procesar
                </div>
            </div>
            {% endif %}
        </form>
        
        <!--  PANTALLA DE CARGA CREATIVA CON VERSCULOS BBLICOS -->
        <div class="loading-overlay" id="loadingOverlay">
            <div class="loading-container">
                <div class="loading-logo">
                    <span class="ml-animation"></span> ML Bulk Mapper Pro
                </div>
                <div class="progress-container">
                    <div class="progress-bar" id="progressBar">0%</div>
                </div>
                <div class="loading-steps" id="loadingSteps">
                    <div class="loading-step" id="step1">
                        <span class="loading-step-icon"></span>
                        <span>Cargando archivos...</span>
                    </div>
                    <div class="loading-step" id="step2">
                        <span class="loading-step-icon"></span>
                        <span>Analizando estructura de datos...</span>
                    </div>
                    <div class="loading-step" id="step3">
                        <span class="loading-step-icon"></span>
                        <span>Procesando con IA...</span>
                    </div>
                    <div class="loading-step" id="step4">
                        <span class="loading-step-icon"></span>
                        <span>Aplicando configuraci贸n manual...</span>
                    </div>
                    <div class="loading-step" id="step5">
                        <span class="loading-step-icon"></span>
                        <span>Generando archivo ML...</span>
                    </div>
                    <div class="loading-step" id="step6">
                        <span class="loading-step-icon"></span>
                        <span>Completado!</span>
                    </div>
                </div>
                
                <!-- VERSCULOS BBLICOS CON EL NOMBRE DE JEHOV -->
                <div class="bible-verse-container" id="bibleVerseContainer" style="
                    margin: 30px 0;
                    padding: 20px;
                    background: linear-gradient(135deg, #f8f9ff, #e8ecff);
                    border-radius: 12px;
                    border-left: 4px solid #3483fa;
                    text-align: center;
                    max-width: 500px;
                ">
                    <div style="font-size: 16px; line-height: 1.6; color: #333; font-style: italic; margin-bottom: 8px;" id="verseText">
                        "Confa en Jehov con todo tu corazn, y no te apoyes en tu propia prudencia."
                    </div>
                    <div style="font-size: 14px; color: #666; font-weight: 600;" id="verseReference">
                        - Proverbios 3:5
                    </div>
                </div>
                
                <p style="color: #666; font-size: 14px; margin-top: 20px;">
                     Este proceso puede tomar entre 10-30 segundos dependiendo del tamao de tu archivo
                </p>
            </div>
        </div>
        
        <script>
        //  FUNCIONES PARA SECCIN IA COLAPSABLE
        let aiSectionExpanded = false;
        
        function toggleAISection() {
            const content = document.getElementById('aiSectionContent');
            const icon = document.getElementById('aiToggleIcon');
            
            if (aiSectionExpanded) {
                // Colapsar
                content.style.maxHeight = '0px';
                icon.style.transform = 'rotate(0deg)';
                icon.textContent = '';
                aiSectionExpanded = false;
            } else {
                // Expandir - Usar scrollHeight + padding extra para el botn
                content.style.maxHeight = (content.scrollHeight + 50) + 'px';
                icon.style.transform = 'rotate(180deg)';
                icon.textContent = '';
                aiSectionExpanded = true;
            }
        }
        
        function updateAIStatus() {
            const badge = document.getElementById('aiStatusBadge');
            const groqSelected = document.getElementById('groq').checked;
            const deepseekSelected = document.getElementById('deepseek').checked;
            const manualSelected = document.getElementById('manual').checked;
            const apiKey = document.getElementById('ai_api_key').value;
            
            if (manualSelected) {
                badge.textContent = 'MANUAL';
                badge.style.background = 'rgba(255,193,7,0.3)';
            } else if ((groqSelected || deepseekSelected) && apiKey.length > 10) {
                badge.textContent = 'CONFIGURADO ';
                badge.style.background = 'rgba(40,167,69,0.3)';
            } else if (groqSelected || deepseekSelected) {
                badge.textContent = 'API KEY FALTA';
                badge.style.background = 'rgba(220,53,69,0.3)';
            } else {
                badge.textContent = 'CONFIGURAR';
                badge.style.background = 'rgba(255,255,255,0.2)';
            }
        }
        
        //  FUNCIONES PARA SECCIN MANUAL COLAPSABLE
        let manualSectionExpanded = false;
        
        function toggleManualSection() {
            const content = document.getElementById('manualSectionContent');
            const icon = document.getElementById('manualToggleIcon');
            
            if (manualSectionExpanded) {
                // Colapsar
                content.style.maxHeight = '0px';
                icon.style.transform = 'rotate(0deg)';
                icon.textContent = '';
                manualSectionExpanded = false;
            } else {
                // Expandir - Usar scrollHeight + padding extra para el botn
                content.style.maxHeight = (content.scrollHeight + 50) + 'px';
                icon.style.transform = 'rotate(180deg)';
                icon.textContent = '';
                manualSectionExpanded = true;
            }
        }
        
        //  FUNCIONES PARA SECCIN IA AVANZADA COLAPSABLE
        let advancedAISectionExpanded = false;
        
        function toggleAdvancedAISection() {
            const content = document.getElementById('advancedAISectionContent');
            const icon = document.getElementById('advancedAIToggleIcon');
            
            if (advancedAISectionExpanded) {
                // Colapsar
                content.style.maxHeight = '0px';
                icon.style.transform = 'rotate(0deg)';
                icon.textContent = '';
                advancedAISectionExpanded = false;
            } else {
                // Expandir - Usar scrollHeight + padding extra para el botn
                content.style.maxHeight = (content.scrollHeight + 50) + 'px';
                icon.style.transform = 'rotate(180deg)';
                icon.textContent = '';
                advancedAISectionExpanded = true;
            }
        }
        
        //  FUNCIONES PARA SECCIN AYUDA COLAPSABLE
        let helpSectionExpanded = false;
        
        function toggleHelpSection() {
            const content = document.getElementById('helpSectionContent');
            
            if (helpSectionExpanded) {
                // Colapsar
                content.style.maxHeight = '0px';
                helpSectionExpanded = false;
            } else {
                // Expandir - Usar scrollHeight + padding extra
                content.style.maxHeight = (content.scrollHeight + 50) + 'px';
                helpSectionExpanded = true;
            }
        }
        
        //  FUNCIONES PARA SECCIN DONACIN COLAPSABLE
        let donationSectionExpanded = false;
        
        function toggleDonationSection() {
            const content = document.getElementById('donationSectionContent');
            
            if (donationSectionExpanded) {
                // Colapsar
                content.style.maxHeight = '0px';
                donationSectionExpanded = false;
            } else {
                // Expandir - Usar scrollHeight + padding extra
                content.style.maxHeight = (content.scrollHeight + 50) + 'px';
                donationSectionExpanded = true;
            }
        }
        
        // Inicializar estado al cargar la pgina
        document.addEventListener('DOMContentLoaded', function() {
            updateAIStatus();
        });
        
        //  VERSCULOS BBLICOS CON EL NOMBRE DE JEHOV
        const bibleVerses = [
            {
                text: "Confa en Jehov con todo tu corazn, y no te apoyes en tu propia prudencia.",
                reference: "Proverbios 3:5"
            },
            {
                text: "Jehov tu Dios est en medio de ti como poderoso salvador.",
                reference: "Sofonas 3:17"
            },
            {
                text: "Jehov es mi pastor; nada me faltar.",
                reference: "Salmo 23:1"
            },
            {
                text: "Cuando pases por las aguas, yo estar contigo, dice Jehov.",
                reference: "Isaas 43:2"
            },
            {
                text: "Jehov pelear por vosotros, y vosotros estaris tranquilos.",
                reference: "xodo 14:14"
            },
            {
                text: "Los ojos de Jehov contemplan toda la tierra para mostrar su poder a favor de los que tienen corazn perfecto para con l.",
                reference: "2 Crnicas 16:9"
            },
            {
                text: "Echad sobre Jehov vuestras cargas, y l os sustentar.",
                reference: "Salmo 55:22"
            },
            {
                text: "Jehov est cerca de todos los que le invocan, de todos los que le invocan de veras.",
                reference: "Salmo 145:18"
            },
            {
                text: "Porque yo soy Jehov tu Dios, quien te sostiene de tu mano derecha, y te dice: No temas, yo te ayudo.",
                reference: "Isaas 41:13"
            },
            {
                text: "As ha dicho Jehov: No temis ni os amedrentis; esforzaos y sed valientes.",
                reference: "Deuteronomio 31:6"
            }
        ];
        
        function getRandomVerse() {
            return bibleVerses[Math.floor(Math.random() * bibleVerses.length)];
        }
        
        function updateBibleVerse() {
            const verse = getRandomVerse();
            document.getElementById('verseText').textContent = `"${verse.text}"`;
            document.getElementById('verseReference').textContent = `- ${verse.reference}`;
        }
        
        //  SISTEMA DE CARGA CREATIVO
        document.querySelector('form').addEventListener('submit', function(e) {
            const templateFile = document.querySelector('input[name="template"]').files[0];
            const contentFile = document.querySelector('input[name="content"]').files[0];
            
            if (!templateFile || !contentFile) {
                e.preventDefault();
                alert('Por favor selecciona ambos archivos antes de continuar.');
                return;
            }
            
            // Mostrar pantalla de carga
            showLoadingScreen();
        });
        
        function showLoadingScreen() {
            const overlay = document.getElementById('loadingOverlay');
            const progressBar = document.getElementById('progressBar');
            const steps = ['step1', 'step2', 'step3', 'step4', 'step5', 'step6'];
            
            overlay.style.display = 'flex';
            
            // Mostrar primer versculo
            updateBibleVerse();
            
            // Cambiar versculos cada 4 segundos
            const verseInterval = setInterval(updateBibleVerse, 4000);
            
            let currentStep = 0;
            let progress = 0;
            
            const progressInterval = setInterval(() => {
                progress += Math.random() * 15 + 5; // Progreso variable
                
                if (progress > 100) progress = 100;
                
                progressBar.style.width = progress + '%';
                progressBar.textContent = Math.round(progress) + '%';
                
                // Activar pasos
                if (progress > 10 && currentStep < 1) {
                    activateStep(steps[currentStep]);
                    currentStep++;
                }
                if (progress > 25 && currentStep < 2) {
                    completeStep(steps[currentStep - 1]);
                    activateStep(steps[currentStep]);
                    currentStep++;
                }
                if (progress > 45 && currentStep < 3) {
                    completeStep(steps[currentStep - 1]);
                    activateStep(steps[currentStep]);
                    currentStep++;
                }
                if (progress > 65 && currentStep < 4) {
                    completeStep(steps[currentStep - 1]);
                    activateStep(steps[currentStep]);
                    currentStep++;
                }
                if (progress > 85 && currentStep < 5) {
                    completeStep(steps[currentStep - 1]);
                    activateStep(steps[currentStep]);
                    currentStep++;
                }
                if (progress >= 100) {
                    completeStep(steps[currentStep - 1]);
                    activateStep(steps[currentStep]);
                    clearInterval(progressInterval);
                    clearInterval(verseInterval); // Detener rotacin de versculos
                }
            }, 800);
        }
        
        function activateStep(stepId) {
            document.getElementById(stepId).classList.add('active');
        }
        
        function completeStep(stepId) {
            const step = document.getElementById(stepId);
            step.classList.remove('active');
            step.classList.add('completed');
            step.querySelector('.loading-step-icon').textContent = '';
        }
        </script>
        
        {% if output_file %}
            <div class="success">
                <strong>Archivo generado exitosamente!</strong><br>
                {{ ai_summary if ai_summary else '' }}<br>
                {{ manual_summary if manual_summary else '' }}<br>
                <a href="{{ url_for('download_file', filename=output_file) }}" style="color: #00a650; font-weight: 600;">Descargar archivo para Mercado Libre</a>
            </div>
        {% endif %}
        
        <div class="creator-signature">
            Desarrollado por <span class="creator-name">Joss Mateo</span><br>
            <small>Herramienta profesional para automatizacin de Mercado Libre</small>
        </div>
    </div>
    
    <!--  LOADING OVERLAY MVIL-OPTIMIZADO -->
    <div id="loadingOverlay" class="loading-overlay">
        <div class="loading-container">
            <div class="loading-logo"> ML Mapper Pro</div>
            <div style="color: #666; margin-bottom: 20px;">Procesando con IA...</div>
            
            <div class="progress-container">
                <div class="progress-bar">0%</div>
            </div>
            
            <div class="loading-steps">
                <div class="loading-step">
                    <span class="step-icon"></span>
                    <span class="step-text">Leyendo archivo...</span>
                </div>
                <div class="loading-step">
                    <span class="step-icon"></span>
                    <span class="step-text">Procesando con IA...</span>
                </div>
                <div class="loading-step">
                    <span class="step-icon"></span>
                    <span class="step-text">Generando template ML...</span>
                </div>
                <div class="loading-step">
                    <span class="step-icon"></span>
                    <span class="step-text">Completado!</span>
                </div>
            </div>
            
            <div style="color: #999; font-size: 12px; margin-top: 16px;">
                 Esto puede tomar unos momentos en dispositivos mviles
            </div>
        </div>
    </div>

</body>
</html>
'''

# Funcin para analizar plantilla ML y extraer estructura
def analyze_ml_template(file_path):
    """Analiza la plantilla de ML y extrae la estructura de categoras con deteccin inteligente"""
    #  FIX CRTICO: Usar data_only=True (compatible con todas las versiones)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except TypeError:
        # Fallback para versiones muy antiguas
        wb = openpyxl.load_workbook(file_path)
    
    # Encontrar hoja de categora (no Ayuda ni Legales)
    category_sheet = None
    
    #  BUSCAR HOJA CORRECTA: Evitar hojas de ayuda/legales
    priority_sheets = []
    for sheet_name in wb.sheetnames:
        sheet_name_lower = sheet_name.lower()
        if not any(skip in sheet_name_lower for skip in ['ayuda', 'help', 'legal', 'info', 'extra']):
            priority_sheets.append(sheet_name)
    
    if priority_sheets:
        category_sheet = wb[priority_sheets[0]]  # Usar la primera hoja vlida
    else:
        category_sheet = wb.active  # Fallback
    
    if not category_sheet:
        raise ValueError("No se encontr hoja de categora en la plantilla ML")

    #  EXTRAER HEADERS DE FILA 3 CON MAPEO INTELIGENTE
    headers = {}
    obligatory_fields = {}
    
    # Headers conocidos de ML para mapeo inteligente
    ml_header_mapping = {
    # keep common non-accented variants for backwards compatibility
    'cdigo de catlogo': 'catalogo',
    'codigo de catalogo': 'catalogo',
    'c贸digo de cat谩logo': 'catalogo',
    'ttulo': 'titulo',
    'titulo': 'titulo',
    't铆tulo': 'titulo',
    'condicin': 'condicion',
    'condicion': 'condicion',
    'cdigo universal': 'codigo_universal',
    'codigo universal': 'codigo_universal',
    'c贸digo universal': 'codigo_universal',
    'color': 'color',
    'nombre comercial': 'color_comercial',
    'fotos': 'fotos',
    'sku': 'sku',
    'stock': 'stock',
    'precio': 'precio',
    'moneda': 'moneda',
    'descripcin': 'descripcion',
    'descripcion': 'descripcion',
    'descripci贸n': 'descripcion',
    'cargo por venta': 'cargo_venta',
    'forma de envo': 'forma_envio',
    'forma de envio': 'forma_envio',
    'forma de env铆o': 'forma_envio',
    'costo de envo': 'costo_envio',
    'costo de envio': 'costo_envio',
    'costo de env铆o': 'costo_envio',
    'retiro en persona': 'retiro_persona',
    'tipo de garanta': 'garantia_tipo',
    'tipo de garantia': 'garantia_tipo',
    'tipo de garant铆a': 'garantia_tipo',
    'tiempo de garanta': 'garantia_tiempo',
    'tiempo de garantia': 'garantia_tiempo',
    'tiempo de garant铆a': 'garantia_tiempo',
    'unidad de tiempo': 'garantia_unidad',
    'unidad de tiempo de la garanta': 'garantia_unidad',
    'unidad de tiempo de la garant铆a': 'garantia_unidad',
    'tienda oficial': 'tienda_oficial',
    'marca': 'marca',
    'modelo': 'modelo',
    'formato de venta': 'formato_venta',
    'unidades por pack': 'unidades_pack',
        'dimetro': 'diametro',
        'unidad de dimetro': 'unidad_diametro',
        'largo': 'largo',
        'unidad de largo': 'unidad_largo',
        'material': 'material',
        'resistencia': 'resistencia',
        'unidad de resistencia': 'unidad_resistencia',
        'peso': 'peso',
        'unidad de peso': 'unidad_peso'
    }
    
    for col in range(1, category_sheet.max_column + 1):
        header = safe_get_cell_value(category_sheet, 3, col)  # FILA 3 es donde estn los headers
        obligatory = safe_get_cell_value(category_sheet, 4, col)  # FILA 4 indica obligatoriedad
        
        if header and len(str(header).strip()) > 0:
            headers[col] = str(header)
            # Verificar si es obligatorio
            obligatory_fields[col] = 'obligatorio' in str(obligatory).lower() if obligatory else False
            
            # Mapeo inteligente para identificar el tipo de campo
            header_lower = str(header).lower()
            for key_phrase, field_type in ml_header_mapping.items():
                if key_phrase in header_lower:
                    headers[f"{col}_type"] = field_type
                    break

    return category_sheet, headers, obligatory_fields
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() not in ['ayuda', 'legales', 'extra info']:
            category_sheet = wb[sheet_name]
            break
    
    if not category_sheet:
        raise ValueError("No se encontr hoja de categora en la plantilla ML")
    
    # Extraer headers (fila 3) con manejo seguro de MergedCell
    headers = {}
    for col in range(1, category_sheet.max_column + 1):
        header = safe_get_cell_value(category_sheet, 3, col)
        if header and len(str(header).strip()) > 0:
            headers[col] = str(header)
    
    return category_sheet, headers

# Funcin mejorada para leer datos de productos
def read_product_data(file_path, file_ext):
    """Lee datos de productos desde diferentes formatos con manejo de MergedCell"""
    if file_ext in ['xlsx', 'xls']:
        #  FIX CRTICO: Manejo de MergedCell y data_only=True para frmulas
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except TypeError:
            # Fallback para versiones antiguas de openpyxl
            wb = openpyxl.load_workbook(file_path)
        
        sheet = wb.active
        
        if not sheet:
            return []
        
        # Obtener headers con manejo seguro de MergedCell
        headers = []
        for col in range(1, (sheet.max_column or 0) + 1):
            header_value = safe_get_cell_value(sheet, 1, col)
            if header_value:
                headers.append(str(header_value).lower().strip())
            else:
                headers.append("")
        
        if not headers:
            return []
        
        # Obtener datos (todas las filas) con manejo seguro de MergedCell
        products = []
        for row in range(2, (sheet.max_row or 1) + 1):
            product = {}
            for col in range(1, len(headers) + 1):
                if col <= len(headers):
                    value = safe_get_cell_value(sheet, row, col)
                    
                    if value is not None:
                        #  Para precios, asegurar que sean nmeros vlidos
                        if 'precio' in headers[col-1]:
                            try:
                                # Convertir a float y formatear
                                if isinstance(value, str) and value.startswith('='):
                                    # Si todava hay frmula, evaluar manualmente o usar valor por defecto
                                    print(f"  Frmula detectada en precio: {value}")
                                    product[headers[col-1]] = "0"  # Valor por defecto
                                else:
                                    value = float(value)
                                    product[headers[col-1]] = str(value)
                            except (ValueError, TypeError):
                                product[headers[col-1]] = str(value)
                        else:
                            product[headers[col-1]] = str(value)
            if product:  # Solo agregar si tiene datos
                products.append(product)
        
        return products
    
    elif file_ext == 'csv':
        products = []
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Normalizar keys a lowercase
                product = {k.lower().strip(): v for k, v in row.items() if v}
                if product:
                    products.append(product)
        return products
    
    else:
        # Para PDF, DOCX, TXT - retornar datos bsicos
        return [{'content': f'Datos extrados de {file_path}'}]

@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    if request.method == 'POST':
        try:
            # Obtener archivos
            template_file = request.files.get('template')
            content_file = request.files.get('content')
            
            if not template_file or not content_file or not template_file.filename or not content_file.filename:
                return render_template_string(HTML_TEMPLATE, 
                                           message="Por favor selecciona ambos archivos", 
                                           message_type="error",
                                           user_info=session)
            
            # Obtener configuraci贸n IA
            ai_provider = request.form.get('ai_provider', 'manual')
            
            # Validar funcionalidades premium
            if ai_provider != 'manual':
                is_premium, premium_message = validate_premium_feature('IA para autocompletar datos')
                if not is_premium:
                    return render_template_string(HTML_TEMPLATE, 
                                               message=premium_message, 
                                               message_type="warning",
                                               user_info=session)
            ai_api_key = request.form.get('ai_api_key', '')
            ai_fields = request.form.getlist('ai_fields')
            
            # Configurar IA
            local_ai_enhancer = AIProductEnhancer(
                provider=ai_provider if ai_provider != 'manual' else 'groq',
                api_key=ai_api_key if ai_api_key else None
            )
            
            # Guardar archivos
            template_filename = secure_filename(template_file.filename)
            content_filename = secure_filename(content_file.filename)
            
            template_path = os.path.join(app.config['UPLOAD_FOLDER'], template_filename)
            content_path = os.path.join(app.config['UPLOAD_FOLDER'], content_filename)
            
            template_file.save(template_path)
            content_file.save(content_path)
            
            # Analizar plantilla ML
            category_sheet, ml_headers, obligatory_fields = analyze_ml_template(template_path)
            
            # Leer datos de productos
            content_ext = content_filename.split('.')[-1].lower()
            products = read_product_data(content_path, content_ext)
            
            # Obtener configuraci贸n de mapeo
            selected_fields = request.form.getlist('map_fields')
            condicion = request.form.get('condicion', 'Nuevo')
            moneda = request.form.get('moneda', '$')
            
            #  NUEVA CONFIGURACIN MANUAL MASIVA
            manual_config = {
                'stock_global': request.form.get('stock_global'),
                'stock_selective': request.form.get('stock_selective'),
                'marca_global': request.form.get('marca_global'),
                'modelo_global': request.form.get('modelo_global'),
                'marca_modelo_selective': request.form.get('marca_modelo_selective'),
                'retiro_persona': request.form.get('retiro_persona'),
                'tipo_garantia': request.form.get('tipo_garantia'),
                'tiempo_garantia': request.form.get('tiempo_garantia'),
                'unidad_garantia': request.form.get('unidad_garantia'),
                'tiene_catalogo': request.form.get('tiene_catalogo'),
                'numero_catalogo_selective': request.form.get('numero_catalogo_selective'),
                'descripcion_global': request.form.get('descripcion_global'),
                #  NUEVAS CONFIGURACIONES
                'forma_envio_global': request.form.get('forma_envio_global'),
                'forma_envio_selective': request.form.get('forma_envio_selective'),
                'costo_envio_global': request.form.get('costo_envio_global'),
                'costo_envio_selective': request.form.get('costo_envio_selective'),
                'color_global': request.form.get('color_global'),
                'color_comercial_global': request.form.get('color_comercial_global'),
                'color_selective': request.form.get('color_selective'),
                #  CDIGOS UNIVERSALES
                'codigo_universal_masivo': request.form.get('codigo_universal_masivo'),
                'codigo_universal_secuencial': request.form.get('codigo_universal_secuencial') == 'on',
                'codigo_universal_offset': int(request.form.get('codigo_universal_offset', 1)) if request.form.get('codigo_universal_offset') else 1,
                'codigo_universal_selective': request.form.get('codigo_universal_selective')
            }
            
            #  NUEVA CONFIGURACIN AVANZADA DE IA
            ai_config = {
                'titulo_prompt': request.form.get('ai_titulo_prompt', ''),
                'descripcion_prompt': request.form.get('ai_descripcion_prompt', ''),
                'research_prompt': request.form.get('ai_research_prompt', ''),
                'auto_research': request.form.get('ai_auto_research') == 'on',
                'enforce_titulo': request.form.get('ai_enforce_titulo') == 'on',
                'enforce_precio': request.form.get('ai_enforce_precio') == 'on',
                'enforce_stock': request.form.get('ai_enforce_stock') == 'on',
                'enforce_marca': request.form.get('ai_enforce_marca') == 'on',
                'enforce_modelo': request.form.get('ai_enforce_modelo') == 'on',
                'enforce_envio': request.form.get('ai_enforce_envio') == 'on',
                'add_peso': request.form.get('ai_add_peso') == 'on',
                'add_dimensiones': request.form.get('ai_add_dimensiones') == 'on',
                'add_material': request.form.get('ai_add_material') == 'on',
                'add_compatibilidad': request.form.get('ai_add_compatibilidad') == 'on',
                'add_garantia': request.form.get('ai_add_garantia') == 'on',
                'add_usos': request.form.get('ai_add_usos') == 'on',
                'tone': request.form.get('ai_tone', 'amigable'),
                'target': request.form.get('ai_target', 'general')
            }
            
            #  PROCESAR CON IA - Enriquecer datos de productos
            ai_summary = ""
            debug_info = []  #  NUEVO: Log de debug detallado
            
            debug_info.append(" INICIANDO PROCESAMIENTO...")
            debug_info.append(f" Total productos detectados: {len(products)}")
            debug_info.append(f" Campos a mapear: {selected_fields}")
            debug_info.append(f" IA habilitada: {'S' if ai_provider != 'manual' else 'No'}")
            debug_info.append(f" Configuraci贸n manual: {'Activa' if any(manual_config.values()) else 'No configurada'}")
            
            #  Debug precios ANTES del procesamiento IA
            precios_detectados_input = 0
            for i, product in enumerate(products):
                precio_value = find_product_value(product, 'precio')
                if precio_value:
                    precios_detectados_input += 1
                    debug_info.append(f" INPUT Precio fila {i+2}: {precio_value}")
            debug_info.append(f" Total precios detectados en INPUT: {precios_detectados_input}")
            
            if ai_provider != 'manual' and ai_fields:
                enhanced_products = []
                ai_stats = {'enhanced': 0, 'total': len(products)}
                
                for product in products:
                    # Determinar qu campos faltan
                    missing_fields = []
                    for field in ai_fields:
                        if not find_product_value(product, field):
                            missing_fields.append(field)
                    
                    if missing_fields:
                        # Generar datos faltantes con IA
                        try:
                            ai_data = local_ai_enhancer.generate_missing_data(product, missing_fields)
                            product.update(ai_data)
                            ai_stats['enhanced'] += 1
                            debug_info.append(f" IA mejor producto: {missing_fields}")
                        except Exception as e:
                            debug_info.append(f" Error IA para producto {product}: {e}")
                    
                    enhanced_products.append(product)
                
                products = enhanced_products
                ai_summary = f" IA proces {ai_stats['enhanced']}/{ai_stats['total']} productos, completando {len(ai_fields)} tipos de campos."
                debug_info.append(ai_summary)
            
            # Crear archivo de salida
            output_filename = f"ML_AI_Output_{template_filename}"
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            
            # Copiar plantilla original como base
            shutil.copy2(template_path, output_path)
            
            # Procesar y llenar datos
            wb_output = openpyxl.load_workbook(output_path)
            
            #  PROCESAR CONFIGURACIONES MANUALES SELECTIVAS
            debug_info.append(" PROCESANDO CONFIGURACIONES MANUALES...")
            
            # Configuraciones selectivas por fila
            stock_overrides = {}
            marca_modelo_overrides = {}
            forma_envio_overrides = {}
            costo_envio_overrides = {}
            color_overrides = {}
            
            # Procesar stock selectivo
            if manual_config.get('stock_selective'):
                try:
                    for item in manual_config['stock_selective'].split(','):
                        if ':' in item:
                            row, stock = item.strip().split(':', 1)
                            stock_overrides[int(row)] = stock.strip()
                            debug_info.append(f" Stock selectivo fila {row}: {stock}")
                except Exception as e:
                    debug_info.append(f" Error en stock selectivo: {e}")
            
            # Procesar marca/modelo selectivo
            if manual_config.get('marca_modelo_selective'):
                try:
                    for item in manual_config['marca_modelo_selective'].split(','):
                        if ':' in item:
                            parts = item.strip().split(':')
                            if len(parts) >= 3:
                                row, marca, modelo = parts[0], parts[1], ':'.join(parts[2:])
                                marca_modelo_overrides[int(row)] = {'marca': marca.strip(), 'modelo': modelo.strip()}
                                debug_info.append(f" Marca/Modelo fila {row}: {marca}/{modelo}")
                except Exception as e:
                    debug_info.append(f" Error en marca/modelo selectivo: {e}")
            
            #  Procesar forma de envo selectiva
            if manual_config.get('forma_envio_selective'):
                try:
                    for item in manual_config['forma_envio_selective'].split(','):
                        if ':' in item:
                            row, forma = item.strip().split(':', 1)
                            forma_envio_overrides[int(row)] = forma.strip()
                            debug_info.append(f" Forma env铆o fila {row}: {forma}")
                except Exception as e:
                    debug_info.append(f" Error en forma env铆o selectiva: {e}")
            
            #  Procesar costo de envo selectivo
            if manual_config.get('costo_envio_selective'):
                try:
                    for item in manual_config['costo_envio_selective'].split(','):
                        if ':' in item:
                            row, costo = item.strip().split(':', 1)
                            costo_envio_overrides[int(row)] = costo.strip()
                            debug_info.append(f" Costo env铆o fila {row}: {costo}")
                except Exception as e:
                    debug_info.append(f" Error en costo env铆o selectivo: {e}")
            
            #  Procesar color selectivo
            if manual_config.get('color_selective'):
                try:
                    for item in manual_config['color_selective'].split(','):
                        if ':' in item:
                            parts = item.strip().split(':')
                            if len(parts) >= 3:
                                row, color, nombre_comercial = parts[0], parts[1], ':'.join(parts[2:])
                                color_overrides[int(row)] = {'color': color.strip(), 'nombre_comercial': nombre_comercial.strip()}
                                debug_info.append(f" Color fila {row}: {color}/{nombre_comercial}")
                except Exception as e:
                    debug_info.append(f" Error en color selectivo: {e}")
            
            # Encontrar hoja de categora
            category_sheet_name = None
            for sheet_name in wb_output.sheetnames:
                if sheet_name.lower() not in ['ayuda', 'legales', 'extra info']:
                    category_sheet_name = sheet_name
                    break
            
            if not category_sheet_name:
                raise ValueError("No se encontr hoja de categora en la plantilla")
            
            output_sheet = wb_output[category_sheet_name]
            
            # Mapear y llenar datos con debug info
            debug_info = fill_ml_template(output_sheet, ml_headers, products, selected_fields, condicion, moneda, manual_config, obligatory_fields, debug_info)
            
            wb_output.save(output_path)
            
            success_message = f"Archivo generado exitosamente! {len(products)} productos procesados."
            if ai_provider != 'manual':
                success_message += f" Con IA {ai_provider.upper()}."
            
            # Agregar informacin sobre configuraci贸n manual aplicada
            manual_summary = ""
            if manual_config:
                applied_configs = []
                if manual_config.get('stock_global'):
                    applied_configs.append(f"Stock global: {manual_config['stock_global']}")
                if manual_config.get('stock_selective'):
                    applied_configs.append("Stock selectivo aplicado")
                if manual_config.get('marca_global'):
                    applied_configs.append(f"Marca global: {manual_config['marca_global']}")
                if manual_config.get('modelo_global'):
                    applied_configs.append(f"Modelo global: {manual_config['modelo_global']}")
                if manual_config.get('retiro_persona'):
                    applied_configs.append(f"Retiro persona: {manual_config['retiro_persona']}")
                if manual_config.get('tipo_garantia'):
                    applied_configs.append(f"Garant铆a: {manual_config['tipo_garantia']}")
                if manual_config.get('descripcion_global'):
                    applied_configs.append("Descripci贸n global agregada")
                #  NUEVAS CONFIGURACIONES
                if manual_config.get('forma_envio_global'):
                    applied_configs.append(f"Forma env铆o: {manual_config['forma_envio_global']}")
                if manual_config.get('costo_envio_global'):
                    applied_configs.append(f"Costo env铆o: {manual_config['costo_envio_global']}")
                if manual_config.get('color_global'):
                    applied_configs.append(f"Color: {manual_config['color_global']}")
                
                if applied_configs:
                    manual_summary = f" Configuraci贸n manual aplicada: {', '.join(applied_configs)}"
            
            # Convertir debug_info a string para mostrar
            debug_info_str = '\n'.join(debug_info) if debug_info else None
            
            return render_template_string(HTML_TEMPLATE, 
                                       message=success_message,
                                       message_type="success",
                                       output_file=output_filename,
                                       ai_summary=ai_summary,
                                       manual_summary=manual_summary,
                                       debug_info=debug_info_str)
            
        except Exception as e:
            return render_template_string(HTML_TEMPLATE, 
                                       message=f"Error al procesar archivos: {str(e)}", 
                                       message_type="error")
    
    return render_template_string(HTML_TEMPLATE, user_info=session)

def fill_ml_template(sheet, headers, products, selected_fields, condicion, moneda, manual_config=None, obligatory_fields=None, debug_info=None):
    """Llena la plantilla ML con datos de productos y configuraci贸n manual masiva con mapeo inteligente"""
    
    if debug_info is None:
        debug_info = []
        
    debug_info.append(" INICIANDO MAPEO INTELIGENTE ML...")
    debug_info.append(f" Headers ML detectados: {len(headers)}")
    debug_info.append(f" Productos a mapear: {len(products)}")
    debug_info.append(f" Campos obligatorios detectados: {sum(1 for v in (obligatory_fields or {}).values() if v)}")
    
    #  MAPEO DINMICO INTELIGENTE basado en headers detectados
    ml_columns = {}
    
    # Buscar columnas por contenido de header (ms flexible)
    for col, header in headers.items():
        if isinstance(header, str):
            header_lower = header.lower()
            
            # Mapeo inteligente basado en palabras clave
            if 'ttulo' in header_lower and 'incluye producto' in header_lower:
                ml_columns['titulo'] = col
            elif header_lower.strip() == 'condicin':
                ml_columns['condicion'] = col
            elif 'cdigo universal' in header_lower:
                ml_columns['codigo_universal'] = col
            elif 'vara por' in header_lower and 'color' in header_lower and 'comercial' not in header_lower:
                ml_columns['color'] = col
            elif 'nombre comercial' in header_lower and 'color' in header_lower:
                ml_columns['color_comercial'] = col
            elif header_lower.strip() == 'sku':
                ml_columns['sku'] = col
            elif header_lower.strip() == 'stock':
                ml_columns['stock'] = col
            elif header_lower.strip() == 'precio':
                ml_columns['precio'] = col
            elif header_lower.strip() == 'moneda':
                ml_columns['moneda'] = col
            elif header_lower.strip() == 'descripcin':
                ml_columns['descripcion'] = col
            elif 'forma de envo' in header_lower:
                ml_columns['forma_envio'] = col
            elif 'costo de envo' in header_lower:
                ml_columns['costo_envio'] = col
            elif 'retiro en persona' in header_lower:
                ml_columns['retiro_persona'] = col
            elif 'tipo de garanta' in header_lower:
                ml_columns['garantia_tipo'] = col
            elif 'tiempo de garanta' in header_lower:
                ml_columns['garantia_tiempo'] = col
            elif 'unidad de tiempo' in header_lower and 'garanta' in header_lower:
                ml_columns['garantia_unidad'] = col
            elif header_lower.strip() == 'marca':
                ml_columns['marca'] = col
            elif header_lower.strip() == 'modelo':
                ml_columns['modelo'] = col
            elif 'cdigo de catlogo' in header_lower:
                ml_columns['catalogo'] = col
            elif header_lower.strip() == 'material':
                ml_columns['material'] = col
            elif header_lower.strip() == 'peso':
                ml_columns['peso'] = col
            elif 'dimetro' in header_lower and 'unidad' not in header_lower:
                ml_columns['diametro'] = col
            elif 'largo' in header_lower and 'unidad' not in header_lower:
                ml_columns['largo'] = col
    
    debug_info.append(" MAPEO DINMICO COMPLETADO:")
    for field, col in ml_columns.items():
        obligatorio = obligatory_fields.get(col, False) if obligatory_fields else False
        status = " OBLIGATORIO" if obligatorio else " Opcional"
        col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
        debug_info.append(f"  {field} -> Columna {col_letter} ({col}) - {status}")
    
    # Contadores para estadsticas
    productos_mapeados = 0
    precios_mapeados = 0
    campos_aplicados = {field: 0 for field in ml_columns.keys()}
    
    #  Procesar configuraciones selectivas (acceder desde manual_config)
    stock_overrides = {}
    marca_modelo_overrides = {}
    forma_envio_overrides = {}
    costo_envio_overrides = {}
    color_overrides = {}
    
    if manual_config:
        # Procesar stock selectivo
        if manual_config.get('stock_selective'):
            try:
                for item in manual_config['stock_selective'].split(','):
                    if ':' in item:
                        row, stock = item.strip().split(':', 1)
                        stock_overrides[int(row)] = stock.strip()
            except:
                pass
        
        # Procesar marca/modelo selectivo  
        if manual_config.get('marca_modelo_selective'):
            try:
                for item in manual_config['marca_modelo_selective'].split(','):
                    if ':' in item:
                        parts = item.strip().split(':')
                        if len(parts) >= 3:
                            row, marca, modelo = parts[0], parts[1], ':'.join(parts[2:])
                            marca_modelo_overrides[int(row)] = {'marca': marca.strip(), 'modelo': modelo.strip()}
            except:
                pass
        
        # Procesar forma de envo selectiva
        if manual_config.get('forma_envio_selective'):
            try:
                for item in manual_config['forma_envio_selective'].split(','):
                    if ':' in item:
                        row, forma = item.strip().split(':', 1)
                        forma_envio_overrides[int(row)] = forma.strip()
            except:
                pass
        
        # Procesar costo de envo selectivo
        if manual_config.get('costo_envio_selective'):
            try:
                for item in manual_config['costo_envio_selective'].split(','):
                    if ':' in item:
                        row, costo = item.strip().split(':', 1)
                        costo_envio_overrides[int(row)] = costo.strip()
            except:
                pass
        
        # Procesar color selectivo
        if manual_config.get('color_selective'):
            try:
                for item in manual_config['color_selective'].split(','):
                    if ':' in item:
                        parts = item.strip().split(':')
                        if len(parts) >= 3:
                            row, color, nombre_comercial = parts[0], parts[1], ':'.join(parts[2:])
                            color_overrides[int(row)] = {'color': color.strip(), 'nombre_comercial': nombre_comercial.strip()}
            except:
                pass
    
    # Mapear cada producto
    for i, product in enumerate(products):
        row_num = i + 8  #  FIX CRTICO: Fila en Excel (empezando desde la 8, NO desde la 4)
        producto_mapeado = False
        
        debug_info.append(f" Procesando producto {i+1}/{len(products)} (fila {row_num})")
        
        # Mapear campos bsicos
        for field in selected_fields:
            if field in ml_columns:
                col = ml_columns[field]
                value = find_product_value(product, field)
                
                if value:
                    safe_set_cell_value(sheet, row_num, col, value)
                    campos_aplicados[field] += 1
                    producto_mapeado = True
                    
                    #  Debug especial para precios
                    if field == 'precio':
                        precios_mapeados += 1
                        debug_info.append(f" PRECIO MAPEADO fila {row_num}: {value}")
        
        #  APLICAR CONFIGURACIONES MANUALES GLOBALES Y SELECTIVAS
        
        # Stock
        if manual_config.get('stock_global') and 'stock' in ml_columns:
            if row_num not in stock_overrides:  # No override selectivo
                safe_set_cell_value(sheet, row_num, ml_columns['stock'], manual_config['stock_global'])
                campos_aplicados['stock'] += 1
                producto_mapeado = True
        
        # Stock selectivo
        if row_num in stock_overrides and 'stock' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['stock'], stock_overrides[row_num])
            campos_aplicados['stock'] += 1
            producto_mapeado = True
        
        # Marca y Modelo globales
        if manual_config.get('marca_global') and 'marca' in ml_columns:
            if row_num not in marca_modelo_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['marca'], manual_config['marca_global'])
                campos_aplicados['marca'] += 1
                producto_mapeado = True
                
        if manual_config.get('modelo_global') and 'modelo' in ml_columns:
            if row_num not in marca_modelo_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['modelo'], manual_config['modelo_global'])
                campos_aplicados['modelo'] += 1
                producto_mapeado = True
        
        # Marca/Modelo selectivos
        if row_num in marca_modelo_overrides:
            override = marca_modelo_overrides[row_num]
            if 'marca' in ml_columns:
                safe_set_cell_value(sheet, row_num, ml_columns['marca'], override['marca'])
                campos_aplicados['marca'] += 1
                producto_mapeado = True
            if 'modelo' in ml_columns:
                safe_set_cell_value(sheet, row_num, ml_columns['modelo'], override['modelo'])
                campos_aplicados['modelo'] += 1
                producto_mapeado = True
        
        #  NUEVAS CONFIGURACIONES
        
        # Forma de envo
        if manual_config.get('forma_envio_global') and 'forma_envio' in ml_columns:
            if row_num not in forma_envio_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['forma_envio'], manual_config['forma_envio_global'])
                campos_aplicados['forma_envio'] += 1
                producto_mapeado = True
        
        if row_num in forma_envio_overrides and 'forma_envio' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['forma_envio'], forma_envio_overrides[row_num])
            campos_aplicados['forma_envio'] += 1
            producto_mapeado = True
        
        # Costo de envo
        if manual_config.get('costo_envio_global') and 'costo_envio' in ml_columns:
            if row_num not in costo_envio_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['costo_envio'], manual_config['costo_envio_global'])
                campos_aplicados['costo_envio'] += 1
                producto_mapeado = True
        
        if row_num in costo_envio_overrides and 'costo_envio' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['costo_envio'], costo_envio_overrides[row_num])
            campos_aplicados['costo_envio'] += 1
            producto_mapeado = True
        
        # Color
        if manual_config.get('color_global') and 'color' in ml_columns:
            if row_num not in color_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['color'], manual_config['color_global'])
                campos_aplicados['color'] += 1
                producto_mapeado = True
        
        if manual_config.get('color_comercial_global') and 'color_comercial' in ml_columns:
            if row_num not in color_overrides:
                safe_set_cell_value(sheet, row_num, ml_columns['color_comercial'], manual_config['color_comercial_global'])
                campos_aplicados['color_comercial'] += 1
                producto_mapeado = True
        
        if row_num in color_overrides:
            override = color_overrides[row_num]
            if 'color' in ml_columns:
                safe_set_cell_value(sheet, row_num, ml_columns['color'], override['color'])
                campos_aplicados['color'] += 1
                producto_mapeado = True
            if 'color_comercial' in ml_columns:
                safe_set_cell_value(sheet, row_num, ml_columns['color_comercial'], override['nombre_comercial'])
                campos_aplicados['color_comercial'] += 1
                producto_mapeado = True
        
        # Retiro en persona
        if manual_config.get('retiro_persona') and 'retiro_persona' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['retiro_persona'], manual_config['retiro_persona'])
            campos_aplicados['retiro_persona'] += 1
            producto_mapeado = True
        
        # Garanta
        if manual_config.get('tipo_garantia') and 'garantia_tipo' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['garantia_tipo'], manual_config['tipo_garantia'])
            campos_aplicados['garantia_tipo'] += 1
            producto_mapeado = True
            
        if manual_config.get('tiempo_garantia') and 'garantia_tiempo' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['garantia_tiempo'], manual_config['tiempo_garantia'])
            campos_aplicados['garantia_tiempo'] += 1
            producto_mapeado = True
            
        if manual_config.get('unidad_garantia') and 'garantia_unidad' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['garantia_unidad'], manual_config['unidad_garantia'])
            campos_aplicados['garantia_unidad'] += 1
            producto_mapeado = True
        
        # Cdigos universales masivos
        if manual_config.get('codigo_universal_masivo') and 'codigo_universal' in ml_columns:
            codigo_masivo = manual_config['codigo_universal_masivo']
            # Agregar nmero secuencial si se especifica
            if manual_config.get('codigo_universal_secuencial', False):
                # Usar el ndice actual + offset si se especifica
                offset = manual_config.get('codigo_universal_offset', 1)
                numero_secuencial = row_num - 7 + offset  #  FIX: Ajustar porque empezamos en fila 8 (no 4)
                codigo_final = f"{codigo_masivo}{numero_secuencial:04d}"  # Formato con 4 dgitos
            else:
                codigo_final = codigo_masivo
            
            safe_set_cell_value(sheet, row_num, ml_columns['codigo_universal'], codigo_final)
            campos_aplicados['codigo_universal'] += 1
            producto_mapeado = True
        
        # Valores por defecto (condicin, moneda, envo)
        if 'condicion' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['condicion'], condicion)
            campos_aplicados['condicion'] += 1
            producto_mapeado = True
            
        if 'moneda' in ml_columns:
            safe_set_cell_value(sheet, row_num, ml_columns['moneda'], moneda)
            campos_aplicados['moneda'] += 1
            producto_mapeado = True
        
        #  VALORES POR DEFECTO PARA ENVO (CAMPOS OBLIGATORIOS)
        if 'forma_envio' in ml_columns and not manual_config.get('forma_envio_global'):
            safe_set_cell_value(sheet, row_num, ml_columns['forma_envio'], 'Mercado Envos')
            campos_aplicados['forma_envio'] += 1
            producto_mapeado = True
            
        if 'costo_envio' in ml_columns and not manual_config.get('costo_envio_global'):
            safe_set_cell_value(sheet, row_num, ml_columns['costo_envio'], 'A cargo del comprador')
            campos_aplicados['costo_envio'] += 1
            producto_mapeado = True
            
        if 'retiro_persona' in ml_columns and not manual_config.get('retiro_persona'):
            safe_set_cell_value(sheet, row_num, ml_columns['retiro_persona'], 'Acepto')
            campos_aplicados['retiro_persona'] += 1
            producto_mapeado = True
        
        # Descripcin global
        if manual_config.get('descripcion_global') and 'descripcion' in ml_columns:
            existing_desc = find_product_value(product, 'descripcion') or ""
            combined_desc = f"{existing_desc}\n\n{manual_config['descripcion_global']}" if existing_desc else manual_config['descripcion_global']
            safe_set_cell_value(sheet, row_num, ml_columns['descripcion'], combined_desc)
            campos_aplicados['descripcion'] += 1
            producto_mapeado = True
        
        if producto_mapeado:
            productos_mapeados += 1
    
    #  Generar estadsticas finales
    debug_info.append(" ESTADSTICAS FINALES DE MAPEO:")
    debug_info.append(f" Productos mapeados: {productos_mapeados}/{len(products)}")
    debug_info.append(f" Precios mapeados: {precios_mapeados}")
    
    for campo, cantidad in campos_aplicados.items():
        if cantidad > 0:
            debug_info.append(f" {campo}: {cantidad} valores aplicados")
    
    return debug_info
    
    # Mapeo de campos
    field_mapping = {
        'titulo': 'T铆tulo: incluye producto, marca, modelo y destaca sus caracter铆sticas principales',
        'precio': 'Precio',
        'stock': 'Stock',
        'sku': 'SKU',
        'descripcion': 'Descripci贸n',
        'marca': 'Marca',
        'modelo': 'Modelo',
        'codigo_universal': 'C贸digo universal de producto',
        'color': 'Color',
        'peso': 'Peso',
        'retiro_persona': 'Acepto retiro del comprador en domicilio del vendedor',
        'tipo_garantia': 'Tipo de garant铆a',
        'tiempo_garantia': 'Tiempo de garant铆a',
        'unidad_garantia': 'Unidad de tiempo de la garant铆a',
        'catalogo': 'El producto tiene cat谩logo',
        'numero_catalogo': 'N煤mero de cat谩logo'
    }
    
    # Encontrar columnas en la plantilla
    column_map = {}
    for col_num, header in headers.items():
        for field, ml_field in field_mapping.items():
            if ml_field in header:
                column_map[field] = col_num
                break
    
    # Procesar configuraci贸n selectiva de la configuraci贸n manual
    stock_selective_dict = {}
    marca_modelo_selective_dict = {}
    catalogo_selective_dict = {}
    codigo_universal_selective_dict = {}
    
    if manual_config:
        # Procesar stock selectivo
        if manual_config.get('stock_selective'):
            try:
                for item in manual_config['stock_selective'].split(','):
                    if ':' in item:
                        fila, stock = item.strip().split(':')
                        stock_selective_dict[int(fila)] = int(stock)
            except Exception as e:
                print(f"Error procesando stock selectivo: {e}")
        
        # Procesar marca/modelo selectivo
        if manual_config.get('marca_modelo_selective'):
            try:
                for item in manual_config['marca_modelo_selective'].split(','):
                    if ':' in item:
                        parts = item.strip().split(':')
                        if len(parts) >= 3:
                            fila, marca, modelo = parts[0], parts[1], parts[2]
                            marca_modelo_selective_dict[int(fila)] = {'marca': marca, 'modelo': modelo}
            except Exception as e:
                print(f"Error procesando marca/modelo selectivo: {e}")
        
        # Procesar catlogo selectivo
        if manual_config.get('numero_catalogo_selective'):
            try:
                for item in manual_config['numero_catalogo_selective'].split(','):
                    if ':' in item:
                        fila, numero = item.strip().split(':')
                        catalogo_selective_dict[int(fila)] = numero
            except Exception as e:
                print(f"Error procesando catlogo selectivo: {e}")
        
        # Procesar cdigos universales selectivos
        if manual_config.get('codigo_universal_selective'):
            try:
                for item in manual_config['codigo_universal_selective'].split(','):
                    if ':' in item:
                        fila, codigo = item.strip().split(':')
                        codigo_universal_selective_dict[int(fila)] = codigo
            except Exception as e:
                print(f"Error procesando cdigo universal selectivo: {e}")
    
    # Llenar datos para cada producto
    start_row = 8  # Comenzar despus de las filas de headers e instrucciones
    
    for i, product in enumerate(products):
        row_num = start_row + i
        excel_row = row_num  # Fila real en Excel para configuraci贸n selectiva
        
        # Llenar campos seleccionados del archivo de datos
        for field in selected_fields:
            if field in column_map:
                col_num = column_map[field]
                
                # Buscar valor en datos del producto
                value = find_product_value(product, field)
                if value:
                    safe_set_cell_value(sheet, row_num, col_num, value)
        
        #  APLICAR CONFIGURACIN MANUAL MASIVA
        if manual_config:
            
            # Stock: selectivo tiene prioridad sobre global
            if excel_row in stock_selective_dict:
                if 'stock' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['stock'], stock_selective_dict[excel_row])
            elif manual_config.get('stock_global'):
                if 'stock' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['stock'], int(manual_config['stock_global']))
            
            # Marca y Modelo: selectivo tiene prioridad sobre global
            if excel_row in marca_modelo_selective_dict:
                selective_data = marca_modelo_selective_dict[excel_row]
                if 'marca' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['marca'], selective_data['marca'])
                if 'modelo' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['modelo'], selective_data['modelo'])
            else:
                # Aplicar valores globales
                if manual_config.get('marca_global') and 'marca' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['marca'], manual_config['marca_global'])
                if manual_config.get('modelo_global') and 'modelo' in column_map:
                    safe_set_cell_value(sheet, row_num, column_map['modelo'], manual_config['modelo_global'])
            
            # Retiro en persona
            if manual_config.get('retiro_persona') and 'retiro_persona' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['retiro_persona'], manual_config['retiro_persona'])
            
            # Garanta
            if manual_config.get('tipo_garantia') and 'tipo_garantia' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['tipo_garantia'], manual_config['tipo_garantia'])
            if manual_config.get('tiempo_garantia') and 'tiempo_garantia' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['tiempo_garantia'], int(manual_config['tiempo_garantia']))
            if manual_config.get('unidad_garantia') and 'unidad_garantia' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['unidad_garantia'], manual_config['unidad_garantia'])
            
            # Catlogo
            if manual_config.get('tiene_catalogo') and 'catalogo' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['catalogo'], manual_config['tiene_catalogo'])
            
            # Nmero de catlogo selectivo
            if excel_row in catalogo_selective_dict and 'numero_catalogo' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['numero_catalogo'], catalogo_selective_dict[excel_row])
            
            # Cdigo universal selectivo (sobrescribe el masivo si existe)
            if excel_row in codigo_universal_selective_dict and 'codigo_universal' in column_map:
                safe_set_cell_value(sheet, row_num, column_map['codigo_universal'], codigo_universal_selective_dict[excel_row])
            
            # Descripcin global: agregar al final de descripcin existente
            if manual_config.get('descripcion_global') and 'descripcion' in column_map:
                existing_desc = safe_get_cell_value(sheet, row_num, column_map['descripcion']) or ""
                if existing_desc:
                    new_desc = f"{existing_desc}\n\n{manual_config['descripcion_global']}"
                else:
                    new_desc = manual_config['descripcion_global']
                safe_set_cell_value(sheet, row_num, column_map['descripcion'], new_desc)
        
        # Llenar valores por defecto estndar
        if 'Condicin' in [headers.get(col) for col in headers]:
            for col_num, header in headers.items():
                if 'Condicin' in header:
                    safe_set_cell_value(sheet, row_num, col_num, condicion)
                    break
        
        if 'Moneda' in [headers.get(col) for col in headers]:
            for col_num, header in headers.items():
                if 'Moneda' in header:
                    safe_set_cell_value(sheet, row_num, col_num, moneda)
                    break

def find_product_value(product, field):
    """Encuentra el valor correspondiente en los datos del producto"""
    
    # Mapeo flexible de nombres de campos
    field_variations = {
        'titulo': ['titulo', 'title', 'nombre', 'product_name', 'name', 'producto'],
        'precio': ['precio', 'price', 'cost', 'costo'],
        'stock': ['stock', 'cantidad', 'inventory', 'qty'],
        'sku': ['sku', 'codigo', 'code', 'id'],
        'descripcion': ['descripcion', 'description', 'desc'],
        'marca': ['marca', 'brand'],
        'modelo': ['modelo', 'model'],
        'codigo_universal': ['codigo_universal', 'upc', 'ean', 'barcode'],
        'color': ['color', 'colour'],
        'peso': ['peso', 'weight']
    }
    
    if field in field_variations:
        for variation in field_variations[field]:
            for key in product.keys():
                if variation in key.lower():
                    value = product[key]
                    
                    #  CONVERSIN ESPECIAL PARA PRECIOS Y NMEROS
                    if field in ['precio', 'stock']:
                        try:
                            # Remover smbolos de moneda y espacios
                            if isinstance(value, str):
                                cleaned_value = value.replace('$', '').replace(',', '').replace(' ', '').strip()
                                return float(cleaned_value)
                            elif isinstance(value, (int, float)):
                                return float(value)
                        except (ValueError, TypeError):
                            return value  # Devolver valor original si no se puede convertir
                    
                    return value
    
    return None

@app.route('/download/<filename>')
def download_file(filename):
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return f"Error descargando archivo: {str(e)}", 404

@app.route('/generate-from-prompt', methods=['POST'])
@login_required
@premium_required
def generate_from_prompt():
    """NUEVA FUNCIONALIDAD: Generar plantilla ML desde prompt de IA - SOLO PREMIUM"""
    try:
        # Log de actividad premium
        user_manager.db.log_activity(
            session['user_id'], 
            'generate_from_prompt_used', 
            f"Prompt length: {len(request.form.get('product_prompt', ''))}", 
            request.remote_addr
        )
        
        # Obtener datos del formulario
        product_prompt = request.form.get('product_prompt', '').strip()
        total_products = int(request.form.get('total_products', 10))
        main_category = request.form.get('main_category', 'electronica')
        ai_provider = request.form.get('ai_provider_prompt', 'groq')
        ai_api_key = request.form.get('ai_api_key_prompt', '').strip()
        generation_style = request.form.get('generation_style', 'professional')
        
        if not product_prompt:
            return render_template_string(HTML_TEMPLATE, 
                error=" Por favor describe los productos que quieres generar.",
                debug_info="Prompt vaco")
        
        if not ai_api_key:
            return render_template_string(HTML_TEMPLATE, 
                error=" Por favor ingresa tu API Key para usar la IA.",
                debug_info="API Key requerida")
        
        # Validar lmites
        if total_products > 100:
            total_products = 100  # Lmite de seguridad
        
        # Inicializar AI enhancer para generacin
        from ai_enhancer import AIProductEnhancer
        generator = AIProductEnhancer(provider=ai_provider, api_key=ai_api_key)
        
        #  PROMPT MEJORADO PARA GENERACIN COMPLETA
        generation_prompt = f"""
    MISIN: Genera exactamente {total_products} productos reales para vender en Mercado Libre Uruguay.
        
    DESCRIPCIN DEL USUARIO: {product_prompt}
        
        CATEGORA: {main_category}
        ESTILO: {generation_style}
        
        INSTRUCCIONES ESPECFICAS:
        1. Genera productos REALES y EXISTENTES que se puedan comprar en Uruguay
        2. Usa precios realistas en pesos uruguayos ($)
        3. Incluye variaciones realistas (colores, tamaos, capacidades)
        4. Cada producto debe tener datos completos y nicos
        
        FORMATO REQUERIDO - RESPONDE SOLO CON JSON:
        [
          {{
            "titulo": "Nombre completo del producto",
            "precio": 25000,
            "moneda": "$",
            "marca": "Samsung",
            "modelo": "Galaxy S24",
            "descripcion": "Descripci贸n detallada del producto",
            "condicion": "Nuevo",
            "categoria": "Electrnicos > Celulares",
            "codigo_universal": "1234567890123",
            "peso": "200g",
            "color": "Negro",
            "garantia": "12 meses"
          }}
        ]
        
        IMPORTANTE: 
        - Responde SOLO con el JSON vlido
        - NO agregues texto adicional 
        - Genera exactamente {total_products} productos
    - Los c贸digos EAN-13 deben tener 13 d铆gitos
        """
        
        # Generar productos con IA
        try:
            ai_response = generator.api_call(generation_prompt)
            
            # Intentar parsear JSON
            import json
            try:
                products_data = json.loads(ai_response)
            except json.JSONDecodeError:
                # Si no es JSON vlido, intentar extraer JSON del texto
                import re
                json_match = re.search(r'\[.*\]', ai_response, re.DOTALL)
                if json_match:
                    products_data = json.loads(json_match.group())
                else:
                    raise ValueError("No se pudo extraer JSON vlido de la respuesta")
            
            if not isinstance(products_data, list):
                raise ValueError("La respuesta debe ser una lista de productos")
            
            #  CREAR PLANTILLA ML CON PRODUCTOS GENERADOS
            template_filename = f"ML_Generated_Prompt_{datetime.now().strftime('%d-%m-%H_%M_%S')}.xlsx"
            template_path = os.path.join(app.config['UPLOAD_FOLDER'], template_filename)
            
            # Crear workbook y hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos ML"
            
            #  ESTRUCTURA COMPLETA DE MERCADO LIBRE
            headers = [
                "T铆tulo: incluye producto, marca, modelo y destaca sus caracter铆sticas principales",
                "Precio",
                "Moneda",
                "C贸digo universal de producto",
                "Marca",
                "Modelo",
                "Condici贸n",
                "Descripci贸n",
                "Fotos",
                "Video",
                "Garant铆a",
                "Peso",
                "Color",
                "Material principal",
                "Altura de la unidad",
                "Ancho de la unidad", 
                "Profundidad de la unidad",
                "Peso de la unidad",
                "C贸digo de color",
                "Edad mnima recomendada",
                "Edad mxima recomendada",
                "Caracter铆sticas del producto",
                "Kit",
                "Formato",
                "Unidades por pack",
                "Lnea",
                "SKU"
            ]
            
            # Escribir headers con formato
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, size=11)
                cell.fill = openpyxl.styles.PatternFill(start_color="3483FA", end_color="3483FA", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")
            
            # Escribir datos de productos
            for row, product in enumerate(products_data, 2):
                ws.cell(row=row, column=1, value=product.get('titulo', ''))
                ws.cell(row=row, column=2, value=product.get('precio', 0))
                ws.cell(row=row, column=3, value=product.get('moneda', '$'))
                ws.cell(row=row, column=4, value=product.get('codigo_universal', ''))
                ws.cell(row=row, column=5, value=product.get('marca', ''))
                ws.cell(row=row, column=6, value=product.get('modelo', ''))
                ws.cell(row=row, column=7, value=product.get('condicion', 'Nuevo'))
                ws.cell(row=row, column=8, value=product.get('descripcion', ''))
                ws.cell(row=row, column=9, value="")  # Fotos - vaco para que el usuario agregue
                ws.cell(row=row, column=10, value="")  # Video - vaco
                ws.cell(row=row, column=11, value=product.get('garantia', '12 meses'))
                ws.cell(row=row, column=12, value=product.get('peso', ''))
                ws.cell(row=row, column=13, value=product.get('color', ''))
                # Resto de columnas pueden quedar vacas o con valores por defecto
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Guardar archivo
            wb.save(template_path)
            
            # Crear estadsticas de generacin
            total_generated = len(products_data)
            success_message = f"""
             <strong>Plantilla generada exitosamente!</strong><br><br>
             <strong>Estadsticas:</strong><br>
             Productos generados: {total_generated}<br>
             Categora principal: {main_category}<br>
             Estilo: {generation_style}<br>
             Proveedor IA: {ai_provider.upper()}<br><br>
             <strong>Productos incluidos:</strong><br>
            """
            
            # Agregar resumen de productos
            for i, product in enumerate(products_data[:5], 1):  # Mostrar solo los primeros 5
                success_message += f" {product.get('marca', 'N/A')} {product.get('modelo', 'N/A')} - ${product.get('precio', 0):,}<br>"
            
            if total_generated > 5:
                success_message += f" ... y {total_generated - 5} productos ms<br>"
            
            success_message += f"""<br>
             <a href="/download/{template_filename}" style="color: #3483fa; font-weight: 600;">
                DESCARGAR PLANTILLA ML GENERADA
            </a>
            """
            
            return render_template_string(HTML_TEMPLATE, 
                success=success_message,
                debug_info=f" Generacin exitosa: {total_generated} productos creados con {ai_provider}")
                
        except Exception as ai_error:
            error_msg = f" Error generando productos con IA: {str(ai_error)}"
            return render_template_string(HTML_TEMPLATE, 
                error=error_msg,
                debug_info=f"AI Error: {str(ai_error)}")
        
    except Exception as e:
        error_msg = f" Error en generacin por prompt: {str(e)}"
        return render_template_string(HTML_TEMPLATE, 
            error=error_msg,
            debug_info=f"General Error: {str(e)}")

@app.get('/api/health')
def api_health():
    return jsonify({
        'status': 'ok',
        'react': USE_REACT_UI,
        'dev': os.getenv('ML_EXTRACTOR_DEV') == '1'
    })

if USE_REACT_UI:
    # Serve React index for any unknown route (SPA fallback)
    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve_react(path):  # pragma: no cover
        if path.startswith('api/'):
            return jsonify({'error': 'Not found'}), 404
        # Determine active build directory
        base_dir = FRONTEND_DIST
        index_path = os.path.join(base_dir, 'index.html')
        if os.path.exists(index_path):
            return send_from_directory(base_dir, 'index.html')
        return "React build not found. Run npm run build in frontend/ (creates dist/).", 500

if __name__ == '__main__':
    mode = 'React SPA' if USE_REACT_UI else 'Legacy Template'
    print(f" Iniciando Mercado Libre Bulk Mapper Pro ({mode}) en http://localhost:5003")
    app.run(debug=True, host='localhost', port=5003)
