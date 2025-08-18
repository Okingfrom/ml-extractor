#!/usr/bin/env python3
"""
Aplicación Flask DEMO para mapeo de productos a Mercado Libre
VERSION SIN AUTENTICACIÓN para testing en VS Code Simple Browser
"""

from flask import Flask, request, render_template_string, send_file, redirect, url_for, jsonify, session, render_template, flash
import openpyxl
from openpyxl.cell import MergedCell
import csv
import docx
import PyPDF2
import yaml
import os
import tempfile
from datetime import datetime, timedelta

#  FUNCCIÓN HELPER PARA MANEJO SEGURO DE CELDAS
def safe_get_cell_value(sheet, row, col):
    """Obtiene el valor de una celda manejando MergedCell de forma segura"""
    try:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # Buscar el valor en la celda principal del rango fusionado
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
                    return top_left.value
            return None
        else:
            return cell.value
    except Exception as e:
        print(f" Error accediendo celda ({row},{col}): {e}")
        return None

def safe_set_cell_value(sheet, row, col, value):
    """Establece el valor de una celda manejando MergedCell de forma segura"""
    try:
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # Para celdas fusionadas, escribir en la celda principal
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                    top_left_cell.value = value
                    return
        else:
            cell.value = value
    except Exception as e:
        print(f" Error escribiendo celda ({row},{col}): {e}")

app = Flask(__name__)
app.secret_key = 'demo_secret_key_for_testing'

# Simular usuario premium para demo
def get_current_user():
    return {
        'user_id': 1,
        'username': 'demo_user',
        'email': 'demo@test.com',
        'first_name': 'Demo',
        'last_name': 'User',
        'account_type': 'premium',
        'user_type': 'premium'
    }

# HTML TEMPLATE CON ESTILOS MERCADO LIBRE
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Mercado Libre Bulk Mapper Pro - DEMO</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #fff159 0%, #ffed69 50%, #3483fa 100%);
            min-height: 100vh;
            line-height: 1.6;
        }

        .container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: rgba(255, 255, 255, 1);
            backdrop-filter: blur(10px);
            border-radius: 20px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
            margin-top: 20px;
            margin-bottom: 20px;
        }

        .header {
            text-align: center;
            margin-bottom: 30px;
            padding: 30px 0;
            background: linear-gradient(135deg, #3483fa, #2968c8);
            color: white;
            border-radius: 15px;
            box-shadow: 0 4px 15px rgba(52, 131, 250, 0.3);
        }

        .header h1 {
            font-size: 2.5rem;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.3);
        }

        .header .subtitle {
            font-size: 1.2rem;
            opacity: 0.9;
            font-weight: 300;
        }

        .demo-banner {
            background: linear-gradient(135deg, #fff159, #ffed69);
            color: #333;
            padding: 15px;
            text-align: center;
            border-radius: 10px;
            margin-bottom: 20px;
            font-weight: bold;
            box-shadow: 0 2px 10px rgba(255, 241, 89, 0.3);
        }

        .user-header {
            background: rgba(52, 131, 250, 0.1);
            backdrop-filter: blur(10px);
            padding: 20px;
            border-radius: 15px;
            margin-bottom: 25px;
            border: 1px solid rgba(52, 131, 250, 0.2);
        }

        .user-info {
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
        }

        .user-details h2 {
            color: #3483fa;
            margin-bottom: 5px;
        }

        .user-details p {
            color: #666;
            font-size: 0.9rem;
        }

        .premium-badge {
            background: linear-gradient(135deg, #fff159, #ffed69);
            color: #333;
            padding: 8px 15px;
            border-radius: 20px;
            font-weight: bold;
            font-size: 0.9rem;
            box-shadow: 0 2px 10px rgba(255, 241, 89, 0.3);
        }

        .premium-alert {
            background: rgba(255, 241, 89, 0.2);
            backdrop-filter: blur(10px);
            border: 2px solid #fff159;
            color: #333;
            padding: 20px;
            border-radius: 15px;
            margin-bottom: 25px;
            text-align: center;
        }

        .premium-alert h3 {
            color: #3483fa;
            margin-bottom: 10px;
        }

        .ai-section {
            background: rgba(52, 131, 250, 0.05);
            backdrop-filter: blur(10px);
            padding: 25px;
            border-radius: 15px;
            margin-bottom: 30px;
            border: 1px solid rgba(52, 131, 250, 0.1);
        }

        .ai-section h3 {
            color: #3483fa;
            margin-bottom: 15px;
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .ai-icon {
            font-size: 1.5rem;
        }

        .manual-config-section {
            background: rgba(255, 255, 255, 0.9);
            backdrop-filter: blur(10px);
            padding: 25px;
            border-radius: 15px;
            margin-bottom: 30px;
            border: 1px solid rgba(52, 131, 250, 0.1);
        }

        .manual-config-section h3 {
            color: #3483fa;
            margin-bottom: 15px;
        }

        .form-group {
            margin-bottom: 20px;
        }

        .form-group label {
            display: block;
            margin-bottom: 8px;
            color: #333;
            font-weight: 500;
        }

        .form-control {
            width: 100%;
            padding: 12px 15px;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-size: 1rem;
            transition: all 0.3s ease;
            background: rgba(255, 255, 255, 0.9);
        }

        .form-control:focus {
            outline: none;
            border-color: #3483fa;
            box-shadow: 0 0 0 3px rgba(52, 131, 250, 0.1);
        }

        .btn {
            background: linear-gradient(135deg, #3483fa, #2968c8);
            color: white;
            padding: 12px 25px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 1rem;
            font-weight: 500;
            transition: all 0.3s ease;
            text-decoration: none;
            display: inline-block;
            box-shadow: 0 4px 15px rgba(52, 131, 250, 0.3);
        }

        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(52, 131, 250, 0.4);
        }

        .btn-secondary {
            background: linear-gradient(135deg, #fff159, #ffed69);
            color: #333;
            box-shadow: 0 4px 15px rgba(255, 241, 89, 0.3);
        }

        .btn-secondary:hover {
            box-shadow: 0 6px 20px rgba(255, 241, 89, 0.4);
        }

        .progress-container {
            background: rgba(255, 255, 255, 0.9);
            backdrop-filter: blur(10px);
            padding: 20px;
            border-radius: 15px;
            margin-top: 20px;
            display: none;
            border: 1px solid rgba(52, 131, 250, 0.1);
        }

        .progress-bar {
            width: 100%;
            height: 20px;
            background-color: #e0e0e0;
            border-radius: 10px;
            overflow: hidden;
            margin-bottom: 10px;
        }

        .progress-fill {
            height: 100%;
            background: linear-gradient(135deg, #3483fa, #2968c8);
            width: 0%;
            transition: width 0.3s ease;
        }

        .file-info {
            background: rgba(52, 131, 250, 0.05);
            backdrop-filter: blur(10px);
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
            border: 1px solid rgba(52, 131, 250, 0.1);
        }

        .footer {
            text-align: center;
            padding: 30px 0;
            color: rgba(255, 255, 255, 0.8);
            background: rgba(52, 131, 250, 0.1);
            backdrop-filter: blur(10px);
            border-radius: 15px;
            margin-top: 30px;
        }

        @media (max-width: 768px) {
            .container {
                margin: 10px;
                padding: 15px;
            }
            
            .header h1 {
                font-size: 2rem;
            }
            
            .user-info {
                flex-direction: column;
                gap: 15px;
                text-align: center;
            }
        }

        .alert {
            padding: 15px;
            margin-bottom: 20px;
            border-radius: 10px;
            border: 1px solid transparent;
        }

        .alert-success {
            background-color: rgba(76, 175, 80, 0.1);
            border-color: #4caf50;
            color: #2e7d32;
        }

        .alert-danger {
            background-color: rgba(244, 67, 54, 0.1);
            border-color: #f44336;
            color: #c62828;
        }

        .alert-warning {
            background-color: rgba(255, 193, 7, 0.1);
            border-color: #ffc107;
            color: #f57c00;
        }

        .tab-container {
            margin-bottom: 30px;
        }

        .tabs {
            display: flex;
            gap: 5px;
            margin-bottom: 20px;
            background: rgba(52, 131, 250, 0.05);
            padding: 5px;
            border-radius: 15px;
        }

        .tab {
            flex: 1;
            padding: 12px 20px;
            background: transparent;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 1rem;
            font-weight: 500;
            color: #666;
            transition: all 0.3s ease;
        }

        .tab.active {
            background: linear-gradient(135deg, #3483fa, #2968c8);
            color: white;
            box-shadow: 0 4px 15px rgba(52, 131, 250, 0.3);
        }

        .tab-content {
            display: none;
        }

        .tab-content.active {
            display: block;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Mercado Libre Bulk Mapper Pro</h1>
            <div class="subtitle">Mapeo Inteligente de Productos con IA - VERSIÓN DEMO</div>
        </div>

        <div class="demo-banner">
            ✨ MODO DEMO ACTIVO - Sin autenticación para testing en VS Code Simple Browser ✨
        </div>

        <!-- User Header -->
        <div class="user-header">
            <div class="user-info">
                <div class="user-details">
                    <h2>👋 Bienvenido, Demo User</h2>
                    <p>📧 demo@test.com | 🕒 Última conexión: Ahora</p>
                </div>
                <div class="premium-badge">
                    ⭐ PREMIUM DEMO
                </div>
            </div>
        </div>

        <!-- Premium Alert -->
        <div class="premium-alert">
            <h3>🎯 Acceso Premium Completo - DEMO</h3>
            <p>Todas las funcionalidades premium están habilitadas en esta versión demo para testing</p>
        </div>

        <!-- Tab Container -->
        <div class="tab-container">
            <div class="tabs">
                <button class="tab active" onclick="switchTab('ai')">🤖 IA Automática</button>
                <button class="tab" onclick="switchTab('manual')">⚙️ Config Manual</button>
            </div>

            <!-- AI Tab Content -->
            <div id="ai" class="tab-content active">
                <div class="ai-section">
                    <h3><span class="ai-icon">🤖</span> Mapeo Automático con Inteligencia Artificial</h3>
                    <p style="margin-bottom: 20px;">Sube tu archivo y deja que la IA complete automáticamente los campos faltantes usando las mejores prácticas de Mercado Libre.</p>
                    
                    <form action="/upload_ai" method="post" enctype="multipart/form-data">
                        <div class="form-group">
                            <label for="file">📁 Selecciona tu archivo de productos:</label>
                            <input type="file" id="file" name="file" class="form-control" 
                                   accept=".xlsx,.xls,.csv,.txt,.pdf,.docx" required>
                        </div>
                        
                        <div class="form-group">
                            <label for="ai_completeness">🎯 Nivel de completado automático:</label>
                            <select id="ai_completeness" name="ai_completeness" class="form-control">
                                <option value="conservative">Conservador (solo campos obvios)</option>
                                <option value="balanced" selected>Balanceado (recomendado)</option>
                                <option value="aggressive">Agresivo (máximo completado)</option>
                            </select>
                        </div>
                        
                        <button type="submit" class="btn">🚀 Procesar con IA</button>
                    </form>
                </div>
            </div>

            <!-- Manual Tab Content -->
            <div id="manual" class="tab-content">
                <div class="manual-config-section">
                    <h3>⚙️ Configuración Manual de Mapeo</h3>
                    <p style="margin-bottom: 20px;">Configura manualmente cómo mapear los campos de tu archivo a la plantilla de Mercado Libre.</p>
                    
                    <form action="/upload" method="post" enctype="multipart/form-data">
                        <div class="form-group">
                            <label for="manual_file">📁 Selecciona tu archivo de productos:</label>
                            <input type="file" id="manual_file" name="file" class="form-control" 
                                   accept=".xlsx,.xls,.csv,.txt,.pdf,.docx" required>
                        </div>
                        
                        <div class="form-group">
                            <label for="config_file">📋 Archivo de configuración (opcional):</label>
                            <input type="file" id="config_file" name="config_file" class="form-control" accept=".yaml,.yml">
                            <small style="color: #666; font-size: 0.9rem; margin-top: 5px; display: block;">
                                Si no seleccionas un archivo, se usará la configuración por defecto
                            </small>
                        </div>
                        
                        <button type="submit" class="btn btn-secondary">📊 Procesar Manualmente</button>
                    </form>
                </div>
            </div>
        </div>

        <!-- Progress Container -->
        <div class="progress-container" id="progressContainer">
            <h4>🔄 Procesando archivo...</h4>
            <div class="progress-bar">
                <div class="progress-fill" id="progressFill"></div>
            </div>
            <p id="progressText">Inicializando...</p>
        </div>

        <!-- Footer -->
        <div class="footer">
            <p>© 2024 Mercado Libre Bulk Mapper Pro | Powered by AI | DEMO VERSION</p>
        </div>
    </div>

    <script>
        function switchTab(tabName) {
            // Hide all tab contents
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.remove('active');
            });
            
            // Remove active class from all tabs
            document.querySelectorAll('.tab').forEach(tab => {
                tab.classList.remove('active');
            });
            
            // Show selected tab content
            document.getElementById(tabName).classList.add('active');
            
            // Add active class to clicked tab
            event.target.classList.add('active');
        }

        // Simulate progress for uploads
        function simulateProgress() {
            const container = document.getElementById('progressContainer');
            const fill = document.getElementById('progressFill');
            const text = document.getElementById('progressText');
            
            container.style.display = 'block';
            
            let progress = 0;
            const steps = [
                'Analizando archivo...',
                'Extrayendo datos...',
                'Aplicando IA...',
                'Mapeando campos...',
                'Generando resultado...',
                '¡Completado!'
            ];
            
            const interval = setInterval(() => {
                progress += Math.random() * 20;
                if (progress > 100) progress = 100;
                
                fill.style.width = progress + '%';
                text.textContent = steps[Math.floor(progress / 20)] || steps[steps.length - 1];
                
                if (progress >= 100) {
                    clearInterval(interval);
                    setTimeout(() => {
                        container.style.display = 'none';
                    }, 2000);
                }
            }, 500);
        }

        // Add progress simulation to form submissions
        document.querySelectorAll('form').forEach(form => {
            form.addEventListener('submit', function(e) {
                simulateProgress();
            });
        });
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/upload_ai', methods=['POST'])
def upload_ai():
    flash('🤖 Funcionalidad de IA en desarrollo - próximamente disponible', 'warning')
    return redirect(url_for('index'))

@app.route('/upload', methods=['POST'])
def upload():
    flash('📊 Funcionalidad de procesamiento manual en desarrollo', 'warning')
    return redirect(url_for('index'))

if __name__ == '__main__':
    print(" 🚀 Iniciando Mercado Libre Bulk Mapper Pro DEMO en http://localhost:5005")
    app.run(debug=True, host='localhost', port=5005)
