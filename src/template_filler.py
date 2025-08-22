"""Utilities to fill Mercado Libre templates from row 8 downward while preserving
the template header rows and formatting.

Functions:
- preserve_template_header(workbook, header_rows=7)
- detect_columns(sheet, header_row=8)
- fill_products_from_row(sheet, start_row=8, products_data, default_values, overwrite=False)
- generate_fill_report(products_data, filled_columns)

This module uses openpyxl and works with openpyxl.workbook.workbook.Workbook or
worksheet objects.
"""
from typing import List, Dict, Tuple, Optional, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import re


def _normalize(text: str) -> str:
    if text is None:
        return ""
    text = str(text).strip().lower()
    # remove accents (very small normalization)
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n', 'ü': 'u', 'ç': 'c'
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text


HEADER_KEYWORDS = {
    'title': ['title', 'titulo', 'título', 'nombre', 'titulo del producto', 'nombre producto', 'producto nombre'],
    'sku': ['sku', 'codigo', 'código', 'codigo del producto', 'codigo_sku', 'id', 'id sku', 'referencia'],
    'price': ['price', 'precio', 'valor', 'precio unitario', 'precio venta', 'pvp'],
    'shipping': ['envio', 'envío', 'envio tipo', 'envio_tipo', 'fulfillment', 'metodo envio', 'metodo de envio'],
    'stock': ['stock', 'cantidad', 'available_quantity', 'cantidad disponible', 'quantity', 'stock disponible'],
    'condition': ['condicion', 'condición', 'condition', 'estado', 'condicion del producto', 'estado del producto'],
    'images': ['imagen', 'imagenes', 'imagen principal', 'image', 'images', 'picture', 'pictures', 'image_url', 'url imagen']
}

# Add common programmatic headers used in sample files (e.g. product_title)
HEADER_SUFFIX_MAP = {
    'title': ['product_title', 'title'],
    'price': ['product_price', 'price'],
    'description': ['product_desc', 'description', 'desc', 'product_description'],
    'stock': ['product_stock', 'stock']
}


def detect_header_row(sheet: Worksheet, search_start: int = 1, search_end: int = 12) -> Tuple[int, Dict[str, int]]:
    """Find the most likely header row between search_start and search_end (inclusive).

    Returns (header_row, mapping) where mapping is logical_field -> column index.
    The function prefers the row with the most keyword matches; ties resolved by
    taking the earlier row. If no matches found, returns (search_start, {}).
    """
    best_row = search_start
    best_map: Dict[str, int] = {}
    best_matches = 0
    for r in range(max(1, search_start), min(sheet.max_row, search_end) + 1):
        m, c = {}, 0
        for col_idx in range(1, sheet.max_column + 1):
            text = _normalize(sheet.cell(row=r, column=col_idx).value)
            if not text:
                continue
            # keyword matches
            for key, keywords in HEADER_KEYWORDS.items():
                for kw in keywords:
                    if kw in text:
                        if key not in m:
                            m[key] = col_idx
                            c += 1
                        break
                if key in m and m[key] == col_idx:
                    break
            # suffix/programmatic matches
            for logical, variants in HEADER_SUFFIX_MAP.items():
                if text in variants or any(text.endswith('_' + v) for v in variants):
                    if logical not in m:
                        m[logical] = col_idx
                        c += 1
        if c > best_matches:
            best_matches = c
            best_map = m
            best_row = r

    return best_row, best_map


def preserve_template_header(workbook_or_path: Any, header_rows: int = 7) -> None:
    """No-op helper that documents intent: workbook rows 1..header_rows must not be
    modified by any filling process. This function does not change the workbook; it
    is provided so callers can make intent explicit and for future hook points (e.g.,
    copying header ranges or validating they exist).

    Args:
        workbook_or_path: openpyxl Workbook or a path to a workbook file.
        header_rows: number of rows to preserve (1-based count).
    """
    # The preservation is enforced by filling functions which start writing at start_row >= header_rows+1
    return


def _ensure_workbook(workbook_or_path: Any) -> Workbook:
    if isinstance(workbook_or_path, Workbook):
        return workbook_or_path
    if isinstance(workbook_or_path, str):
        return load_workbook(workbook_or_path)
    raise TypeError('workbook_or_path must be openpyxl Workbook or path string')


def detect_columns(sheet: Worksheet, header_row: Optional[int] = 8) -> Dict[str, int]:
    """Scan header rows to detect logical column indices.

    Behavior:
      - Try the provided header_row first (1-based). If no keywords are found,
        scan a small window around header_row (header_row-2 .. header_row+2)
        and pick the row with the most keyword matches.
      - Returns a mapping logical_field -> 1-based column index.

    Returns:
        dict: { 'title': 3, 'sku': 5, ... }
    """
    def scan_row(r: int) -> Tuple[Dict[str, int], int]:
        """Scan a single row r and return (mapping, matches_count)."""
        mapping: Dict[str, int] = {}
        matches = 0
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=col_idx)
            text = _normalize(cell.value)
            if not text:
                continue
            for key, keywords in HEADER_KEYWORDS.items():
                for kw in keywords:
                    if kw in text:
                        if key not in mapping:
                            mapping[key] = col_idx
                            matches += 1
                        break
                if key in mapping and mapping[key] == col_idx:
                    break
        return mapping, matches

    # If header_row is None or <=0, perform an automatic header search
    if header_row is None or (isinstance(header_row, int) and header_row <= 0):
        header_row = 1

    # If header_row is specifically 'auto' (or caller passed None previously), use detect_header_row
    if header_row == 'auto' or header_row is None:
        r, m = detect_header_row(sheet, search_start=1, search_end=min(12, sheet.max_row))
        return m

    # If a numeric header_row is provided, try it, otherwise scan nearby rows (±2)
    try:
        mapping, matches = scan_row(int(header_row))
    except Exception:
        mapping, matches = {}, 0

    if matches > 0:
        return mapping

    # scan nearby rows (within +/-2 rows) to find the best header row
    best_mapping: Dict[str, int] = {}
    best_matches = 0
    for r in range(max(1, int(header_row) - 2), min(sheet.max_row, int(header_row) + 2) + 1):
        try:
            m, c = scan_row(r)
        except Exception:
            continue
        if c > best_matches:
            best_matches = c
            best_mapping = m

    # fallback: programmatic header names
    if not best_mapping:
        for r in range(max(1, int(header_row) - 2), min(sheet.max_row, int(header_row) + 2) + 1):
            for col_idx in range(1, sheet.max_column + 1):
                text = _normalize(sheet.cell(row=r, column=col_idx).value)
                if not text:
                    continue
                for logical, variants in HEADER_SUFFIX_MAP.items():
                    if text in variants or any(text.endswith('_' + v) for v in variants):
                        if logical not in best_mapping:
                            best_mapping[logical] = col_idx

    return best_mapping


def _coerce_price(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    # strip currency symbols and thousands separators
    s = re.sub(r'[,$€¥£]', '', s)
    s = s.replace('\xa0', '')
    s = s.replace(' ', '')
    try:
        return float(s)
    except Exception:
        return None


def fill_products_from_row(sheet: Worksheet,
                           start_row: Any = 8,
                           products_data: List[Dict[str, Any]] = None,
                           default_values: Dict[str, Any] = None,
                           overwrite: bool = False) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Fill worksheet cells starting at `start_row` (1-based) with products_data.

    Each product in products_data is a dict with keys that match the logical fields
    detected by `detect_columns`, e.g. { 'title': 'Zapato', 'sku': '123', 'price': 10 }

    Behavior:
      - The function writes products sequentially: the first product goes to start_row,
        the next to start_row+1, etc.
      - It will not write any cells in rows < start_row, preserving the template header.
      - If overwrite=False, it will not overwrite a non-empty cell in the target.
      - For missing product fields, it will use default_values if provided.

    Returns:
        Tuple of (filled_per_product, skipped_per_product) where each is a list of dicts
        describing which columns were filled or skipped for each product.
    """
    products = products_data or []
    defaults = default_values or {}
    # Always preserve rows 1..7. Regardless of detected header row, we will write
    # starting at row 8 to ensure the Mercado Libre template's top rows remain intact.
    requested_header = start_row
    # If caller asked for auto detection, detect header row and mapping
    if start_row in (None, 'auto'):
        header_row_found, mapping = detect_header_row(sheet, search_start=1, search_end=min(12, sheet.max_row))
        detected = mapping
        # default write start when auto: ensure minimum row 8
        start_row_val = 8
    else:
        # Allow callers to pass a numeric start_row to write at a different row
        # while still enforcing preservation of rows 1..7. Use the row immediately
        # above the write start as the header row to detect logical columns. This
        # prevents accidentally scanning the first product row as headers when
        # callers pass the Excel data start (e.g. 8).
        try:
            header_to_scan = int(start_row) - 1 if isinstance(start_row, int) else 8 - 1
            if header_to_scan < 1:
                header_to_scan = 1
        except Exception:
            header_to_scan = 7
        detected = detect_columns(sheet, header_row=header_to_scan)
        # If nothing detected near the requested header_to_scan, try a broader
        # automatic header detection across the top of the sheet (rows 1..12).
        if not detected:
            try:
                hdr_row, hdr_map = detect_header_row(sheet, search_start=1, search_end=min(12, sheet.max_row))
                if hdr_map:
                    detected = hdr_map
            except Exception:
                # ignore and keep detected as-is
                pass
        # Normalize requested start_row into an integer if possible
        try:
            if isinstance(start_row, str) and start_row.isdigit():
                start_row_val = int(start_row)
            elif isinstance(start_row, int):
                start_row_val = start_row
            else:
                start_row_val = 8
        except Exception:
            start_row_val = 8

    # Enforce that writes start at least at Excel row 8 to preserve template top rows
    try:
        start_row = max(8, int(start_row_val))
    except Exception:
        start_row = 8

    filled_report: List[Dict[str, Any]] = []
    skipped_report: List[Dict[str, Any]] = []

    for idx, product in enumerate(products):
        excel_row = start_row + idx
        filled = {}
        skipped = {}
        for logical_field, col_idx in detected.items():
            target_cell = sheet.cell(row=excel_row, column=col_idx)
            # Determine value: product value -> default -> None
            value = product.get(logical_field) if product else None
            if value is None:
                value = defaults.get(logical_field)
            # Coerce price if needed
            if logical_field == 'price' and value is not None:
                coerced = _coerce_price(value)
                if coerced is None:
                    # invalid price, skip
                    skipped[logical_field] = {'reason': 'invalid_price', 'value': value}
                    continue
                value = coerced
            # If not allowed to overwrite and cell already has a value, skip
            if not overwrite and target_cell.value not in (None, ''):
                skipped[logical_field] = {'reason': 'cell_not_empty', 'existing': target_cell.value}
                continue
            # If value is None, skip writing
            if value is None:
                skipped[logical_field] = {'reason': 'no_value'}
                continue
            # Write the value
            target_cell.value = value
            filled[logical_field] = value

        filled_report.append({'row': excel_row, 'filled': filled})
        skipped_report.append({'row': excel_row, 'skipped': skipped})

    return filled_report, skipped_report


def generate_fill_report(products_data: List[Dict[str, Any]],
                         filled_columns: List[Dict[str, Any]],
                         skipped_columns: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Produce a concise summary report describing what was filled and skipped.

    Args:
        products_data: original list of products provided by the user.
        filled_columns: list returned by `fill_products_from_row` (filled_report)
        skipped_columns: list returned by `fill_products_from_row` (skipped_report)

    Returns:
        dict with counts and per-row details.
    """
    total_products = len(products_data or [])
    details = []
    total_filled = 0
    total_skipped = 0
    for i, prod in enumerate(products_data or []):
        filled = filled_columns[i]['filled'] if i < len(filled_columns) else {}
        skipped = skipped_columns[i]['skipped'] if i < len(skipped_columns) else {}
        total_filled += len(filled)
        total_skipped += len(skipped)
        details.append({'product_index': i, 'filled': filled, 'skipped': skipped})

    return {
        'total_products': total_products,
        'total_fields_filled': total_filled,
        'total_fields_skipped': total_skipped,
        'details': details
    }


if __name__ == '__main__':
    # quick local smoke test when invoked directly
    import json
    wb = None
    try:
        wb = load_workbook('../samples/sample_output.xlsx')
    except Exception:
        try:
            wb = load_workbook('..\\samples\\sample_output.xlsx')
        except Exception:
            print('No sample file found for smoke test')
    if wb:
        ws = wb.active
        print('Detected columns:', detect_columns(ws, header_row=8))
        products = [{'title': 'Prueba', 'sku': 'ABC123', 'price': '99.90'}]
        filled, skipped = fill_products_from_row(ws, start_row=8, products_data=products, default_values={'condition': 'Nuevo', 'stock': 1}, overwrite=False)
        print(json.dumps(generate_fill_report(products, filled, skipped), indent=2, ensure_ascii=False))
