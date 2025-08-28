import pandas as pd
import os
import logging
import openpyxl
from openpyxl.cell import MergedCell

logger = logging.getLogger(__name__)

def read_excel(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    if ext == ".xlsx":
        try:
            # 🔧 FIX CRÍTICO: Manejo especial de celdas fusionadas y fórmulas
            logger.info(f"🔍 Leyendo archivo Excel: {file_path}")
            
            # Primer intento: lectura directa con pandas
            try:
                df = pd.read_excel(file_path, engine="openpyxl")
                logger.info(f"✅ Lectura exitosa con pandas - Filas: {len(df)}, Columnas: {len(df.columns)}")
                return df
            except Exception as pandas_error:
                logger.warning(f"⚠️ Pandas falló, intentando con openpyxl manual: {pandas_error}")
                
                # Segundo intento: lectura manual con manejo de MergedCell
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Extraer datos manualmente manejando MergedCell
                data = []
                for row in ws.iter_rows():
                    row_data = []
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            # Para celdas fusionadas, obtener el valor de la celda superior izquierda
                            for merged_range in ws.merged_cells.ranges:
                                if cell.coordinate in merged_range:
                                    # Obtener la celda superior izquierda del rango fusionado
                                    top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                                    row_data.append(top_left.value)
                                    break
                            else:
                                row_data.append(None)
                        else:
                            row_data.append(cell.value)
                    data.append(row_data)
                
                # Convertir a DataFrame
                if data:
                    # Usar la primera fila como headers
                    headers = data[0] if data else []
                    df_data = data[1:] if len(data) > 1 else []
                    df = pd.DataFrame(df_data, columns=headers)
                    logger.info(f"✅ Lectura manual exitosa - Filas: {len(df)}, Columnas: {len(df.columns)}")
                    return df
                else:
                    raise ValueError("No se pudieron extraer datos del archivo")
                    
        except Exception as e:
            print(f"❌ Error crítico leyendo Excel: {e}")
            raise ValueError(f"Failed to read .xlsx file: {e}")
    elif ext == ".xls":
        try:
            return pd.read_excel(file_path, engine="xlrd")
        except Exception as e:
            raise ValueError(f"Failed to read .xls file: {e}")
    else:
        raise ValueError(f"Unsupported Excel file extension: {ext}")
