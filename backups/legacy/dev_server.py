#!/usr/bin/env python3
"""
Complete backend for ML Extractor FastAPI
"""

import uvicorn
import os
import tempfile
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel

# Setup logging
try:
    from core.logging_config import setup_logging, get_logger
    setup_logging()
    logger = get_logger(__name__)
except ImportError:
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

app = FastAPI(
    title="ML Extractor API",
    version="2.0.0-dev",
    description="ML Extractor Backend - Complete File Processing System",
    docs_url="/docs",
    redoc_url="/redoc"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://localhost:3002", "http://127.0.0.1:3000", "http://127.0.0.1:3001", "http://127.0.0.1:3002", "*"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Create uploads directory
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- Optional admin settings router (development convenience) ---
# Try to load admin_settings.py directly by file path to avoid executing
# package-level imports in `backend/api/__init__.py` (that may pull in
# SQLAlchemy or other heavy modules). This keeps the dev server lightweight.
try:
    import importlib.util as _importlib_util
    import sys as _sys
    _admin_path = os.path.join(os.path.dirname(__file__), 'api', 'admin_settings.py')
    if os.path.exists(_admin_path):
        spec = _importlib_util.spec_from_file_location('admin_settings_local', _admin_path)
        if spec and spec.loader:
            mod = _importlib_util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            admin_router = getattr(mod, 'router', None)
            if admin_router is not None:
                app.include_router(admin_router, prefix="/api")
                logger.info("Admin router registered at /api/admin")
            else:
                logger.info("Admin module loaded but no 'router' attribute found")
    else:
        logger.info(f"Admin module not found at {_admin_path}")
except Exception as _err:
    # Don't fail startup if admin router isn't importable in this lightweight dev server
    logger.info(f"Admin router not registered (optional): {_err}")

# Pydantic models
class ProcessingResponse(BaseModel):
    success: bool
    message: str
    output_file: Optional[str] = None
    processing_time: Optional[float] = None
    records_processed: Optional[int] = None

class HealthResponse(BaseModel):
    status: str
    mode: str
    message: str
    timestamp: str

class LoginRequest(BaseModel):
    email: str
    password: str

class AuthResponse(BaseModel):
    success: bool
    message: str
    user: Optional[dict] = None
    token: Optional[str] = None

# Demo users database
DEMO_USERS = {
    "premium@test.com": {
        "id": 1,
        "email": "premium@test.com",
        "password": "Premium123!",
        "name": "Premium User",
        "first_name": "Premium",
        "last_name": "User",
        "account_type": "premium",
        "is_verified": True
    },
    "free@test.com": {
        "id": 2,
        "email": "free@test.com", 
        "password": "Free123!",
        "name": "Free User",
        "first_name": "Free",
        "last_name": "User",
        "account_type": "free",
        "is_verified": True
    }
}

# Root endpoint
@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "message": "ML Extractor API - Complete Development Mode",
        "version": "2.0.0-dev",
        "status": "development",
        "docs": "/docs",
        "health": "/health",
        "endpoints": {
            "file_upload": "/api/files/upload",
            "file_process": "/api/files/process",
            "product_data_analysis": "/api/files/analyze-product-data",
            "health": "/health"
        }
    }

# Health check endpoint
@app.get("/health", response_model=HealthResponse)
async def health():
    """Health check endpoint"""
    return HealthResponse(
        status="healthy",
        mode="development",
        message="Backend is running in development mode",
        timestamp=datetime.now().isoformat()
    )

# API Status endpoint
@app.get("/api/status")
async def api_status():
    """API status endpoint"""
    return {
        "api_version": "2.0.0-dev",
        "backend": "FastAPI",
        "frontend": "React + Tailwind",
        "features": {
            "authentication": "in_development",
            "file_processing": "available",
            "ml_template_analysis": "available",
            "ai_enhancement": "planned",
            "mapping_config": "in_development"
        },
        "supported_formats": ["xlsx", "xls", "csv", "pdf", "docx", "txt"]
    }

# File upload endpoint
@app.post("/api/files/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload a file for processing"""
    try:
        # Validate file type
        allowed_extensions = {'.xlsx', '.xls', '.csv', '.pdf', '.docx', '.txt'}
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"File type {file_extension} not supported. Allowed types: {', '.join(allowed_extensions)}"
            )
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f"{timestamp}_{file.filename}"
        file_path = os.path.join(UPLOAD_DIR, safe_filename)
        
        # Save file
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        return {
            "success": True,
            "message": "File uploaded successfully",
            "filename": safe_filename,
            "original_name": file.filename,
            "size": len(content),
            "type": file_extension,
            "upload_time": datetime.now().isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

# File processing endpoint
@app.post("/api/files/process", response_model=ProcessingResponse)
async def process_file(
    filename: str = Form(...),
    output_format: str = Form(default="xlsx"),
    ai_enhance: bool = Form(default=False)
):
    """Process uploaded file"""
    try:
        start_time = datetime.now()
        
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Simple processing simulation
        # In real implementation, this would call the actual processing logic
        output_filename = f"processed_{filename}"
        output_path = os.path.join(UPLOAD_DIR, output_filename)
        
        # For demo, just copy the file (real processing would happen here)
        import shutil
        shutil.copy2(file_path, output_path)
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        return ProcessingResponse(
            success=True,
            message=f"File processed successfully{'with AI enhancement' if ai_enhance else ''}",
            output_file=output_filename,
            processing_time=processing_time,
            records_processed=100  # Simulated
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

# File download endpoint
@app.get("/api/files/download/{filename}")
async def download_file(filename: str):
    """Download processed file"""
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )

# List files endpoint
@app.get("/api/files/")
async def list_files():
    """List uploaded files"""
    try:
        files = []
        for filename in os.listdir(UPLOAD_DIR):
            file_path = os.path.join(UPLOAD_DIR, filename)
            if os.path.isfile(file_path):
                stat = os.stat(file_path)
                files.append({
                    "filename": filename,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "is_processed": filename.startswith("processed_")
                })
        
        return {
            "files": files,
            "count": len(files)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")

# Product Data Analysis endpoint (replaces ML Template Analysis)
@app.post("/api/files/analyze-product-data")
async def analyze_product_data_endpoint(file: UploadFile = File(...)):
    """Analyze uploaded product data file and suggest mappings to ML template"""
    try:
        # Import the new product analyzer
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))
        
        try:
            from services.product_data_analyzer import analyze_product_data
        except ImportError:
            # Fallback to simple analysis if import fails
            import pandas as pd
            
            # Validate file type
            allowed_extensions = {'.xlsx', '.xls', '.csv'}
            file_extension = os.path.splitext(file.filename)[1].lower()
            
            if file_extension not in allowed_extensions:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tipo de archivo no soportado. Use: {', '.join(allowed_extensions)}"
                )
            
            # Save temporary file for analysis
            temp_filename = f"temp_product_analysis_{file.filename}"
            temp_path = os.path.join(UPLOAD_DIR, temp_filename)
            
            # Save file temporarily
            with open(temp_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)
            
            try:
                # Basic analysis with pandas
                if file_extension == '.csv':
                    df = pd.read_csv(temp_path)
                else:
                    df = pd.read_excel(temp_path)
                
                # Basic product data detection
                columns = df.columns.tolist()
                total_products = len(df.dropna(how='all'))
                sample_data = df.head(3).to_dict('records') if not df.empty else []
                
                # Simple mapping suggestions
                ml_fields = ['Título', 'Precio', 'Stock', 'Categoría', 'Marca', 'Modelo', 'Descripción', 'Peso', 'Garantía', 'Condición', 'Imagen Principal']
                product_indicators = {
                    'Título': ['titulo', 'title', 'nombre', 'product_name', 'name'],
                    'Precio': ['precio', 'price', 'cost', 'valor', 'costo'],
                    'Stock': ['stock', 'cantidad', 'qty', 'quantity'],
                    'Categoría': ['categoria', 'category', 'tipo'],
                    'Marca': ['marca', 'brand'],
                    'Modelo': ['modelo', 'model'],
                    'Descripción': ['descripcion', 'description', 'detalle']
                }
                
                field_mappings = []
                for col in columns:
                    col_lower = col.lower()
                    best_match = "Sin mapear"
                    confidence = "low"
                    
                    for ml_field, patterns in product_indicators.items():
                        if any(pattern in col_lower for pattern in patterns):
                            best_match = ml_field
                            confidence = "high" if any(pattern == col_lower for pattern in patterns) else "medium"
                            break
                    
                    field_mappings.append({
                        "source_column": col,
                        "target_ml_field": best_match,
                        "confidence": confidence,
                        "examples": [str(df[col].iloc[i]) for i in range(min(2, len(df))) if pd.notna(df[col].iloc[i])]
                    })
                
                mapped_fields = len([m for m in field_mappings if m['target_ml_field'] != "Sin mapear"])
                confidence_score = mapped_fields / len(columns) if columns else 0
                
                # Clean up temporary file
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
                return {
                    "file_analysis": {
                        "total_products": total_products,
                        "detected_columns": columns,
                        "column_count": len(columns),
                        "sample_data": sample_data,
                        "file_type": file_extension
                    },
                    "ml_template_mapping": {
                        "field_mappings": field_mappings,
                        "mapping_summary": {
                            "total_source_fields": len(columns),
                            "mapped_fields": mapped_fields,
                            "unmapped_fields": len(columns) - mapped_fields,
                            "overall_confidence_score": round(confidence_score, 2),
                            "ready_for_processing": confidence_score >= 0.6
                        },
                        "ml_template_fields": ml_fields,
                        "data_insertion_row": 8
                    },
                    "recommendations": [
                        f"� Se detectaron {total_products} productos en el archivo",
                        f"🏷️ {len(columns)} columnas encontradas",
                        f"🎯 {mapped_fields} campos mapeados automáticamente",
                        "✅ Revise los mapeos sugeridos antes de procesar" if confidence_score >= 0.5 else "⚠️ Se recomienda mapeo manual",
                        "🚀 Listo para generar plantilla ML" if confidence_score >= 0.6 else "🔧 Ajuste mapeos antes de continuar"
                    ],
                    "status": "success",
                    "message": f"✅ Análisis completado: {total_products} productos detectados con {mapped_fields} campos mapeados"
                }
                
            except Exception as analysis_error:
                # Clean up temporary file
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
                raise HTTPException(
                    status_code=500,
                    detail=f"Error analizando datos de productos: {str(analysis_error)}"
                )


# Close outer try for analyze_product_data_endpoint
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando solicitud: {str(e)}"
        )

    # end of analyze_product_data_endpoint
# ML Template Analysis endpoint
@app.post("/api/files/analyze-ml-template")
async def analyze_ml_template_endpoint(file: UploadFile = File(...)):
    """Analyze uploaded ML template structure and detect fields/categories"""
    try:
        # Import the ML template analyzer
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))

        try:
            from services.ml_template_analyzer import MLTemplateAnalyzer
        except ImportError:
            # Fallback to basic analysis if import fails
            raise HTTPException(
                status_code=500,
                detail="ML Template Analyzer no disponible. Instale las dependencias necesarias."
            )

        # Validate file type
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_extension = os.path.splitext(file.filename)[1].lower()

        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Tipo de archivo no soportado. Use: {', '.join(allowed_extensions)}"
            )

        # Save temporary file for analysis
        temp_filename = f"temp_ml_template_{file.filename}"
        temp_path = os.path.join(UPLOAD_DIR, temp_filename)

        # Save file temporarily
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        try:
            # Analyze ML template with specialized analyzer
            analyzer = MLTemplateAnalyzer()
            analysis_result = analyzer.analyze_template(temp_path)
            
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            # Format response
            if analysis_result.get('success', False):
                analysis = analysis_result.get('analysis', {})
                
                return {
                    "status": "success",
                    "message": analysis_result.get('message', 'Análisis completado'),
                    "analysis": analysis,
                    "validation": {
                        "is_ml_template": analysis.get('is_ml_template', False),
                        "confidence_score": analysis.get('confidence_score', 0.0),
                        "field_count": len(analysis.get('ml_fields', [])),
                        "category_count": len(analysis.get('categories', [])),
                        "structure_valid": analysis.get('validation', {}).get('is_valid', False)
                    },
                    "recommendations": analysis.get('recommendations', []),
                    "file_info": {
                        "filename": file.filename,
                        "size": len(content),
                        "type": file_extension
                    }
                }
            else:
                raise HTTPException(
                    status_code=400,
                    detail=analysis_result.get('error', 'Error en análisis de plantilla ML')
                )
            
        except Exception as analysis_error:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            raise HTTPException(
                status_code=500,
                detail=f"Error analizando plantilla ML: {str(analysis_error)}"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando solicitud de análisis ML: {str(e)}"
        )

# Product Data Analysis endpoint
@app.post("/api/files/analyze-product-data")
async def analyze_product_data_endpoint(file: UploadFile = File(...)):
    """Analyze uploaded product data file and extract structure/fields"""
    try:
        # Validate file type
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Tipo de archivo no soportado. Use: {', '.join(allowed_extensions)}"
            )
        
        # Save temporary file for analysis
        temp_filename = f"temp_product_data_{file.filename}"
        temp_path = os.path.join(UPLOAD_DIR, temp_filename)
        
        # Save file temporarily
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        try:
            import pandas as pd
            
            # Read file based on extension
            if file_extension in ['.xlsx', '.xls']:
                df = pd.read_excel(temp_path)
            else:  # .csv
                df = pd.read_csv(temp_path)
            
            # Analyze structure
            total_rows = len(df)
            total_columns = len(df.columns)
            column_names = df.columns.tolist()
            
            # Analyze data types and sample values
            field_analysis = []
            for col in df.columns:
                sample_values = df[col].dropna().head(3).tolist()
                data_type = str(df[col].dtype)
                non_null_count = df[col].count()
                
                field_analysis.append({
                    "field_name": col,
                    "data_type": data_type,
                    "non_null_count": non_null_count,
                    "null_percentage": round((total_rows - non_null_count) / total_rows * 100, 2),
                    "sample_values": [str(val) for val in sample_values]
                })
            
            # Get sample data for preview (first 5 rows)
            sample_data = []
            for idx, row in df.head(5).iterrows():
                sample_data.append(row.to_dict())
            
            # Calculate compatibility score based on common field patterns
            compatibility_score = 0
            common_patterns = ['titulo', 'title', 'precio', 'price', 'stock', 'cantidad', 'descripcion', 'description']
            
            for col in column_names:
                col_lower = col.lower()
                if any(pattern in col_lower for pattern in common_patterns):
                    compatibility_score += 1
            
            compatibility_percentage = min(round((compatibility_score / len(common_patterns)) * 100), 100)
            
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            return {
                "status": "success",
                "message": "Análisis de productos completado",
                "analysis": {
                    "file_info": {
                        "filename": file.filename,
                        "size": len(content),
                        "type": file_extension,
                        "total_rows": total_rows,
                        "total_columns": total_columns
                    },
                    "structure": {
                        "totalProducts": total_rows,
                        "fieldsDetected": total_columns,
                        "compatibility": f"{compatibility_percentage}%",
                        "fields": column_names
                    },
                    "fields_analysis": field_analysis,
                    "sampleData": sample_data,
                    "validation": {
                        "is_valid": total_rows > 0 and total_columns > 0,
                        "has_data": total_rows > 0,
                        "has_headers": total_columns > 0,
                        "estimated_compatibility": compatibility_percentage
                    }
                }
            }
            
        except Exception as analysis_error:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            raise HTTPException(
                status_code=500,
                detail=f"Error analizando archivo de productos: {str(analysis_error)}"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando archivo de productos: {str(e)}"
        )

# ML File Preview endpoint
@app.post("/api/files/preview-ml-file")
async def preview_ml_file_endpoint(
    ml_template: UploadFile = File(...),
    product_data: UploadFile = File(...),
    mapping_config: str = Form(...),
    default_settings: str = Form(...)
):
    """Generate preview of final ML file without saving - shows first 5 rows"""
    try:
        import json
        import pandas as pd
        import tempfile
        
        # Parse configuration data
        try:
            mapping = json.loads(mapping_config)
            settings = json.loads(default_settings)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error parsing configuration: {str(e)}"
            )
        
        # Validate file types
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        ml_ext = os.path.splitext(ml_template.filename)[1].lower()
        product_ext = os.path.splitext(product_data.filename)[1].lower()
        
        if ml_ext not in allowed_extensions or product_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail="Tipos de archivo no soportados. Use: .xlsx, .xls, .csv"
            )
        
        # Save temporary files
        temp_dir = tempfile.mkdtemp()
        ml_temp_path = os.path.join(temp_dir, f"ml_template_{ml_template.filename}")
        product_temp_path = os.path.join(temp_dir, f"product_data_{product_data.filename}")
        
        # Save uploaded files
        with open(ml_temp_path, "wb") as buffer:
            ml_content = await ml_template.read()
            buffer.write(ml_content)
            
        with open(product_temp_path, "wb") as buffer:
            product_content = await product_data.read()
            buffer.write(product_content)
        
        try:
            # Read ML template (preserving structure)
            ml_df = pd.read_excel(ml_temp_path, header=None) if ml_ext in ['.xlsx', '.xls'] else pd.read_csv(ml_temp_path, header=None)
            
            # Read product data
            if product_ext in ['.xlsx', '.xls']:
                product_df = pd.read_excel(product_temp_path)
            else:
                product_df = pd.read_csv(product_temp_path)

            # Resolve mapping values using the shared resolver in src/mapping_loader.py
            mapping_resolved = {}
            try:
                import sys
                import importlib.util
                src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src'))
                if src_dir not in sys.path:
                    sys.path.insert(0, src_dir)
                
                # Use importlib to safely import mapping_loader
                spec = importlib.util.spec_from_file_location("mapping_loader", os.path.join(src_dir, "mapping_loader.py"))
                mapping_loader_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mapping_loader_module)

                product_cols = list(product_df.columns)
                mapping_resolved = mapping_loader_module.resolve_column_aliases(product_cols, mapping)
            except Exception:
                # fallback: use literal column names where available
                mapping_resolved = {k: (v if isinstance(v, str) and v in list(product_df.columns) else None) for k, v in mapping.items()}
            
            # Find the data start row in ML template
            # Use zero-based index. Setting default to 7 makes data rows start
            # at Excel row 8 (user requirement: "filas desde la 8 en adelante").
            data_start_row = 7  # Default for ML templates (zero-based index -> Excel row 8)
            for i in range(min(15, len(ml_df))):
                row = ml_df.iloc[i]
                non_null_count = row.count()
                if non_null_count >= len(ml_df.columns) * 0.5:
                    data_start_row = i
                    break
            
            # Get ML template headers for preview (normalized)
            headers = []
            if data_start_row > 0 and data_start_row - 1 < len(ml_df):
                header_row = ml_df.iloc[data_start_row - 1]
                headers = [str(cell).strip() if pd.notna(cell) else f"Columna {i+1}" for i, cell in enumerate(header_row)]
            else:
                headers = [f"Columna {i+1}" for i in range(len(ml_df.columns))]

            # Build normalized lowercase headers for detection
            normalized_headers = [h.lower() for h in headers]

            # Keywords to detect fields in headers
            field_keywords = {
                'title': ['título', 'titulo', 'title', 'nombre', 'product_name', 'name'],
                'price': ['precio', 'price', 'cost', 'valor', 'costo'],
                'stock': ['stock', 'cantidad', 'qty', 'quantity'],
                'category': ['categoria', 'categoría', 'category', 'tipo'],
                'description': ['descripcion', 'description', 'detalle'],
            }

            detected_idx = {}
            for idx, h in enumerate(normalized_headers):
                for field, keywords in field_keywords.items():
                    if any(kw in h for kw in keywords):
                        if field not in detected_idx:
                            detected_idx[field] = idx

            # Helper to resolve index for a logical field (fallback to numeric positions used previously)
            def preview_idx_for(field_name, fallback_pos=None):
                if field_name in detected_idx:
                    i = detected_idx[field_name]
                    return i if i < len(headers) else None
                if fallback_pos is not None and fallback_pos < len(headers):
                    return fallback_pos
                return None
            
            # Generate preview data (first 5 products)
            preview_rows = []
            preview_limit = min(5, len(product_df))
            
            for product_idx in range(preview_limit):
                product_row = product_df.iloc[product_idx]
                preview_row = {}
                
                # Start with empty ML template row
                ml_row_data = [""] * len(headers)

                # Apply field mappings using detected indices where possible
                for ml_field, product_field in mapping.items():
                    if product_field in product_row and pd.notna(product_row[product_field]):
                        value = str(product_row[product_field])
                        key = ml_field.lower()

                        if 'título' in key or 'title' in key or 'nombre' in key:
                            idx = preview_idx_for('title', fallback_pos=1)
                            if idx is not None and idx < len(ml_row_data):
                                ml_row_data[idx] = value
                        elif 'precio' in key or 'price' in key:
                            idx = preview_idx_for('price', fallback_pos=2)
                            if idx is not None and idx < len(ml_row_data):
                                ml_row_data[idx] = value
                        elif 'stock' in key or 'cantidad' in key:
                            idx = preview_idx_for('stock', fallback_pos=3)
                            if idx is not None and idx < len(ml_row_data):
                                ml_row_data[idx] = value
                        elif 'categor' in key or 'category' in key or 'tipo' in key:
                            idx = preview_idx_for('category', fallback_pos=4)
                            if idx is not None and idx < len(ml_row_data):
                                ml_row_data[idx] = value
                        elif 'descrip' in key or 'description' in key:
                            idx = preview_idx_for('description', fallback_pos=5)
                            if idx is not None and idx < len(ml_row_data):
                                ml_row_data[idx] = value
                
                # Apply default settings
                if settings:
                    idx = preview_idx_for('condition', fallback_pos=6)
                    if 'condition' in settings and idx is not None and idx < len(ml_row_data):
                        condition_map = {'new': 'Nuevo', 'used': 'Usado', 'refurbished': 'Reacondicionado'}
                        ml_row_data[idx] = condition_map.get(settings['condition'], 'Nuevo')

                    idx = preview_idx_for('currency', fallback_pos=7)
                    if 'currency' in settings and idx is not None and idx < len(ml_row_data):
                        ml_row_data[idx] = settings['currency']

                    idx = preview_idx_for('free_shipping', fallback_pos=8)
                    if 'free_shipping' in settings and idx is not None and idx < len(ml_row_data):
                        ml_row_data[idx] = 'Sí' if settings['free_shipping'] == 'yes' else 'No'

                    idx = preview_idx_for('accepts_mercado_pago', fallback_pos=9)
                    if 'accepts_mercado_pago' in settings and idx is not None and idx < len(ml_row_data):
                        ml_row_data[idx] = 'Sí' if settings['accepts_mercado_pago'] == 'yes' else 'No'

                    idx = preview_idx_for('pickup_allowed', fallback_pos=10)
                    if 'pickup_allowed' in settings and idx is not None and idx < len(ml_row_data):
                        ml_row_data[idx] = 'Sí' if settings['pickup_allowed'] == 'yes' else 'No'
                
                # Convert to dictionary for JSON response
                row_dict = {}
                for i, header in enumerate(headers):
                    row_dict[header] = ml_row_data[i] if i < len(ml_row_data) else ""
                
                preview_rows.append({
                    'row_number': product_idx + 1,
                    'source_data': product_row.to_dict(),
                    'mapped_data': row_dict
                })
            
            # Clean up temporary files
            import shutil
            shutil.rmtree(temp_dir)

            return {
                "status": "success",
                "message": f"Vista previa generada para {preview_limit} productos",
                "preview": {
                    "headers": headers,
                    "rows": preview_rows,
                    "total_products": len(product_df),
                    "preview_count": preview_limit,
                    # expose 1-based Excel row to callers (more user-friendly)
                    "data_start_row": data_start_row + 1
                },
                "mapping_info": {
                    "applied_mappings": mapping,
                    "default_settings": settings,
                    "mapped_fields_count": len(mapping)
                },
                "file_info": {
                    "ml_template": ml_template.filename,
                    "product_data": product_data.filename,
                    "estimated_output_rows": len(product_df)
                }
            }
            
        except Exception as processing_error:
            # Clean up temporary files
            import shutil
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            raise HTTPException(
                status_code=500,
                detail=f"Error generando vista previa: {str(processing_error)}"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en vista previa ML: {str(e)}"
        )

# ML File Generation endpoint
@app.post("/api/files/generate-ml-file")
async def generate_ml_file_endpoint(
    ml_template: UploadFile = File(...),
    product_data: UploadFile = File(...),
    mapping_config: str = Form(...),
    default_settings: str = Form(...),
    write_mode: str = Form('fill-empty'),  # allowed: fill-empty, append, interactive, overwrite
    edits: str = Form(None)  # optional JSON list of edits when write_mode=interactive or overwrite
):
    """Generate final ML file with mapped product data"""
    try:
        import json
        from datetime import datetime
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        import tempfile
        
        # Parse configuration data
        try:
            mapping = json.loads(mapping_config)
            settings = json.loads(default_settings)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error parsing configuration: {str(e)}"
            )
        
        # Validate file types
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        ml_ext = os.path.splitext(ml_template.filename)[1].lower()
        product_ext = os.path.splitext(product_data.filename)[1].lower()
        
        if ml_ext not in allowed_extensions or product_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail="Tipos de archivo no soportados. Use: .xlsx, .xls, .csv"
            )
        
        # Save temporary files
        temp_dir = tempfile.mkdtemp()
        ml_temp_path = os.path.join(temp_dir, f"ml_template_{ml_template.filename}")
        product_temp_path = os.path.join(temp_dir, f"product_data_{product_data.filename}")
        
        # Save uploaded files
        with open(ml_temp_path, "wb") as buffer:
            ml_content = await ml_template.read()
            buffer.write(ml_content)
            
        with open(product_temp_path, "wb") as buffer:
            product_content = await product_data.read()
            buffer.write(product_content)
        
        try:
            # Read ML template (preserving structure)
            ml_df = pd.read_excel(ml_temp_path, header=None) if ml_ext in ['.xlsx', '.xls'] else pd.read_csv(ml_temp_path, header=None)
            
            # Read product data
            if product_ext in ['.xlsx', '.xls']:
                product_df = pd.read_excel(product_temp_path)
            else:
                product_df = pd.read_csv(product_temp_path)
            
            # Find the data start row in ML template (usually around row 8)
            # Use zero-based index. Default to 7 so data rows start at Excel row 8.
            data_start_row = 7  # Default for ML templates (zero-based index -> Excel row 8)
            for i in range(min(15, len(ml_df))):
                row = ml_df.iloc[i]
                non_null_count = row.count()
                if non_null_count >= len(ml_df.columns) * 0.5:
                    data_start_row = i
                    break
            
            # Build final rows in memory instead of assigning with iloc repeatedly.
            # This avoids shape/resize issues with pandas when expanding DataFrames.
            processed_count = 0

            # Resolve mapping values using shared resolver
            mapping_resolved = {}
            try:
                import sys
                import importlib.util
                src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src'))
                if src_dir not in sys.path:
                    sys.path.insert(0, src_dir)
                
                # Use importlib to safely import mapping_loader
                spec = importlib.util.spec_from_file_location("mapping_loader", os.path.join(src_dir, "mapping_loader.py"))
                mapping_loader_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mapping_loader_module)

                product_cols = list(product_df.columns)
                mapping_resolved = mapping_loader_module.resolve_column_aliases(product_cols, mapping)
            except Exception:
                mapping_resolved = {k: (v if isinstance(v, str) and v in list(product_df.columns) else None) for k, v in mapping.items()}

            # Minimum expected columns for ML template output
            min_expected_cols = 12
            num_cols = max(len(ml_df.columns), min_expected_cols)

            # Helper to normalize a row (list) to have length num_cols
            def normalize_row_list(row_list):
                row = list(row_list) if row_list is not None else []
                if len(row) < num_cols:
                    row = row + [None] * (num_cols - len(row))
                elif len(row) > num_cols:
                    row = row[:num_cols]
                return row

            # Header rows: keep everything before data_start_row
            header_rows = [normalize_row_list(ml_df.iloc[i].tolist()) for i in range(min(data_start_row, len(ml_df)))]

            # Base template row for data (if exists use data_start_row row, else use empty)
            if data_start_row < len(ml_df):
                base_template_row = normalize_row_list(ml_df.iloc[data_start_row].tolist())
            else:
                base_template_row = [None] * num_cols

            # Footer rows: keep rows after the data_start_row (they will be appended after product rows)
            footer_rows = []
            if data_start_row + 1 < len(ml_df):
                footer_rows = [normalize_row_list(ml_df.iloc[i].tolist()) for i in range(data_start_row + 1, len(ml_df))]

            # Build product rows
            import logging
            logging.basicConfig(level=logging.INFO)
            logging.info("generate-ml-file: num_cols=%s, mapping_keys=%s, product_shape=%s", num_cols, list(mapping.keys()), getattr(product_df, 'shape', None))
            data_rows = []

            # Detect column indices from ML template headers for robust mapping.
            # Build headers similarly to preview logic (header row is the row before data_start_row)
            headers = []
            if data_start_row > 0 and data_start_row - 1 < len(ml_df):
                header_row = ml_df.iloc[data_start_row - 1]
                headers = [str(cell).strip().lower() if pd.notna(cell) else f"columna_{i+1}" for i, cell in enumerate(header_row)]
            else:
                headers = [f"columna_{i+1}" for i in range(num_cols)]

            # Keyword sets to detect common ML fields in the template headers
            field_keywords = {
                'title': ['título', 'titulo', 'title', 'nombre', 'nombre_producto'],
                'price': ['precio', 'price', 'cost', 'valor', 'costo'],
                'stock': ['stock', 'cantidad', 'qty', 'quantity'],
                'category': ['categoria', 'categoría', 'category', 'tipo'],
                'condition': ['condición', 'condicion', 'condition'],
                'currency': ['moneda', 'currency'],
                'free_shipping': ['envio_gratis', 'free_shipping', 'envío_gratis', 'envio gratis'],
                'accepts_mercado_pago': ['mercado_pago', 'accepts_mercado_pago', 'acepta_mercadopago']
            }

            # Detect indices by scanning headers for keywords
            detected_idx = {}
            for idx, h in enumerate(headers):
                for field, keywords in field_keywords.items():
                    if any(kw in h for kw in keywords):
                        # only set the first time a field is found
                        if field not in detected_idx:
                            detected_idx[field] = idx

            # Fallback default positions if not detected (keeps previous assumptions but within bounds)
            fallback_positions = {
                'title': 1,
                'price': 2,
                'stock': 3,
                'category': 4,
                'condition': 5,
                'currency': 6,
                'free_shipping': 7,
                'accepts_mercado_pago': 8
            }

            # Final column index resolver
            def idx_for(field_name):
                # prefer detected index
                if field_name in detected_idx:
                    i = detected_idx[field_name]
                    return i if i < num_cols else None
                # fallback
                i = fallback_positions.get(field_name)
                if i is None:
                    return None
                return i if i < num_cols else None
            # iterate product rows and apply mappings
            for product_idx, product_row in product_df.iterrows():
                # Start from base template for each product
                out_row = base_template_row.copy()

                # Apply field mappings
                for ml_field, product_field in mapping.items():
                    try:
                        # resolve to actual product column (if we resolved earlier)
                        resolved_col = mapping_resolved.get(ml_field, product_field)
                        if not resolved_col:
                            # nothing resolved for this mapping
                            continue
                        if resolved_col in product_row and pd.notna(product_row[resolved_col]):
                            value = product_row[resolved_col]
                        else:
                            # If not present as column, skip
                            continue
                    except Exception:
                        continue

                    key = ml_field.lower()
                    # Resolve target index using detected headers or fallback positions
                    if 'título' in key or 'title' in key or 'nombre' in key:
                        idx = idx_for('title')
                        if idx is not None:
                            out_row[idx] = value
                    elif 'precio' in key or 'price' in key:
                        idx = idx_for('price')
                        if idx is not None:
                            out_row[idx] = value
                    elif 'stock' in key or 'cantidad' in key:
                        idx = idx_for('stock')
                        if idx is not None:
                            out_row[idx] = value
                    elif 'categor' in key or 'category' in key or 'tipo' in key:
                        idx = idx_for('category')
                        if idx is not None:
                            out_row[idx] = value

                # Apply default settings to the out_row
                if settings:
                    if 'condition' in settings:
                        condition_map = {'new': 'Nuevo', 'used': 'Usado', 'refurbished': 'Reacondicionado'}
                        idx = idx_for('condition')
                        if idx is not None:
                            out_row[idx] = condition_map.get(settings['condition'], 'Nuevo')

                    if 'currency' in settings:
                        idx = idx_for('currency')
                        if idx is not None:
                            out_row[idx] = settings['currency']

                    if 'free_shipping' in settings:
                        idx = idx_for('free_shipping')
                        if idx is not None:
                            out_row[idx] = 'Sí' if settings['free_shipping'] == 'yes' else 'No'

                    if 'accepts_mercado_pago' in settings:
                        idx = idx_for('accepts_mercado_pago')
                        if idx is not None:
                            out_row[idx] = 'Sí' if settings['accepts_mercado_pago'] == 'yes' else 'No'

                data_rows.append(out_row)
                processed_count += 1

            # Use the template_filler utility to write into the template starting at
            # Excel row 8 and produce a fill report. This preserves rows 1..7 and
            # writes only detected logical columns.
            import sys
            import importlib.util
            src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src'))
            if src_dir not in sys.path:
                sys.path.insert(0, src_dir)

            try:
                # Use importlib to safely import template_filler
                spec = importlib.util.spec_from_file_location("template_filler", os.path.join(src_dir, "template_filler.py"))
                template_filler_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(template_filler_module)

                # Build products_data list using the mapping provided by the wizard.
                products_data = []
                for _, product_row in product_df.iterrows():
                    prod = {}
                    for ml_field, product_field in mapping.items():
                        try:
                            if product_field in product_row and pd.notna(product_row[product_field]):
                                value = product_row[product_field]
                            else:
                                continue
                        except Exception:
                            continue

                        key = ml_field.lower()
                        if 'titul' in key or 'title' in key or 'nombre' in key:
                            prod['title'] = value
                        elif 'sku' in key:
                            prod['sku'] = value
                        elif 'precio' in key or 'price' in key:
                            prod['price'] = value
                        elif 'envio' in key or 'shipping' in key:
                            prod['shipping'] = value
                        elif 'stock' in key or 'cantidad' in key:
                            prod['stock'] = value
                        elif 'cond' in key or 'condition' in key:
                            prod['condition'] = value
                        elif 'image' in key or 'imagen' in key:
                            prod['images'] = value

                    products_data.append(prod)

                # Prepare default values from settings (wizard)
                default_values = {}
                if settings:
                    if 'condition' in settings:
                        condition_map = {'new': 'Nuevo', 'used': 'Usado', 'refurbished': 'Reacondicionado'}
                        default_values['condition'] = condition_map.get(settings['condition'], 'Nuevo')
                    # If wizard indicates free shipping, map to ML fulfillment code 'me2' by default
                    if settings.get('free_shipping') in ('yes', True) and 'shipping' not in default_values:
                        default_values['shipping'] = 'me2'
                    if 'default_stock' in settings:
                        default_values['stock'] = settings['default_stock']
                    else:
                        # sensible default when not provided
                        default_values.setdefault('stock', 1)

                # Load workbook and sheet
                from openpyxl import load_workbook
                wb = load_workbook(ml_temp_path)
                ws = wb.active

                # Enforce writing starts at Excel row 8 by default
                start_row_excel = max(8, data_start_row + 1)

                # Parse edits payload if provided
                edits_payload = None
                try:
                    if edits:
                        import json as _json
                        edits_payload = _json.loads(edits)
                except Exception:
                    raise HTTPException(status_code=400, detail='Invalid edits JSON payload')

                # Validate write_mode
                write_mode = (write_mode or 'fill-empty').lower()
                if write_mode not in ('fill-empty', 'append', 'interactive', 'overwrite'):
                    raise HTTPException(status_code=400, detail='Invalid write_mode')

                # Prepare behavior based on write_mode
                # - fill-empty: default safe behavior, do not overwrite existing cells
                # - append: find first empty row >=8 and append products sequentially
                # - interactive: apply only user-provided edits; do not bulk write products
                # - overwrite: overwrite target cells (server will create a backup copy first)

                # For overwrite mode, create a backup copy of the uploaded template
                backup_path = None
                if write_mode == 'overwrite':
                    import shutil
                    backup_filename = f"backup_{ml_template.filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    backup_path = os.path.join(UPLOAD_DIR, backup_filename)
                    shutil.copy(ml_temp_path, backup_path)

                # append mode: find first row with an empty 'title' or other primary column
                if write_mode == 'append':
                    # detect columns and find first empty target row at or after row 8
                    cols = template_filler_module.detect_columns(ws, header_row=8)
                    # prefer title column, else first detected
                    title_col = cols.get('title') or (min(cols.values()) if cols else 1)
                    append_row = 8
                    for r in range(8, ws.max_row + 2):
                        cell = ws.cell(row=r, column=title_col)
                        if cell.value in (None, ''):
                            append_row = r
                            break
                    # write products starting at append_row
                    filled_report, skipped_report = template_filler_module.fill_products_from_row(
                        ws,
                        start_row=append_row,
                        products_data=products_data,
                        default_values=default_values,
                        overwrite=False
                    )
                elif write_mode == 'interactive':
                    # Apply only edits provided by user
                    filled_report = []
                    skipped_report = []
                    if not edits_payload:
                        raise HTTPException(status_code=400, detail='No edits provided for interactive mode')
                    # edits expected as list of {row: int, field: str, value: any}
                    cols = template_filler_module.detect_columns(ws, header_row=8)
                    for e in edits_payload:
                        try:
                            row = int(e.get('row'))
                            field = e.get('field')
                            value = e.get('value')
                        except Exception:
                            continue
                        # validate row
                        if row < 8:
                            skipped_report.append({'row': row, 'skipped': {'reason': 'row_below_header'}})
                            continue
                        if field not in cols:
                            skipped_report.append({'row': row, 'skipped': {'reason': 'unknown_field', 'field': field}})
                            continue
                        col_idx = cols[field]
                        target_cell = ws.cell(row=row, column=col_idx)
                        # write value
                        target_cell.value = value
                        filled_report.append({'row': row, 'filled': {field: value}})
                else:
                    # fill-empty (default) or overwrite
                    allow_overwrite = True if write_mode == 'overwrite' else False
                    filled_report, skipped_report = template_filler_module.fill_products_from_row(
                        ws,
                        start_row=8,
                        products_data=products_data,
                        default_values=default_values,
                        overwrite=allow_overwrite
                    )

                # Save result
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"ml_productos_mapeados_{timestamp}.xlsx"
                output_path = os.path.join(UPLOAD_DIR, output_filename)
                wb.save(output_path)

                # Clean up temporary files
                import shutil
                shutil.rmtree(temp_dir)

                # Build fill summary
                fill_summary = template_filler_module.generate_fill_report(products_data, filled_report, skipped_report)

                download_url = f"/api/files/download/{output_filename}"

                return {
                    "status": "success",
                    "message": f"Archivo ML generado exitosamente: {len(products_data)} productos procesados",
                    "file_info": {
                        "filename": output_filename,
                        "original_ml_template": ml_template.filename,
                        "original_product_data": product_data.filename,
                        "products_processed": len(products_data),
                        "mapping_fields": len(mapping),
                        "generation_timestamp": timestamp
                    },
                    "download": {
                        "url": download_url,
                        "filename": output_filename,
                        "ready": True
                    },
                    "summary": {
                        "total_products": len(products_data),
                        "mapped_fields": list(mapping.keys()),
                        "default_settings_applied": list(default_values.keys()),
                        "ml_template_preserved": True,
                        "fill_report": fill_summary
                    }
                }

            except Exception as exc:
                # If the structured filler is not available or fails, fallback to original write behavior
                import logging
                logging.error("template_filler integration failed, falling back to legacy writer: %s", str(exc))

                # Fallback: write data_rows as previously implemented
                from openpyxl import load_workbook
                wb = load_workbook(ml_temp_path)
                ws = wb.active

                start_row_excel = data_start_row + 1
                if start_row_excel < 8:
                    start_row_excel = 8

                for i, out_row in enumerate(data_rows):
                    target_row = start_row_excel + i
                    for col_idx in range(min(len(out_row), num_cols)):
                        try:
                            value = out_row[col_idx]
                            ws.cell(row=target_row, column=col_idx + 1, value=value)
                        except Exception:
                            continue

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"ml_productos_mapeados_{timestamp}.xlsx"
                output_path = os.path.join(UPLOAD_DIR, output_filename)
                wb.save(output_path)

                import shutil
                shutil.rmtree(temp_dir)

                download_url = f"/api/files/download/{output_filename}"

                return {
                    "status": "success",
                    "message": f"Archivo ML generado exitosamente (legacy): {processed_count} productos procesados",
                    "file_info": {
                        "filename": output_filename,
                        "original_ml_template": ml_template.filename,
                        "original_product_data": product_data.filename,
                        "products_processed": processed_count,
                        "mapping_fields": len(mapping),
                        "generation_timestamp": timestamp
                    },
                    "download": {
                        "url": download_url,
                        "filename": output_filename,
                        "ready": True
                    },
                    "summary": {
                        "total_products": processed_count,
                        "mapped_fields": list(mapping.keys()),
                        "default_settings_applied": list(settings.keys()) if settings else [],
                        "ml_template_preserved": True,
                        "fallback_mode": True
                    }
                }
            
        except Exception as processing_error:
            # Clean up temporary files
            import shutil, traceback, logging
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

            tb = traceback.format_exc()
            # Log the full traceback for debugging
            logging.error("Error generating ML file:\n%s", tb)

            # Return a shorter message but include that full traceback was logged
            # For debugging, attach some diagnostics
            try:
                prod_shape = getattr(product_df, 'shape', None)
            except Exception:
                prod_shape = None

            # Log diagnostics server-side but do not expose sensitive internals in the HTTP response
            import logging
            logging.error("generate-ml-file error: %s", str(processing_error))
            logging.error("diagnostics num_cols=%s mapping_keys=%s product_shape=%s", locals().get('num_cols', None), (len(mapping) if isinstance(mapping, dict) else None), prod_shape)

            raise HTTPException(
                status_code=500,
                detail="Error procesando archivos en el servidor. Consulte los logs para más detalles."
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generando archivo ML: {str(e)}"
        )

# File Download endpoint
@app.get("/api/files/download/{filename}")
async def download_file_endpoint(filename: str):
    """Download generated ML file"""
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=404,
                detail="Archivo no encontrado"
            )
        
        from fastapi.responses import FileResponse
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error descargando archivo: {str(e)}"
        )

# Authentication endpoints
@app.post("/auth/login", response_model=AuthResponse)
async def login(request: LoginRequest):
    """Demo login endpoint"""
    try:
        # Check if user exists
        if request.email not in DEMO_USERS:
            return AuthResponse(
                success=False,
                message="Email no encontrado"
            )
        
        user = DEMO_USERS[request.email]
        
        # Check password
        if request.password != user["password"]:
            return AuthResponse(
                success=False,
                message="Contraseña incorrecta"
            )
        
        # Generate demo token
        demo_token = f"demo_token_{user['id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Return user data without password
        user_data = {k: v for k, v in user.items() if k != "password"}
        
        return AuthResponse(
            success=True,
            message=f"Bienvenido {user['name']}!",
            user=user_data,
            token=demo_token
        )
        
    except Exception as e:
        return AuthResponse(
            success=False,
            message=f"Error de login: {str(e)}"
        )

@app.get("/auth/me")
async def get_current_user():
    """Get current user info (demo endpoint)"""
    return {
        "user": {
            "id": 1,
            "email": "demo@test.com",
            "name": "Demo User",
            "account_type": "premium"
        },
        "authenticated": True
    }

@app.post("/auth/logout")
async def logout():
    """Demo logout endpoint"""
    return {"success": True, "message": "Sesión cerrada exitosamente"}

# Frontend-compatible API endpoints (matching React frontend expectations)
@app.post("/api/login", response_model=AuthResponse)
async def api_login(request: LoginRequest):
    """Frontend-compatible login endpoint"""
    return await login(request)

@app.get("/api/user")
async def api_get_current_user():
    """Frontend-compatible current user endpoint"""
    return await get_current_user()

@app.post("/api/logout")
async def api_logout():
    """Frontend-compatible logout endpoint"""
    return await logout()

if __name__ == "__main__":
    logger.info("Starting ML Extractor Backend - FastAPI")
    logger.info("Documentation available at: http://localhost:8009/docs")
    logger.info("Health check at: http://localhost:8009/health")
    
    # Start the server
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8009,
        reload=False,
        log_level="info"
    )
