#!/usr/bin/env python3

import os
import sys

print("=== DEBUG TEST ===")
print(f"Current directory: {os.getcwd()}")
print(f"Python executable: {sys.executable}")
print(f"Python version: {sys.version}")

# Test file existence
files_to_check = [
    'uploads/Publicar-08-09-10_39_33.xlsx',
    'uploads/Shimano_Julio.pdf',
    'config/mapping.yaml'
]

for file_path in files_to_check:
    exists = os.path.exists(file_path)
    print(f"File {file_path}: {'EXISTS' if exists else 'NOT FOUND'}")

# Test imports
try:
    import openpyxl
    print("openpyxl: OK")
except ImportError as e:
    print(f"openpyxl: ERROR - {e}")

try:
    import yaml
    print("yaml: OK")
except ImportError as e:
    print(f"yaml: ERROR - {e}")

try:
    import docx
    print("docx: OK")
except ImportError as e:
    print(f"docx: ERROR - {e}")

try:
    import PyPDF2
    print("PyPDF2: OK")
except ImportError as e:
    print(f"PyPDF2: ERROR - {e}")

# Test file reading
try:
    print("\n=== TESTING FILE READING ===")
    
    # Test Excel reading
    if os.path.exists('uploads/Publicar-08-09-10_39_33.xlsx'):
        print("Testing Excel file...")
        wb = openpyxl.load_workbook('uploads/Publicar-08-09-10_39_33.xlsx')
        ws = wb.active
        print(f"Excel sheets: {wb.sheetnames}")
        print(f"Active sheet max row: {ws.max_row}")
        print(f"Active sheet max col: {ws.max_column}")
        
        # Get first few cells
        print("First row values:")
        for i, cell in enumerate(ws[1]):
            if i < 5:  # First 5 columns
                print(f"  Col {i+1}: {cell.value}")
    
    # Test config reading
    if os.path.exists('config/mapping.yaml'):
        print("\nTesting YAML config...")
        with open('config/mapping.yaml', 'r') as f:
            config = yaml.safe_load(f)
        print(f"Config loaded: {config is not None}")
        if config:
            print(f"Config keys: {list(config.keys())}")
            
except Exception as e:
    print(f"ERROR in file testing: {e}")
    import traceback
    traceback.print_exc()

print("\n=== TEST COMPLETE ===")
