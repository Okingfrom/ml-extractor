import pandas as pd
import openpyxl
import os

p = os.path.join(os.path.dirname(__file__), '..', 'samples', 'output_generated_ml.xlsx')
print('File:', os.path.abspath(p))
wb = openpyxl.load_workbook(p, data_only=True)
sheet = wb.sheetnames[0]
print('Sheet:', sheet)
print('\nFirst 10 raw rows (openpyxl values):')
for i, row in enumerate(wb[sheet].iter_rows(min_row=1, max_row=10, values_only=True), start=1):
    print(f'{i:2d}:', row)
print('\n--- pandas preview ---')
df = pd.read_excel(p, sheet_name=0)
print('Columns:', list(df.columns))
print(df.head(5).to_dict(orient='records'))
print('\nTotal rows (pandas):', len(df))
