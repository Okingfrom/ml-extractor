from openpyxl import load_workbook
import json, os, sys

# Usage: python inspect_template_rows.py [path/to/template.xlsx]
# Defaults to the common Mercado Libre template file if present.
default_candidates = [
    os.path.join('samples', 'plantilla_ml_ejemplo.xlsx'),
    os.path.join('samples', 'sample_input.xlsx'),
    os.path.join('samples', 'sample_output.xlsx')
]

given = sys.argv[1] if len(sys.argv) > 1 else None
if given:
    p = os.path.abspath(given)
else:
    # pick the first candidate that exists
    p = None
    for c in default_candidates:
        if os.path.exists(c):
            p = os.path.abspath(c)
            break
    if p is None:
        # fallback to current samples/sample_input.xlsx if nothing found
        p = os.path.abspath(os.path.join('samples', 'sample_input.xlsx'))

print('Inspecting:', p)
wb = load_workbook(p)
ws = wb.active
rows = []
for i in range(1, min(13, ws.max_row+1)):
    row = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column+1)]
    rows.append(row)
# Print only first 12 rows with non-empty cells per row for readability
filtered = [[c for c in row if c is not None] for row in rows]
print(json.dumps(filtered, ensure_ascii=False, indent=2))
