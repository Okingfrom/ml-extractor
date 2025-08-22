from openpyxl import load_workbook
wb = load_workbook('..\\samples\\sample_input.xlsx')
ws = wb.active
for r in range(1,13):
    vals = [str(ws.cell(row=r, column=c).value) if ws.cell(row=r, column=c).value is not None else '' for c in range(1,21)]
    print(f'Row {r}:', vals)
