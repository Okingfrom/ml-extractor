"""
ML Mapper and validation utilities for Mercado Libre massive templates.

Provides:
- check_sheet_exists(sheet_name)
- validate_columns(sheet, required_columns, optional_columns)
- validate_image_urls(url_list)
- validate_price_format(price_str)
- group_variations(rows)

The module focuses on validating two sheets: "Publicaciones" and "Variaciones",
and produces a structured validation report (errors + warnings).
"""
from typing import List, Dict, Tuple, Any
from openpyxl import load_workbook
import re
import requests
import time
import logging

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def check_sheet_exists(wb, sheet_name: str) -> bool:
    """Return True if sheet exists in workbook object or file path.

    wb may be either an openpyxl.Workbook or a path string.
    """
    if isinstance(wb, str):
        wb_obj = load_workbook(wb, read_only=True, data_only=True)
        exists = sheet_name in wb_obj.sheetnames
        wb_obj.close()
        return exists

    # workbook object
    return sheet_name in wb.sheetnames


def validate_columns(sheet, required_columns: List[str], optional_columns: List[str]) -> Tuple[List[str], List[str]]:
    """Validate presence of required and optional columns in an openpyxl worksheet.

    Returns (missing_required, missing_optional_found_missing)
    """
    # Get header row: find the first row with non-empty cells
    header = None
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            break

    if header is None:
        # no header found
        return required_columns.copy(), optional_columns.copy()

    lowered = [h.lower() for h in header]

    missing_required = []
    for req in required_columns:
        synonyms = [req]
        if '/' in req:
            synonyms.extend([s.strip() for s in req.split('/')])
        found = False
        for s in synonyms:
            if any(s.lower() == h or s.lower() in h for h in lowered):
                found = True
                break
        if not found:
            missing_required.append(req)

    missing_optional = []
    for opt in optional_columns:
        if not any(opt.lower() == h or opt.lower() in h for h in lowered):
            missing_optional.append(opt)

    return missing_required, missing_optional


def validate_image_urls(url_list: List[str], timeout: float = 3.0) -> Dict[str, Dict[str, Any]]:
    """Check list of image URLs, return dict with status and http code or error.

    This makes HEAD requests where possible; falls back to GET if HEAD is not allowed.
    """
    results: Dict[str, Dict[str, Any]] = {}
    session = requests.Session()
    session.headers.update({"User-Agent": "ml-mapper-validator/1.0"})

    for url in url_list:
        url = str(url).strip()
        if not url:
            results[url] = {"ok": False, "reason": "empty"}
            continue
        try:
            # prefer HEAD for speed
            resp = session.head(url, timeout=timeout, allow_redirects=True)
            code = resp.status_code
            if code == 405:
                # some servers reject HEAD
                resp = session.get(url, stream=True, timeout=timeout)
                code = resp.status_code

            results[url] = {"ok": code == 200, "status_code": code}
        except Exception as e:
            results[url] = {"ok": False, "reason": str(e)}

        # be polite
        time.sleep(0.05)

    return results


PRICE_RX = re.compile(r"^\d{1,13}(?:\.\d{1,2})?$")


def validate_price_format(price_str: Any) -> Tuple[bool, float]:
    """Validate price format like 1999.99 and return (valid, numeric_value).

    Accepts numeric types and strings. Returns (False, 0.0) on invalid.
    """
    if price_str is None:
        return False, 0.0
    try:
        if isinstance(price_str, (int, float)):
            value = float(price_str)
            if value < 0:
                return False, 0.0
            return True, value

        s = str(price_str).strip()
        s = s.replace(',', '')
        if PRICE_RX.match(s):
            return True, float(s)
        value = float(s)
        return True, value
    except Exception:
        return False, 0.0


def group_variations(rows: List[Dict[str, Any]], variation_id_field: str = 'ID de Variación') -> Dict[str, List[Dict[str, Any]]]:
    """Group rows by variation id when present.

    Returns a dict mapping variation_id -> list of rows.
    If variation_id missing, groups under a generated parent key.
    """
    groups: Dict[str, List[Dict[str, Any]]] = {}
    fallback_counter = 0
    for row in rows:
        vid = None
        for key in [variation_id_field, 'variation_id', 'id_variacion', 'id de variacion']:
            if key in row and row[key] not in (None, ''):
                vid = str(row[key]).strip()
                break
        if not vid:
            vid = f"__no_vid_{fallback_counter}"
            fallback_counter += 1

        groups.setdefault(vid, []).append(row)

    return groups


class MLMapperValidator:
    """High-level validator to check whole workbook for required structure and content."""

    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb = load_workbook(workbook_path, data_only=True)
        self.report: Dict[str, Any] = {"errors": [], "warnings": [], "info": {}}

    def check_sheet_exists(self, sheet_name: str) -> bool:
        exists = sheet_name in self.wb.sheetnames
        if not exists:
            self.report['errors'].append(f"Sheet not found: {sheet_name}")
        return exists

    def _sheet_to_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        ws = self.wb[sheet_name]
        rows = list(ws.values)
        if not rows:
            return []
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        dicts = []
        for r in rows[1:]:
            d = {}
            for i, h in enumerate(header):
                d[h] = r[i] if i < len(r) else None
            dicts.append(d)
        return dicts

    def validate(self) -> Dict[str, Any]:
        # Define required and optional columns (allow multiple sensible names)
        pubs_required = ['ID de Variación', 'SKU', 'Categoría', 'Título', 'Condición', 'Precio', 'Cantidad']
        pubs_optional = ['URLs de Imágenes', 'Marca', 'Modelo', 'EAN', 'UPC', 'Garantía', 'Tipo de Publicación', 'Envío']

        vars_required = ['ID de Variación', 'SKU', 'Stock', 'Precio']
        vars_optional = ['Atributos', 'URLs de Imágenes']

        # Check sheets
        if not self.check_sheet_exists('Publicaciones') or not self.check_sheet_exists('Variaciones'):
            self.report['info']['sheets'] = self.wb.sheetnames
            return self.report

        pubs = self._sheet_to_dicts('Publicaciones')
        vars_ = self._sheet_to_dicts('Variaciones')

        # Validate columns using first row headers
        ws_pubs = self.wb['Publicaciones']
        missing_pub_required, missing_pub_optional = validate_columns(ws_pubs, pubs_required, pubs_optional)
        if missing_pub_required:
            self.report['errors'].append(f"Publicaciones: columnas requeridas faltantes: {missing_pub_required}")

        ws_vars = self.wb['Variaciones']
        missing_var_required, missing_var_optional = validate_columns(ws_vars, vars_required, vars_optional)
        if missing_var_required:
            self.report['errors'].append(f"Variaciones: columnas requeridas faltantes: {missing_var_required}")

        # If there are blocking errors, stop
        if self.report['errors']:
            return self.report

        # Further validations
        # 1) Validate Publicaciones rows: price and quantity
        for idx, row in enumerate(pubs, start=2):
            # price
            p = None
            for key in ['Precio', 'Price', 'precio', 'price']:
                if key in row:
                    p = row[key]
                    break
            valid_price, numeric_price = validate_price_format(p)
            if not valid_price:
                self.report['warnings'].append(f"Publicaciones fila {idx}: precio inválido -> {p}")

            # cantidad
            q = None
            for key in ['Cantidad', 'cantidad', 'Stock', 'stock']:
                if key in row:
                    q = row[key]
                    break
            try:
                if q is None or int(float(q)) != float(q):
                    raise ValueError()
            except Exception:
                self.report['warnings'].append(f"Publicaciones fila {idx}: cantidad inválida -> {q}")

        # 2) Group variations and validate SKUs uniqueness per parent
        var_groups = group_variations(vars_, variation_id_field='ID de Variación')
        for vid, group_rows in var_groups.items():
            skus = [str(r.get('SKU') or r.get('sku') or '').strip() for r in group_rows]
            if len(skus) != len(set(skus)):
                self.report['warnings'].append(f"Variaciones grupo {vid}: SKUs repetidos dentro del grupo")

            # price consistency
            prices = []
            for r in group_rows:
                for k in ['Precio', 'Price', 'precio', 'price']:
                    if k in r and r[k] not in (None, ''):
                        prices.append(r[k])
                        break
            validated = [validate_price_format(p)[1] for p in prices if validate_price_format(p)[0]]
            if validated and len(set(validated)) > 1:
                self.report['errors'].append(f"Variaciones grupo {vid}: precios inconsistentes entre variaciones: {set(validated)}")

        # 3) Validate image URLs in both sheets if present
        def collect_image_urls(rows, keys_candidates):
            urls = []
            for r in rows:
                for k in keys_candidates:
                    if k in r and r[k]:
                        raw = str(r[k])
                        parts = [u.strip() for u in raw.split(',') if u.strip()]
                        urls.extend(parts)
            return urls

        image_keys = ['URLs de Imágenes', 'URLs', 'Images', 'Imagenes', 'imagenes', 'urls']
        pub_urls = collect_image_urls(pubs, image_keys)
        var_urls = collect_image_urls(vars_, image_keys)
        all_urls = list(dict.fromkeys(pub_urls + var_urls))
        if all_urls:
            url_results = validate_image_urls(all_urls)
            invalids = [u for u, r in url_results.items() if not r.get('ok')]
            if invalids:
                self.report['warnings'].append(f"URLs de imágenes inválidas detectadas: {len(invalids)} (ej: {invalids[:3]})")

        # Final report info
        self.report['info']['publicaciones_count'] = len(pubs)
        self.report['info']['variaciones_count'] = len(vars_)
        self.report['info']['variation_groups'] = len(var_groups)

        return self.report
