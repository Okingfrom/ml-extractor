"""
Mercado Libre Template Processor
Handles specific ML template structure and validation
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime

try:
    from .mapping_pattern_generator import generate_mapping_from_analysis
    MAPPING_GENERATOR_AVAILABLE = True
except ImportError:
    print("Warning: Mapping pattern generator not available")
    MAPPING_GENERATOR_AVAILABLE = False


@dataclass
class MLTemplateStructure:
    """Structure definition for ML templates"""
    obligatory_columns: Dict[str, str]  # column_name: data_type
    data_start_row: int = 8  # Row where product data starts (1-indexed)
    obligatory_row: int = 4  # Row with "obligatorio" markers (1-indexed)
    data_type_row: int = 5   # Row with data type definitions (1-indexed)
    

@dataclass
class MLTemplateAnalysis:
    """Analysis result of ML template"""
    is_ml_template: bool
    template_structure: Optional[MLTemplateStructure]
    detected_columns: Dict[str, str]
    validation_errors: List[str]
    recommendations: List[str]
    total_products: int
    sample_data: List[Dict[str, Any]]


class MLTemplateProcessor:
    """Processor for Mercado Libre templates"""
    
    # Known ML template column patterns
    ML_COLUMN_PATTERNS = {
        'titulo': ['título', 'title', 'nombre', 'product_name'],
        'precio': ['precio', 'price', 'cost', 'valor'],
        'stock': ['stock', 'cantidad', 'inventory', 'qty'],
        'categoria': ['categoría', 'category', 'cat'],
        'marca': ['marca', 'brand'],
        'modelo': ['modelo', 'model'],
        'descripcion': ['descripción', 'description', 'desc'],
        'peso': ['peso', 'weight'],
        'garantia': ['garantía', 'warranty'],
        'condicion': ['condición', 'condition'],
        'imagen_principal': ['imagen', 'image', 'foto', 'picture'],
        'imagen_1': ['imagen_1', 'image_1', 'foto_1'],
        'imagen_2': ['imagen_2', 'image_2', 'foto_2'],
        'imagen_3': ['imagen_3', 'image_3', 'foto_3'],
    }
    
    # Data type validation patterns
    DATA_TYPE_PATTERNS = {
        'numero': ['número', 'number', 'numeric', 'decimal', 'entero'],
        'texto': ['texto', 'text', 'string', 'cadena'],
        'fecha': ['fecha', 'date', 'datetime'],
        'boolean': ['sí/no', 'si/no', 'yes/no', 'true/false', 'boolean'],
        'url': ['url', 'link', 'enlace', 'http'],
        'email': ['email', 'correo', 'mail'],
    }

    def __init__(self):
        self.file_processor = None

    def analyze_ml_template(self, file_path: str) -> MLTemplateAnalysis:
        """
        Analyze uploaded file to determine if it's an ML template
        and extract its structure
        """
        try:
            # Determine file type and read accordingly
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension in ['.xlsx', '.xls']:
                return self._analyze_excel_template(file_path)
            elif file_extension == '.csv':
                return self._analyze_csv_template(file_path)
            else:
                return MLTemplateAnalysis(
                    is_ml_template=False,
                    template_structure=None,
                    detected_columns={},
                    validation_errors=[f"Formato de archivo no soportado: {file_extension}"],
                    recommendations=["Sube un archivo Excel (.xlsx) o CSV"],
                    total_products=0,
                    sample_data=[]
                )
                
        except Exception as e:
            return MLTemplateAnalysis(
                is_ml_template=False,
                template_structure=None,
                detected_columns={},
                validation_errors=[f"Error procesando archivo: {str(e)}"],
                recommendations=["Verifica que el archivo no esté corrupto"],
                total_products=0,
                sample_data=[]
            )

    def _analyze_excel_template(self, file_path: str) -> MLTemplateAnalysis:
        """Analyze Excel file for ML template structure"""
        try:
            # Read with openpyxl for better control
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Check for ML template indicators
            obligatory_columns = self._find_obligatory_columns(sheet)
            data_types = self._extract_data_types(sheet)
            
            # Determine if this is an ML template
            is_ml_template = len(obligatory_columns) > 0 and self._has_ml_structure(sheet)
            
            if is_ml_template:
                # Extract template structure
                template_structure = MLTemplateStructure(
                    obligatory_columns=obligatory_columns,
                    data_start_row=8,
                    obligatory_row=4,
                    data_type_row=5
                )
                
                # Analyze product data
                df = pd.read_excel(file_path, skiprows=7, header=0)
                total_products = len(df.dropna(how='all'))
                sample_data = self._extract_sample_data(df, limit=5)
                
                # Validation
                validation_errors = self._validate_template_data(df, obligatory_columns)
                recommendations = self._generate_recommendations(df, obligatory_columns)
                
            else:
                template_structure = None
                total_products = 0
                sample_data = []
                validation_errors = ["El archivo no parece ser una plantilla de Mercado Libre"]
                recommendations = [
                    "Verifica que uses la plantilla oficial de ML",
                    "Las columnas obligatorias deben estar marcadas en la fila 4",
                    "Los datos deben comenzar en la fila 8"
                ]
            
            return MLTemplateAnalysis(
                is_ml_template=is_ml_template,
                template_structure=template_structure,
                detected_columns=data_types,
                validation_errors=validation_errors,
                recommendations=recommendations,
                total_products=total_products,
                sample_data=sample_data
            )
            
        except Exception as e:
            return MLTemplateAnalysis(
                is_ml_template=False,
                template_structure=None,
                detected_columns={},
                validation_errors=[f"Error leyendo Excel: {str(e)}"],
                recommendations=["Verifica que el archivo Excel no esté protegido o corrupto"],
                total_products=0,
                sample_data=[]
            )

    def _analyze_csv_template(self, file_path: str) -> MLTemplateAnalysis:
        """Analyze CSV file for ML template structure"""
        try:
            # Read CSV
            df = pd.read_csv(file_path)
            
            # For CSV, we need to detect ML patterns differently
            # as there's no "obligatorio" marking like in Excel
            detected_columns = self._detect_ml_columns_from_headers(df.columns.tolist())
            
            is_ml_template = len(detected_columns) >= 3  # At least 3 ML-like columns
            
            if is_ml_template:
                # Assume CSV follows ML data structure
                obligatory_columns = {col: 'texto' for col in detected_columns.keys()}
                
                template_structure = MLTemplateStructure(
                    obligatory_columns=obligatory_columns,
                    data_start_row=2,  # CSV typically starts at row 2
                    obligatory_row=1,
                    data_type_row=1
                )
                
                total_products = len(df.dropna(how='all'))
                sample_data = self._extract_sample_data(df, limit=5)
                validation_errors = self._validate_csv_data(df)
                recommendations = self._generate_csv_recommendations(df)
                
            else:
                template_structure = None
                total_products = 0
                sample_data = []
                validation_errors = ["El CSV no parece seguir la estructura de ML"]
                recommendations = [
                    "Verifica que las columnas sigan los nombres estándar de ML",
                    "Considera usar la plantilla Excel oficial de ML"
                ]
            
            return MLTemplateAnalysis(
                is_ml_template=is_ml_template,
                template_structure=template_structure,
                detected_columns=detected_columns,
                validation_errors=validation_errors,
                recommendations=recommendations,
                total_products=total_products,
                sample_data=sample_data
            )
            
        except Exception as e:
            return MLTemplateAnalysis(
                is_ml_template=False,
                template_structure=None,
                detected_columns={},
                validation_errors=[f"Error leyendo CSV: {str(e)}"],
                recommendations=["Verifica que el CSV esté en formato UTF-8"],
                total_products=0,
                sample_data=[]
            )

    def _find_obligatory_columns(self, sheet) -> Dict[str, str]:
        """Find columns marked as obligatory in row 4"""
        obligatory_columns = {}
        
        # Check row 4 for "obligatorio" markers
        for col_idx, cell in enumerate(sheet[4], 1):
            if cell.value and 'obligatorio' in str(cell.value).lower():
                # Get column header from row 1
                header_cell = sheet.cell(row=1, column=col_idx)
                if header_cell.value:
                    # Get data type from row 5
                    data_type_cell = sheet.cell(row=5, column=col_idx)
                    data_type = self._classify_data_type(data_type_cell.value) if data_type_cell.value else 'texto'
                    
                    obligatory_columns[str(header_cell.value).strip()] = data_type
        
        return obligatory_columns

    def _extract_data_types(self, sheet) -> Dict[str, str]:
        """Extract data type information from row 5"""
        data_types = {}
        
        # Get all column headers and their data types
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            header_cell = sheet.cell(row=1, column=col_idx)
            data_type_cell = sheet.cell(row=5, column=col_idx)
            
            if header_cell.value:
                data_type = self._classify_data_type(data_type_cell.value) if data_type_cell.value else 'texto'
                data_types[str(header_cell.value).strip()] = data_type
        
        return data_types

    def _classify_data_type(self, value: str) -> str:
        """Classify data type based on row 5 content"""
        if not value:
            return 'texto'
        
        value_lower = str(value).lower()
        
        for data_type, patterns in self.DATA_TYPE_PATTERNS.items():
            if any(pattern in value_lower for pattern in patterns):
                return data_type
        
        return 'texto'

    def _has_ml_structure(self, sheet) -> bool:
        """Check if sheet has typical ML template structure"""
        # Check if there are headers in row 1
        row1_has_headers = any(cell.value for cell in sheet[1])
        
        # Check if there's data starting around row 8
        row8_has_data = any(cell.value for cell in sheet[8] if cell.value)
        
        # Check for some ML-specific column names
        headers = [str(cell.value).lower() if cell.value else '' for cell in sheet[1]]
        has_ml_columns = any(
            any(pattern in header for pattern in patterns)
            for patterns in self.ML_COLUMN_PATTERNS.values()
            for header in headers
        )
        
        return row1_has_headers and row8_has_data and has_ml_columns

    def _detect_ml_columns_from_headers(self, headers: List[str]) -> Dict[str, str]:
        """Detect ML column types from header names"""
        detected = {}
        
        for header in headers:
            header_lower = header.lower().strip()
            for ml_column, patterns in self.ML_COLUMN_PATTERNS.items():
                if any(pattern in header_lower for pattern in patterns):
                    detected[header] = ml_column
                    break
        
        return detected

    def _extract_sample_data(self, df: pd.DataFrame, limit: int = 5) -> List[Dict[str, Any]]:
        """Extract sample data for preview"""
        sample_df = df.head(limit).dropna(how='all')
        return sample_df.to_dict('records')

    def _validate_template_data(self, df: pd.DataFrame, obligatory_columns: Dict[str, str]) -> List[str]:
        """Validate data in obligatory columns"""
        errors = []
        
        for col_name, data_type in obligatory_columns.items():
            if col_name in df.columns:
                series = df[col_name].dropna()
                
                # Check for empty obligatory columns
                if len(series) == 0:
                    errors.append(f"Columna obligatoria '{col_name}' está vacía")
                    continue
                
                # Validate data types
                if data_type == 'numero':
                    non_numeric = series[~pd.to_numeric(series, errors='coerce').notna()]
                    if len(non_numeric) > 0:
                        errors.append(f"Columna '{col_name}' contiene valores no numéricos")
                
            else:
                errors.append(f"Columna obligatoria '{col_name}' no encontrada")
        
        return errors

    def _validate_csv_data(self, df: pd.DataFrame) -> List[str]:
        """Validate CSV data structure"""
        errors = []
        
        if len(df) == 0:
            errors.append("El archivo CSV está vacío")
        
        if len(df.columns) < 3:
            errors.append("El CSV debe tener al menos 3 columnas")
        
        return errors

    def _generate_recommendations(self, df: pd.DataFrame, obligatory_columns: Dict[str, str]) -> List[str]:
        """Generate recommendations for improving the template"""
        recommendations = []
        
        # Check completion rate
        total_rows = len(df)
        if total_rows == 0:
            recommendations.append("Agrega datos de productos a partir de la fila 8")
        else:
            # Check obligatory column completion
            for col_name in obligatory_columns.keys():
                if col_name in df.columns:
                    completion_rate = df[col_name].notna().sum() / total_rows
                    if completion_rate < 0.8:
                        recommendations.append(f"Completa más datos en la columna obligatoria '{col_name}' ({completion_rate:.1%} completado)")
        
        # Check for common ML best practices
        if 'precio' in [col.lower() for col in df.columns]:
            price_col = next((col for col in df.columns if 'precio' in col.lower()), None)
            if price_col:
                prices = pd.to_numeric(df[price_col], errors='coerce')
                if prices.min() < 0:
                    recommendations.append("Algunos precios son negativos, revisa los datos")
        
        return recommendations

    def _generate_csv_recommendations(self, df: pd.DataFrame) -> List[str]:
        """Generate recommendations for CSV files"""
        recommendations = []
        
        if len(df.columns) < 5:
            recommendations.append("Considera agregar más columnas como descripción, marca, modelo")
        
        # Check for essential columns
        essential_columns = ['titulo', 'precio', 'stock']
        missing_essential = []
        
        for essential in essential_columns:
            if not any(essential in col.lower() for col in df.columns):
                missing_essential.append(essential)
        
        if missing_essential:
            recommendations.append(f"Considera agregar columnas esenciales: {', '.join(missing_essential)}")
        
        return recommendations

    def prepare_for_ai_processing(self, analysis: MLTemplateAnalysis, file_path: str) -> Dict[str, Any]:
        """
        Prepare data structure for future AI processing with DeepSeek API
        """
        return {
            'template_info': {
                'is_ml_template': analysis.is_ml_template,
                'total_products': analysis.total_products,
                'obligatory_columns': analysis.template_structure.obligatory_columns if analysis.template_structure else {},
                'detected_columns': analysis.detected_columns,
            },
            'validation_status': {
                'has_errors': len(analysis.validation_errors) > 0,
                'errors': analysis.validation_errors,
                'recommendations': analysis.recommendations,
            },
            'data_preview': {
                'sample_data': analysis.sample_data,
                'columns': list(analysis.detected_columns.keys()),
            },
            'file_info': {
                'path': file_path,
                'format': Path(file_path).suffix,
                'processed_at': datetime.now().isoformat(),
            },
            'ai_ready': analysis.is_ml_template and len(analysis.validation_errors) == 0,
        }


# Utility functions for the API
def process_ml_template(file_path: str) -> Dict[str, Any]:
    """Main function to process ML template"""
    processor = MLTemplateProcessor()
    analysis = processor.analyze_ml_template(file_path)
    
    # Prepare for future AI integration
    ai_data = processor.prepare_for_ai_processing(analysis, file_path)
    
    # Generate mapping patterns if it's a valid ML template
    mapping_data = None
    if analysis.is_ml_template and MAPPING_GENERATOR_AVAILABLE:
        try:
            mapping_data = generate_mapping_from_analysis(analysis.__dict__)
        except Exception as e:
            print(f"Warning: Could not generate mapping patterns: {e}")
    elif analysis.is_ml_template:
        print("Warning: Mapping generator not available, skipping pattern generation")
    
    return {
        'analysis': {
            'is_ml_template': analysis.is_ml_template,
            'template_structure': analysis.template_structure.__dict__ if analysis.template_structure else None,
            'detected_columns': analysis.detected_columns,
            'validation_errors': analysis.validation_errors,
            'recommendations': analysis.recommendations,
            'total_products': analysis.total_products,
            'sample_data': analysis.sample_data,
        },
        'mapping_patterns': mapping_data,
        'ai_processing_data': ai_data,
        'status': 'success' if analysis.is_ml_template else 'warning',
        'message': 'Plantilla ML detectada y analizada' if analysis.is_ml_template else 'No se detectó estructura de plantilla ML'
    }
