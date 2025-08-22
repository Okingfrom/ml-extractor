from openpyxl import load_workbook
import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))
from template_filler import detect_columns, _normalize

p = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'samples', 'plantilla_ml_ejemplo.xlsx'))
print('Template path:', p)
wb = load_workbook(p)
ws = wb.active
print('Sheet:', wb.sheetnames[0])
print('Max row, Max col:', ws.max_row, ws.max_column)

print('\nFirst 12 rows (raw values):')
for r in range(1, min(13, ws.max_row+1)):
    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
    print(r, vals)

print('\nNormalized header rows (1..12):')
for r in range(1, min(13, ws.max_row+1)):
    texts = [(_normalize(ws.cell(row=r, column=c).value)) for c in range(1, ws.max_column+1)]
    print(r, texts)

print('\nDetect columns with header_row=8:')
cols = detect_columns(ws, header_row=8)
print(cols)

print('\nDetect header row auto:')
from template_filler import detect_header_row
hdr_row, hdr_map = detect_header_row(ws, search_start=1, search_end=min(12, ws.max_row))
print('detected header row:', hdr_row)
print('hdr_map:', hdr_map)
